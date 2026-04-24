"""
core/framework.py
═════════════════
Κοινή υποδομή για όλους τους ελέγχους MySchool.

Παρέχει:
  - read_csv_fixed / read_input  : φόρτωση αρχείων με διόρθωση offset MySchool
  - build_sheet / save_workbook  : δημιουργία Excel με ενιαία μορφοποίηση
  - send_email                   : αποστολή με SSL/STARTTLS fallback
  - run_check                    : ενιαίο run loop (ερωτήσεις → επεξεργασία → email)
"""

import csv, os, smtplib, ssl
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from email.header import Header
from email.utils import formataddr

ENCODING = 'iso-8859-7'
SEP      = ';'
_active_test_body     = None   # ορίζεται από run_check πριν κάθε εκτέλεση
_multi_run_mode       = False  # True κατά τη διάρκεια multi-run
_multi_run_results    = []     # Συλλέγει (title, body_text, result_type) σε multi-run mode
_current_check_title  = ''     # Τίτλος τρέχοντος ελέγχου — εμφανίζεται στα dialogs

def _get_tk_root():
    """Επιστρέφει το υπάρχον Tk root window."""
    import tkinter as tk
    # Ψάχνει existing Tk instance
    try:
        for widget in tk._default_root.winfo_children() if tk._default_root else []:
            pass
        return tk._default_root
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════
# ΦΟΡΤΩΣΗ ΑΡΧΕΙΩΝ
# ═══════════════════════════════════════════════════════════════════

def read_csv_fixed(path):
    """CSV με 1 επιπλέον πεδίο ανά data row (MySchool offset)."""
    rows = []
    with open(path, encoding=ENCODING) as f:
        reader = csv.reader(f, delimiter=SEP)
        headers = next(reader)
        n = len(headers)
        for row in reader:
            rows.append(row[:n])
    return pd.DataFrame(rows, columns=headers)

def read_input(path):
    """Φορτώνει CSV ή Excel αυτόματα."""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        return read_csv_fixed(path)
    elif ext in ('.xlsx', '.xls'):
        return pd.read_excel(path)
    else:
        raise ValueError(f'Μη υποστηριζόμενος τύπος αρχείου: {ext}')

def clean_field(series):
    """Καθαρισμός πεδίων τύπου ="..." από MySchool."""
    return (series.astype(str)
            .str.replace('="', '', regex=False)
            .str.replace('"', '', regex=False)
            .str.strip())


# ═══════════════════════════════════════════════════════════════════
# ΔΗΜΙΟΥΡΓΙΑ EXCEL
# ═══════════════════════════════════════════════════════════════════

def build_sheet(ws, df_sheet, title, columns, center_cols, today,
                subtitle_extra='', highlight_col=None,
                highlight_colors=None, status_colors=None, status_col=None):
    """
    Δημιουργεί φύλλο Excel με ενιαία μορφοποίηση.

    Παράμετροι:
      columns        : list of (name, width) ή (name, width, alias)
      center_cols    : set με ονόματα στηλών που κεντράρονται
      highlight_col  : στήλη με κεχριμπαρί τόνισμα (π.χ. Υπόλοιπο)
      highlight_colors: (header_color, even_color, odd_color)
      status_colors  : dict {status_value: (even_color, odd_color)} για χρωματισμό γραμμών
      status_col     : στήλη για status_colors lookup
    """
    thin  = Side(style='thin', color='CCCCCC')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Normalize columns σε (name, width)
    col_defs = [(c[0], c[1]) for c in columns]
    ncols    = len(col_defs)

    # Γραμμή τίτλου
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ws['A1'] = f'{title}  —  {today.strftime("%d/%m/%Y")}{subtitle_extra}'
    ws['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', start_color='1F4E79')
    ws['A1'].alignment = ctr
    ws.row_dimensions[1].height = 24

    # Γραμμή συνόλου
    ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
    ws['A2'] = f'Σύνολο εγγραφών: {len(df_sheet)}{subtitle_extra and "" or ""}'
    ws['A2'].font      = Font(name='Arial', italic=True, size=9)
    ws['A2'].fill      = PatternFill('solid', start_color='D6E4F0')
    ws['A2'].alignment = ctr
    ws.row_dimensions[2].height = 16

    # Headers
    hl_col_idx = None
    hl_colors  = highlight_colors or ('F4B942', 'FFF3CD', 'FFF8E1')
    for ci, (name, width) in enumerate(col_defs, 1):
        c = ws.cell(row=3, column=ci, value=name)
        if highlight_col and name == highlight_col:
            c.fill    = PatternFill('solid', start_color=hl_colors[0])
            c.font    = Font(name='Arial', bold=True, color='000000', size=10)
            hl_col_idx = ci
        else:
            c.fill = PatternFill('solid', start_color='1F4E79')
            c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.alignment = ctr
        c.border    = brd
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 28

    # Δεδομένα
    alt_fill = PatternFill('solid', start_color='EBF3FB')
    col_keys = [c[0] for c in columns]
    for ri, (_, row) in enumerate(df_sheet.iterrows(), start=4):
        # Καθορισμός χρώματος γραμμής
        if status_colors and status_col and status_col in row:
            colors = status_colors.get(row[status_col], ('EBF3FB', 'FFFFFF'))
            fill   = PatternFill('solid', start_color=colors[ri % 2])
        else:
            fill = alt_fill if ri % 2 == 0 else PatternFill()

        for ci, key in enumerate(col_keys, 1):
            val = row.get(key, '')
            c   = ws.cell(row=ri, column=ci, value=val)
            c.border = brd
            # Τόνισμα στήλης highlight
            if hl_col_idx and ci == hl_col_idx:
                c.fill = PatternFill('solid', start_color=hl_colors[1] if ri % 2 == 0 else hl_colors[2])
                c.font = Font(name='Arial', size=9, bold=True)
            else:
                c.fill = fill
                c.font = Font(name='Arial', size=9)
            c.alignment = ctr if key in center_cols else lft
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(ncols)}3'

def save_workbook(df_sheet, title, columns, center_cols, today, output_path,
                  subtitle_extra='', highlight_col=None, highlight_colors=None,
                  status_colors=None, status_col=None):
    """Δημιουργεί και αποθηκεύει workbook με ένα sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Αποτελέσματα'
    build_sheet(ws, df_sheet, title, columns, center_cols, today,
                subtitle_extra=subtitle_extra,
                highlight_col=highlight_col,
                highlight_colors=highlight_colors,
                status_colors=status_colors,
                status_col=status_col)
    wb.save(output_path)


# ═══════════════════════════════════════════════════════════════════
# ΑΠΟΣΤΟΛΗ EMAIL
# ═══════════════════════════════════════════════════════════════════

def send_email(config, to_addr, subject, body, attachment_path):
    """Αποστολή email με SSL/STARTTLS fallback. to_addr: str ή list."""
    if isinstance(to_addr, list):
        recipients = to_addr
        to_header  = recipients[0]
    else:
        recipients = [to_addr]
        to_header  = to_addr
    msg = MIMEMultipart()
    from_name = getattr(config, 'FROM_NAME', config.FROM_EMAIL)
    msg['From']    = formataddr((from_name, config.FROM_EMAIL))
    msg['To']      = to_header
    msg['Subject'] = Header(subject, 'utf-8')
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    filename = os.path.basename(attachment_path)
    with open(attachment_path, 'rb') as f:
        part = MIMEBase('application', 'vnd.ms-excel')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename=filename)
    part.add_header('Content-Type', 'application/vnd.ms-excel', name=filename)
    msg.attach(part)

    recipients = list(dict.fromkeys(recipients))  # dedup
    msg_str = msg.as_string()

    def _lenient_ctx():
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode    = ssl.CERT_NONE
        return ctx

    sent = False
    # Πρώτη προσπάθεια: πλήρης επαλήθευση πιστοποιητικού
    try:
        with smtplib.SMTP_SSL(config.SMTP_HOST, 465,
                               context=ssl.create_default_context()) as s:
            s.login(config.FROM_EMAIL, config.FROM_PASSWORD)
            s.sendmail(config.FROM_EMAIL, recipients, msg_str)
            sent = True
    except Exception:
        pass

    # Δεύτερη προσπάθεια: SSL χωρίς επαλήθευση (π.χ. εσωτερικός CA)
    if not sent:
        try:
            with smtplib.SMTP_SSL(config.SMTP_HOST, 465,
                                   context=_lenient_ctx()) as s:
                s.login(config.FROM_EMAIL, config.FROM_PASSWORD)
                s.sendmail(config.FROM_EMAIL, recipients, msg_str)
                sent = True
        except Exception:
            pass

    # Τρίτη προσπάθεια: STARTTLS port 587
    if not sent:
        with smtplib.SMTP(config.SMTP_HOST, 587) as s:
            s.starttls(context=_lenient_ctx())
            s.login(config.FROM_EMAIL, config.FROM_PASSWORD)
            s.sendmail(config.FROM_EMAIL, recipients, msg_str)


# ═══════════════════════════════════════════════════════════════════
# ΒΟΗΘΗΤΙΚΕΣ ΕΙΣΑΓΩΓΗ
# ═══════════════════════════════════════════════════════════════════

def ask_file(prompt, required=True, csv_only=False):
    """Ανοίγει παράθυρο επιλογής αρχείου (Windows file dialog)."""
    import tkinter as tk
    from tkinter import filedialog

    if csv_only:
        filetypes = [("CSV", "*.csv"), ("All files", "*.*")]
    else:
        filetypes = [
            ("CSV & Excel", "*.csv *.xlsx *.xls"),
            ("CSV", "*.csv"),
            ("Excel", "*.xlsx *.xls"),
            ("All files", "*.*"),
        ]

    while True:
        print(f'\n{prompt}')

        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        path = filedialog.askopenfilename(title=prompt, filetypes=filetypes, parent=root)
        root.destroy()

        if not path:
            if not required:
                print('  - Paralipthike')
                return None
            print('  X Den epilexthike arxeio - prospathise xana.')
            continue

        if not os.path.exists(path):
            print(f'  X Den vrethike: {path}')
            continue

        if csv_only and not path.lower().endswith('.csv'):
            print('  X Prepei na einai .csv - epilexa xana.')
            continue

        print(f'  OK {os.path.basename(path)}')
        return path

def _show_results_popup(title, body_text, result_type='warn'):
    """
    Εμφανίζει popup παράθυρο με τα αποτελέσματα ελέγχου.
    result_type: 'ok'   → πράσινο (δεν βρέθηκαν θέματα)
                 'warn' → πορτοκαλί (βρέθηκαν εγγραφές)
    Σε multi-run mode τα αποτελέσματα συλλέγονται αντί να εμφανίζονται.
    """
    import core.framework as _fw
    if _fw._multi_run_mode:
        _fw._multi_run_results.append((title, body_text, result_type))
        return

    import tkinter as tk
    from tkinter import scrolledtext

    colors = {
        'ok':   ('#E8F5E9', '#2E7D32', '#388E3C'),   # bg, header_bg, accent
        'warn': ('#FFF8E1', '#E65100', '#F57C00'),
    }
    bg, hdr_bg, accent = colors.get(result_type, colors['warn'])

    win = tk.Toplevel()
    win.title(f'Αποτελέσματα — {title}')
    win.configure(bg=bg)
    win.resizable(True, True)
    win.grab_set()
    win.attributes('-topmost', True)
    _ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '8ball.ico')
    if os.path.exists(_ico):
        try: win.iconbitmap(_ico)
        except Exception: pass
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    win.geometry(f'520x400+{sw//2-260}+{sh//2-200}')

    # Header
    hdr = tk.Frame(win, bg=hdr_bg, pady=8)
    hdr.pack(fill='x')
    icon = '✓' if result_type == 'ok' else '⚠'
    tk.Label(hdr, text=f'{icon}  {title}',
             bg=hdr_bg, fg='white',
             font=('Arial', 11, 'bold')).pack()

    # Κουμπί κλεισίματος — pack πρώτα ώστε να είναι πάντα ορατό
    tk.Button(win, text='Κλείσιμο',
              bg=hdr_bg, fg='white',
              font=('Arial', 10, 'bold'),
              relief='flat', padx=20, pady=6,
              cursor='hand2',
              command=win.destroy).pack(side='bottom', pady=(4, 12))

    # Body text
    txt = scrolledtext.ScrolledText(
        win, wrap=tk.WORD,
        font=('Consolas', 9),
        bg=bg, fg='#212121',
        relief='flat', bd=0,
        padx=14, pady=10
    )
    txt.pack(fill='both', expand=True, padx=10, pady=(10, 4))
    txt.insert('1.0', body_text)
    txt.configure(state='disabled')

    win.wait_window()


def get_ady_xoris_egkrisi(prompt='Αρχείο Αδυνατούντων (υπό έγκριση) [csv / xlsx]:'):
    """
    Επιστρέφει το αρχείο Αδυνατούντων υπό έγκριση.
    Πρώτα ελέγχει το ADY_XORIS_EGKRISI_PATH στο config.
    Αν δεν υπάρχει → επιστρέφει None (θεωρείται άδειο, δεν ζητείται αρχείο).
    """
    try:
        import config as _cfg
        path = getattr(_cfg, 'ADY_XORIS_EGKRISI_PATH', '').strip()
        if path and os.path.exists(path):
            print(f'  ✓ Αδυνατούντες (υπό έγκριση): {os.path.basename(path)}')
            return path
    except Exception as e:
        print(f'  ⚠ Σφάλμα ανάγνωσης config: {e}')
    print('  ℹ Αδυνατούντες (υπό έγκριση): δεν έχει οριστεί — θα αγνοηθούν στους υπολογισμούς.')
    return None


def get_downloaded_file(report_id, prompt=None, required=True, csv_only=False):
    """
    Βρίσκει αυτόματα το κατεβασμένο αρχείο για το report_id (π.χ. '4.8', '8.2')
    από τον φάκελο downloads της σημερινής μέρας.
    Αν δεν βρεθεί, εμφανίζει ενημερωτικό μήνυμα και επιστρέφει None.
    """
    from datetime import datetime as _dt
    import tkinter as tk
    from tkinter import ttk

    import sys as _sys
    if getattr(_sys, 'frozen', False):
        _exe_dir = os.path.dirname(_sys.executable)
        _pf   = os.environ.get('PROGRAMFILES',      r'C:\Program Files').lower()
        _pf86 = os.environ.get('PROGRAMFILES(X86)', r'C:\Program Files (x86)').lower()
        if _exe_dir.lower().startswith(_pf) or _exe_dir.lower().startswith(_pf86):
            base_dir = os.path.join(os.environ.get('LOCALAPPDATA', os.path.expanduser('~')), 'MySchoolChecks')
            os.makedirs(base_dir, exist_ok=True)
        else:
            base_dir = _exe_dir
    else:
        base_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), '..'))
    today = _dt.now().strftime('%Y%m%d')
    today_dir = os.path.normpath(os.path.join(base_dir, 'downloads', today))

    try:
        from core.downloader import find_latest_downloads
        dl = find_latest_downloads(base_dir)
        if report_id in dl and os.path.exists(dl[report_id]):
            path = dl[report_id]
            print(f'  ✓ Αυτόματα [{report_id}]: {os.path.basename(path)}')
            return path
    except Exception as e:
        print(f'  ⚠ Σφάλμα αυτόματης εύρεσης [{report_id}]: {e}')

    from core.downloader import FILE_PREFIX_MAP
    prefix = FILE_PREFIX_MAP.get(report_id, report_id)
    filename = f'{prefix}*'

    print(f'  ⚠ Το αρχείο "{filename}" δεν βρέθηκε στον φάκελο "{today_dir}".')

    root = tk.Tk()
    root.withdraw()
    win = tk.Toplevel(root)
    win.title('Αρχείο δεν βρέθηκε')
    win.resizable(False, False)
    win.grab_set()

    tk.Label(
        win,
        text=f'Το αρχείο  "{filename}"  δεν βρέθηκε\nστον φάκελο:\n{today_dir}',
        padx=20, pady=15, justify='center'
    ).pack()

    ttk.Button(win, text='Κλείσιμο', command=win.destroy).pack(pady=(0, 15))

    win.update_idletasks()
    w, h = win.winfo_width(), win.winfo_height()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    win.geometry(f'{w}x{h}+{(sw-w)//2}+{(sh-h)//2}')

    root.wait_window(win)
    root.destroy()
    return None


def ask_date_yyyymmdd(prompt='Ημερομηνία ελέγχου:'):
    """Ζητά ημερομηνία μέσω GUI popup. Default: σήμερα."""
    return _ask_date_gui(prompt, fmt='%Y-%m-%d')


def _ask_date_gui(prompt, fmt='%Y-%m-%d'):
    """Κοινός date picker — επιστρέφει datetime."""
    import tkinter as tk
    from tkinter import ttk

    result = [datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)]
    today  = result[0]

    win = tk.Toplevel()
    win.title('Επιλογή Ημερομηνίας')
    win.configure(bg='#EEF4F0')
    win.resizable(False, False)
    win.grab_set()
    win.attributes('-topmost', True)
    _ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '8ball.ico')
    if os.path.exists(_ico):
        try: win.iconbitmap(_ico)
        except Exception: pass

    # Κεντράρισμα
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    win.geometry(f'380x280+{sw//2-190}+{sh//2-140}')

    # ── Header παραμετροποίησης ──────────────────────────────────────
    import core.framework as _fw
    tk.Label(win, text='ΠΑΡΑΜΕΤΡΟΠΟΙΗΣΗ ΕΛΕΓΧΟΥ',
             bg='#1F4E79', fg='white',
             font=('Arial', 9, 'bold'), pady=5).pack(fill='x')
    if _fw._current_check_title:
        tk.Label(win, text=_fw._current_check_title,
                 bg='#EEF4F0', fg='#1F4E79',
                 font=('Arial', 9, 'bold'), pady=4).pack(fill='x', padx=10)
    tk.Frame(win, bg='#C5D8E8', height=1).pack(fill='x', padx=10)

    tk.Label(win, text=prompt, bg='#EEF4F0', fg='#1F4E79',
             font=('Arial', 10, 'bold'), pady=8, wraplength=340, justify='center').pack()

    frame = tk.Frame(win, bg='#EEF4F0')
    frame.pack(pady=4)

    # Ημέρα
    tk.Label(frame, text='Ημέρα', bg='#EEF4F0', font=('Arial',8)).grid(row=0,column=0,padx=4)
    tk.Label(frame, text='Μήνας', bg='#EEF4F0', font=('Arial',8)).grid(row=0,column=1,padx=4)
    tk.Label(frame, text='Έτος',  bg='#EEF4F0', font=('Arial',8)).grid(row=0,column=2,padx=4)

    day_var   = tk.StringVar(value=str(today.day))
    month_var = tk.StringVar(value=str(today.month))
    year_var  = tk.StringVar(value=str(today.year))

    days   = [str(i) for i in range(1,32)]
    months = [str(i) for i in range(1,13)]
    years  = [str(i) for i in range(today.year-2, today.year+3)]

    ttk.Combobox(frame, textvariable=day_var,   values=days,   width=4,  state='readonly').grid(row=1,column=0,padx=4)
    ttk.Combobox(frame, textvariable=month_var, values=months, width=4,  state='readonly').grid(row=1,column=1,padx=4)
    ttk.Combobox(frame, textvariable=year_var,  values=years,  width=6,  state='readonly').grid(row=1,column=2,padx=4)

    msg_var = tk.StringVar()
    tk.Label(win, textvariable=msg_var, bg='#EEF4F0', fg='#C62828',
             font=('Arial',8)).pack()

    def confirm():
        try:
            d = int(day_var.get())
            m = int(month_var.get())
            y = int(year_var.get())
            result[0] = datetime(y, m, d)
            win.destroy()
        except ValueError as e:
            msg_var.set(f'Μη έγκυρη ημερομηνία: {e}')

    btn_f = tk.Frame(win, bg='#EEF4F0')
    btn_f.pack(pady=10)
    tk.Button(btn_f, text='Σήμερα', font=('Arial',9),
              bg='#E2EDEA', relief='flat', padx=10, pady=4,
              command=lambda: [day_var.set(str(today.day)),
                               month_var.set(str(today.month)),
                               year_var.set(str(today.year))]).pack(side='left', padx=4)
    tk.Button(btn_f, text='OK', font=('Arial',9,'bold'),
              bg='#1F4E79', fg='white', relief='flat', padx=14, pady=4,
              command=confirm).pack(side='left', padx=4)

    win.wait_window()
    return result[0]

def ask_date_ddmmyyyy(prompt='Ημερομηνία cutoff:'):
    """Ζητά ημερομηνία μέσω GUI popup (DD/MM/YYYY)."""
    return _ask_date_gui(prompt, fmt='%d/%m/%Y')

def _ask_send_options_gui(test_only=False):
    """Dialog επιλογής αποστολής — επιστρέφει (test_mode, do_send).
    test_only=True: εμφανίζει μόνο 'Χωρίς αποστολή' και 'Test mode'."""
    import tkinter as tk
    result = [False, False]  # [test_mode, do_send]

    win = tk.Toplevel()
    win.title('Επιλογές Αποστολής')
    win.configure(bg='#EEF4F0')
    win.resizable(False, False)
    win.grab_set()
    win.attributes('-topmost', True)
    _ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '8ball.ico')
    if os.path.exists(_ico):
        try: win.iconbitmap(_ico)
        except Exception: pass
    # ── Header παραμετροποίησης ──────────────────────────────────────
    import core.framework as _fw
    tk.Label(win, text='ΠΑΡΑΜΕΤΡΟΠΟΙΗΣΗ ΕΛΕΓΧΟΥ',
             bg='#1F4E79', fg='white',
             font=('Arial', 9, 'bold'), pady=5).pack(fill='x')
    if _fw._current_check_title:
        tk.Label(win, text=_fw._current_check_title,
                 bg='#EEF4F0', fg='#1F4E79',
                 font=('Arial', 9, 'bold'), pady=4).pack(fill='x', padx=10)
    tk.Frame(win, bg='#C5D8E8', height=1).pack(fill='x', padx=10)

    tk.Label(win, text='Επιλογές Αποστολής Email',
             bg='#EEF4F0', fg='#1F4E79',
             font=('Arial', 10, 'bold'), pady=6).pack(fill='x', padx=10)

    body = tk.Frame(win, bg='#EEF4F0', padx=20, pady=12)
    body.pack(fill='both', expand=True)

    mode_var = tk.StringVar(value='none')

    import config as _cfg
    _test_label = f'Test mode  (μόνο ενημέρωση στο {_cfg.FROM_EMAIL})' if _cfg.FROM_EMAIL else 'Test mode  (μόνο ενημέρωση στο email αποστολής)'

    tk.Radiobutton(body, text='Χωρίς αποστολή email',
                   variable=mode_var, value='none',
                   bg='#EEF4F0', font=('Arial',10),
                   activebackground='#EEF4F0').pack(anchor='w', pady=3)
    tk.Radiobutton(body, text=_test_label,
                   variable=mode_var, value='test',
                   bg='#EEF4F0', font=('Arial',10),
                   activebackground='#EEF4F0',
                   wraplength=300, justify='left').pack(anchor='w', pady=3)
    if not test_only:
        tk.Radiobutton(body, text='Κανονική αποστολή  (σε όλα τα σχολεία)',
                       variable=mode_var, value='full',
                       bg='#EEF4F0', font=('Arial',10),
                       activebackground='#EEF4F0').pack(anchor='w', pady=3)

    def confirm():
        v = mode_var.get()
        result[0] = (v == 'test')
        result[1] = (v in ('test', 'full'))
        win.destroy()

    tk.Button(win, text='OK', font=('Arial',10,'bold'),
              bg='#1F4E79', fg='white', relief='flat',
              padx=20, pady=6, cursor='hand2',
              command=confirm).pack(pady=(0,12))

    # Auto-size: κεντράρισμα μετά τη δημιουργία όλων των widgets
    win.update_idletasks()
    w = win.winfo_reqwidth() + 20
    h = win.winfo_reqheight() + 20
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    win.geometry(f'{w}x{h}+{(sw-w)//2}+{(sh-h)//2}')

    win.wait_window()
    return result[0], result[1]


def yes_no(prompt):
    """Ερώτηση Ναι/Όχι μέσω GUI popup."""
    import tkinter as tk
    result = [False]
    win = tk.Toplevel()
    win.title('')
    win.configure(bg='#EEF4F0')
    win.resizable(False, False)
    win.grab_set()
    win.attributes('-topmost', True)
    _ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '8ball.ico')
    if os.path.exists(_ico):
        try: win.iconbitmap(_ico)
        except Exception: pass
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    win.geometry(f'340x130+{sw//2-170}+{sh//2-65}')

    tk.Label(win, text=prompt, bg='#EEF4F0', fg='#1F4E79',
             font=('Arial', 10), wraplength=300, pady=16).pack()

    btn_f = tk.Frame(win, bg='#EEF4F0')
    btn_f.pack()

    def on_yes():
        result[0] = True
        win.destroy()
    def on_no():
        result[0] = False
        win.destroy()

    tk.Button(btn_f, text='Ναι', font=('Arial',10,'bold'),
              bg='#1F4E79', fg='white', relief='flat',
              padx=20, pady=6, cursor='hand2',
              command=on_yes).pack(side='left', padx=8)
    tk.Button(btn_f, text='Όχι', font=('Arial',10),
              bg='#E2EDEA', fg='#1F4E79', relief='flat',
              padx=20, pady=6, cursor='hand2',
              command=on_no).pack(side='left', padx=8)

    win.wait_window()
    return result[0]


# ═══════════════════════════════════════════════════════════════════
# ΕΝΙΑΙΟ RUN LOOP
# ═══════════════════════════════════════════════════════════════════

def run_check(check_module, config):
    """
    Εκτελεί έναν έλεγχο με βάση το check_module.

    Αναμενόμενες σταθερές στο module:
      CHECK_TITLE    : str
      COLUMNS        : list of (name, width) ή (name, width, alias)
      SCHOOL_COLUMN  : str  — στήλη για split ανά σχολείο
      EMAIL_COLUMN   : str  — στήλη με email σχολείου (ή None)
      EMAIL_SUBJECT  : str  — θέμα email
      EMAIL_BODY     : str  — σώμα email (ή callable(school) -> str)
      RESULTS_FOLDER : str  — prefix φακέλου αποτελεσμάτων
      CENTER_COLS    : set
      [HIGHLIGHT_COL]: str  — προαιρετικά
      [STATUS_COLORS]: dict — προαιρετικά
      [STATUS_COL]   : str  — προαιρετικά

    Αναμενόμενες συναρτήσεις στο module:
      ask_inputs(fw) -> dict   : ζητά τα αρχεία και παραμέτρους (επιστρέφει context)
      process(context) -> df   : επεξεργασία, επιστρέφει DataFrame εξόδου
    """
    title  = getattr(check_module, 'CHECK_TITLE',    '?')
    cols   = getattr(check_module, 'COLUMNS',        [])
    scol   = getattr(check_module, 'SCHOOL_COLUMN',  'Ονομασία Σχολείου')
    ecol   = getattr(check_module, 'EMAIL_COLUMN',   'Email Σχολείου')
    subj   = getattr(check_module, 'EMAIL_SUBJECT',  '')
    body_t = getattr(check_module, 'EMAIL_BODY',     '')

    # Έλεγξε για custom email template αποθηκευμένο από τον χρήστη
    try:
        import json, sys as _sys
        _mod_name = check_module.__name__.split('.')[-1]
        _exe = _sys.executable
        if getattr(_sys, 'frozen', False):
            _exe_dir = os.path.dirname(_exe)
            if 'program files' in _exe_dir.lower():
                _base = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'MySchoolChecks')
            else:
                _base = _exe_dir
        else:
            _base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        _settings_path = os.path.join(_base, 'data', 'local_settings.json')
        if os.path.exists(_settings_path):
            with open(_settings_path, encoding='utf-8') as _f:
                _sdata = json.load(_f)
            _tmpl = _sdata.get('email_templates', {}).get(_mod_name)
            if _tmpl:
                subj   = _tmpl.get('subject', subj)
                _cbody = _tmpl.get('body', '')
                body_t = lambda school='', _b=_cbody: _b + config.email_signature()
    except Exception:
        pass
    rfold  = getattr(check_module, 'RESULTS_FOLDER', 'results')
    ccols  = getattr(check_module, 'CENTER_COLS',    set())
    hlcol  = getattr(check_module, 'HIGHLIGHT_COL',  None)
    hlclrs = getattr(check_module, 'HIGHLIGHT_COLORS', None)
    sclrs  = getattr(check_module, 'STATUS_COLORS',  None)
    scol2  = getattr(check_module, 'STATUS_COL',     None)
    has_email = getattr(check_module, 'HAS_EMAIL',   ecol is not None)

    print('=' * 62)
    print(f'  {title}')
    print('=' * 62)

    # Αποθήκευση test_body function αν υπάρχει στο module
    import builtins
    _tb = getattr(check_module, 'test_body', None)
    # Περνάμε μέσω global για χρήση στο _send_loop
    import core.framework as _fw
    _fw._active_test_body    = _tb
    _fw._current_check_title = title   # για εμφάνιση στα dialogs

    # Ζητά αρχεία/παραμέτρους από το module
    ctx = check_module.ask_inputs()

    today     = ctx.get('today', datetime.today().replace(hour=0, minute=0, second=0, microsecond=0))
    test_mode = False
    do_send   = False

    if has_email:
        _test_only = getattr(check_module, 'TEST_ONLY', False)
        test_mode, do_send = _ask_send_options_gui(test_only=_test_only)

    print(f'\n  Ημερομηνία : {today.strftime("%d/%m/%Y")}')
    if has_email:
        print(f'  Λειτουργία : {"TEST MODE" if test_mode else "ΚΑΝΟΝΙΚΗ"}')
        print(f'  Αποστολή   : {"ΝΑΙ" if do_send else "ΟΧΙ"}')
    print('-' * 62)

    # Επεξεργασία
    print('\nΕπεξεργασία...')
    df_out = check_module.process(ctx)
    print(f'  → {len(df_out)} εγγραφές βρέθηκαν')

    if df_out.empty:
        print(f'\n✓ Δεν βρέθηκαν εγγραφές.')
        _show_results_popup(
            title,
            f'Ημερομηνία ελέγχου: {today.strftime("%d/%m/%Y")}\n\n'
            f'✓  Δεν βρέθηκαν εγγραφές που χρήζουν προσοχής.\n\n'
            f'Ο έλεγχος ολοκληρώθηκε χωρίς θέματα.',
            result_type='ok'
        )
        return

    # Φάκελος αποτελεσμάτων — Documents\MySchoolChecks\results_YYYYMMDD\
    subfolder = rfold.replace('results_', '') if rfold.startswith('results_') else rfold
    import sys as _sys
    if getattr(_sys, 'frozen', False):
        # Πάντα στα Documents του χρήστη — εύκολο να βρεθεί
        _docs = os.path.join(os.path.expanduser('~'), 'Documents')
        _app_base = os.path.join(_docs, 'MySchoolChecks')
    else:
        # Development: φάκελος του project
        _app_base = os.path.normpath(os.path.join(
            os.path.dirname(os.path.abspath(check_module.__file__)), '..'))
    out_dir = os.path.normpath(os.path.join(
        _app_base, f'results_{today.strftime("%Y%m%d")}', subfolder
    ))
    os.makedirs(out_dir, exist_ok=True)
    print(f'\nΑποθήκευση → {out_dir}')
    print('-' * 62)

    # Συνολικό αρχείο
    path_all = os.path.join(out_dir, f'{today.strftime("%Y%m%d")}_ΣΥΝΟΛΟ.xlsx')
    save_workbook(df_out, title, cols, ccols, today, path_all,
                  highlight_col=hlcol, highlight_colors=hlclrs,
                  status_colors=sclrs, status_col=scol2)
    print(f'  ✓ Συνολικό: {os.path.basename(path_all)}  ({len(df_out)} εγγραφές)')

    # Αρχεία ανά σχολείο — δημιουργούνται μόνο σε κανονική αποστολή (όχι test mode)
    schools      = df_out[scol].unique()
    school_files = {}
    if not do_send or test_mode:
        print(f'\n  Χωρίς αποστολή: τα {len(schools)} αρχεία ανά σχολείο παραλείπονται.')
    else:
        print(f'\n  Δημιουργία {len(schools)} αρχείων ανά σχολείο...')
        for school in sorted(schools):
            df_s      = df_out[df_out[scol] == school].copy()
            safe_name = ''.join(c for c in school if c not in r'\/:*?"<>|').strip()[:60]
            path_s    = os.path.join(out_dir, f'{today.strftime("%Y%m%d")}_{safe_name}.xlsx')
            save_workbook(df_s, title, cols, ccols, today, path_s,
                          subtitle_extra=f'  |  {school}',
                          highlight_col=hlcol, highlight_colors=hlclrs,
                          status_colors=sclrs, status_col=scol2)
            school_files[school] = path_s
            print(f'  ✓ {safe_name}  ({len(df_s)} εγγραφές)')

    # Email
    if do_send and has_email:
        _send_loop(config, test_mode, title, today, subj, body_t,
                   df_out, schools, school_files, scol, ecol, path_all)

    print('─' * 62)
    total_files = 1 + len(school_files)  # συνολικό + ανά σχολείο
    print(f'\n✓ Ολοκληρώθηκε!  {total_files} αρχεία στο φάκελο:')
    print(f'  {out_dir}')

    # Popup αποτελεσμάτων σε όλες τις περιπτώσεις
    _tb = getattr(check_module, 'test_body', None)
    if _tb:
        try:
            summary = _tb(df_out, today, schools)
        except Exception:
            summary = None
    else:
        summary = None

    if not summary:
        summary = (
            f'Σύνοψη ελέγχου — {today.strftime("%d/%m/%Y")}\n'
            f'{"─" * 40}\n'
            f'Βρέθηκαν: {len(df_out)} εγγραφές\n'
            f'Σχολεία:  {len(schools)}\n\n'
            + '\n'.join(f'  • {s}' for s in sorted(str(x) for x in schools))
        )

    summary += (
        f'\n\n{"─" * 40}\n'
        f'Αποτελέσματα αποθηκεύτηκαν στο φάκελο:\n'
        f'{out_dir}\n\n'
        f'📋 Για περισσότερες πληροφορίες δες τα αρχεία\n'
        f'στο φάκελο αποτελεσμάτων.'
    )
    _show_results_popup(title, summary, result_type='warn')

    # Άνοιγμα φακέλου αποτελεσμάτων στον Explorer
    try:
        import subprocess
        subprocess.Popen(['explorer', out_dir])
    except Exception:
        pass



def _send_loop(config, test_mode, title, today, subject_base, body_template,
               df_out, schools, school_files, scol, ecol, path_all):
    """Εσωτερική συνάρτηση αποστολής email."""
    import sys as _sys
    print(f'\n{"─"*62}')

    subject = f'{subject_base} — {today.strftime("%d/%m/%Y")}'

    if test_mode:
        # Αν το module έχει test_body(), τη χρησιμοποιούμε για σύνοψη
        # αλλιώς fallback στο γενικό μήνυμα
        _test_body_fn = globals().get('_active_test_body', None)
        if _test_body_fn:
            body = _test_body_fn(df_out, today, schools)
        else:
            body = (f'Σύνοψη ελέγχου — {today.strftime("%d/%m/%Y")}\n'
                    f'Βρέθηκαν: {len(df_out)} εγγραφές\n'
                    f'Σχολεία: {len(schools)}\n'
                    f'{", ".join(sorted(str(s) for s in schools))}')
        print(f'🧪 TEST MODE → {config.TEST_EMAIL}')
        print(f'   Θέμα: [TEST] {subject}')
        print(f'\n   Προεπισκόπηση:\n{"─"*40}')
        print(body[:400] + ('...' if len(body) > 400 else ''))
        print('─' * 40)
        recipients = [config.TEST_EMAIL]
        cc_extra = getattr(config, 'TEST_EMAIL_CC', None)
        if cc_extra:
            recipients.append(cc_extra)
        try:
            send_email(config, recipients, f'[TEST] {subject}', body, path_all)
            print(f'  ✓ Εστάλη στο {config.TEST_EMAIL}' +
                  (f' + {cc_extra}' if cc_extra else ''))
        except Exception as e:
            print(f'  ✗ Σφάλμα: {e}')
    else:
        print(f'🚀 ΚΑΝΟΝΙΚΗ ΑΠΟΣΤΟΛΗ — {len(schools)} σχολεία')
        print(f'   Θέμα: {subject}\n')
        def _valid_email(e):
            return e and e not in ('', 'nan', 'None') and '@' in e

        no_email = []
        bad_format = []
        for school in sorted(schools):
            df_s    = df_out[df_out[scol] == school]
            email_s = str(df_s[ecol].iloc[0]).strip() if ecol in df_s.columns else ''
            if not email_s or email_s in ('', 'nan', 'None'):
                print(f'   ⚠  {str(school)[:50]} — ΔΕΝ ΥΠΑΡΧΕΙ EMAIL')
                no_email.append(school)
            elif '@' not in email_s:
                print(f'   ⚠  {str(school)[:50]} — ΛΑΘΟΣ FORMAT: {email_s}')
                bad_format.append(school)
            else:
                print(f'   →  {str(school)[:50]} → {email_s}')
        if no_email:
            print(f'\n   ⚠  {len(no_email)} σχολεία χωρίς email — θα παραλειφθούν.')
        if bad_format:
            print(f'   ⚠  {len(bad_format)} σχολεία με λάθος format email — θα παραλειφθούν.')

        ok = fail = 0
        for school, path_s in school_files.items():
            df_s    = df_out[df_out[scol] == school]
            email_s = str(df_s[ecol].iloc[0]).strip() if ecol in df_s.columns else ''
            if not _valid_email(email_s):
                fail += 1
                continue
            body = (body_template if isinstance(body_template, str)
                    else body_template(str(school)))
            try:
                send_email(config, [email_s], subject, body, path_s)
                print(f'  ✓ {str(school)[:50]} → {email_s}')
                ok += 1
            except Exception as e:
                print(f'  ✗ {str(school)[:50]} → {e}')
                fail += 1
        print(f'\n  Αποστολές: {ok} επιτυχείς, {fail} αποτυχίες')

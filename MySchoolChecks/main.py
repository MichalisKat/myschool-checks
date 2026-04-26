"""
main.py
═══════
Κεντρικό σημείο εισόδου — MySchool Αυτοματισμοί.
"""

import sys, os, importlib, threading, queue, time, warnings
warnings.filterwarnings('ignore', message='.*file size.*not 512.*sector size.*')

# DEBUG: Καταγραφή crashes σε αρχείο
def _emergency_log(exc_type, exc_val, exc_tb):
    import traceback
    try:
        _log = os.path.join(os.path.expanduser('~'), 'Desktop', 'crash.log')
        with open(_log, 'w', encoding='utf-8') as _f:
            _f.write(''.join(traceback.format_exception(exc_type, exc_val, exc_tb)))
    except Exception:
        pass
    sys.__excepthook__(exc_type, exc_val, exc_tb)
sys.excepthook = _emergency_log


def _app_base():
    """Επιστρέφει τον βασικό φάκελο για αποθήκευση δεδομένων.
    - Program Files → %LOCALAPPDATA%\MySchoolChecks  (δεν επιτρέπεται εγγραφή στο PF)
    - dist\ (development/portable) → δίπλα στο .exe
    - Development → φάκελος του κώδικα
    """
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        pf  = os.environ.get('PROGRAMFILES',       r'C:\Program Files').lower()
        pf86= os.environ.get('PROGRAMFILES(X86)',  r'C:\Program Files (x86)').lower()
        if exe_dir.lower().startswith(pf) or exe_dir.lower().startswith(pf86):
            data_dir = os.path.join(
                os.environ.get('LOCALAPPDATA', os.path.expanduser('~')),
                'MySchoolChecks')
            os.makedirs(data_dir, exist_ok=True)
            return data_dir
        return exe_dir
    return os.path.dirname(os.path.abspath(__file__))


def _docs_base():
    """Επιστρέφει τον φάκελο Documents\MySchoolChecks — κοινός για downloads και results.
    Εύκολος στην πρόσβαση από τον χρήστη.
    """
    _docs = os.path.join(os.path.expanduser('~'), 'Documents')
    path  = os.path.join(_docs, 'MySchoolChecks')
    os.makedirs(path, exist_ok=True)
    return path
import tkinter as tk
from tkinter import messagebox

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import config
except ImportError:
    import tkinter as tk
    from tkinter import messagebox
    _r = tk.Tk(); _r.withdraw()
    messagebox.showerror('Σφάλμα', 'Δεν βρέθηκε το config.py στον φάκελο του προγράμματος.')
    sys.exit(1)


class GUIStream:
    """Ανακατευθύνει το sys.stdout στο status bar του GUI."""
    def __init__(self):
        self._callback = None
        self._buffer   = []

    def set_callback(self, cb):
        self._callback = cb
        # Εκκρεμή μηνύματα
        for msg in self._buffer:
            cb(msg)
        self._buffer.clear()

    def write(self, text):
        text = text.strip()
        if not text:
            return
        if self._callback:
            self._callback(text)
        else:
            self._buffer.append(text)

    def flush(self):
        pass


_gui_stream = GUIStream()
sys.stdout  = _gui_stream
sys.stderr  = _gui_stream

try:
    import ctypes
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('myschool.checks.1')
except Exception:
    pass

WHITE = '#FFFFFF'
C = {
    'bg'        : '#EEF4F0',
    'bg2'       : '#E2EDEA',
    'border'    : '#C4D9D0',
    'hdr_bg'    : '#1F4E79',
    'hdr_fg'    : '#FFFFFF',
    'hdr_sub'   : '#D6E4F0',
    'btn_bg'    : '#1F4E79',
    'btn_fg'    : '#FFFFFF',
    'btn_act'   : '#2E75B6',
    'btn_dis'   : '#7F9FAF',
    'sel_bg'    : '#D0E8DC',
    'sel_bd'    : '#4CA870',
    'norm_bg'   : '#FFFFFF',
    'norm_bd'   : '#C4D9D0',
    'desc'      : '#4A6860',
    'footer'    : '#7A9A90',
    'status_ok' : '#2E7D32',
    'status_err': '#C62828',
    'status_run': '#E65100',
    'ind_idle'  : '#B0BEC5',
    'ind_run'   : '#FB8C00',
    'ind_ok'    : '#43A047',
    'ind_err'   : '#E53935',
    'ind_out'   : '#ECEFF1',
    'warn'      : '#E65100',
}


CHECK_ORDER = [
    'forma_82',
    'orario_diafora',
    'arnhtika_ypoloipa',
    'adies_aneu',
    'adies',
    'analipsi',
    'dioikitiko_ergo',
    'ypoloipa',
]

def load_checks():
    base = os.path.dirname(os.path.abspath(__file__))
    if base not in sys.path:
        sys.path.insert(0, base)

    checks = []

    if getattr(sys, 'frozen', False):
        # ── Frozen exe (PyInstaller) ──────────────────────────────────
        # Τα checks είναι compiled μέσα στο exe. Τα φορτώνουμε με
        # τη σειρά του CHECK_ORDER + όποια άλλα γνωστά.
        # Δεν χρειαζόμαστε os.listdir() — απευθείας import.
        all_known = list(CHECK_ORDER) + [
            'forma_82', 'orario_diafora', 'arnhtika_ypoloipa',
            'adies_aneu', 'adies', 'analipsi', 'dioikitiko_ergo', 'ypoloipa',
        ]
        seen = set()
        ordered = []
        for m in all_known:
            if m not in seen:
                seen.add(m)
                ordered.append(m)
    else:
        # ── Development mode ─────────────────────────────────────────
        checks_dir = os.path.join(base, 'checks')
        available = {fname[:-3] for fname in os.listdir(checks_dir)
                     if fname.endswith('.py') and not fname.startswith('_')}
        ordered = CHECK_ORDER + sorted(available - set(CHECK_ORDER))

    _log = os.path.join(os.path.expanduser('~'), 'Desktop', 'checks_errors.log')
    with open(_log, 'w', encoding='utf-8') as _f:
        _f.write(f'frozen={getattr(sys, "frozen", False)}\n')
        _f.write(f'sys.path={sys.path}\n\n')

    for mod_base in ordered:
        mod_name = f'checks.{mod_base}'
        try:
            mod = importlib.import_module(mod_name)
            title = getattr(mod, 'CHECK_TITLE', None)
            if title:
                desc = getattr(mod, 'CHECK_DESCRIPTION', '')
                checks.append((title, desc, mod))
        except Exception as e:
            import traceback as _tb
            with open(_log, 'a', encoding='utf-8') as _f:
                _f.write(f'\n--- {mod_base} ---\n')
                _f.write(_tb.format_exc())
            print(f'  !! Den fortothike to {mod_base}.py: {e}')
    return checks


def password_is_set():
    """Ελέγχει αν ο κωδικός email έχει οριστεί (keyring ή config)."""
    try:
        import keyring
        val = keyring.get_password('MySchoolChecks', 'FROM_PASSWORD')
        if val:
            return True
    except Exception:
        pass
    return bool(getattr(config, 'FROM_PASSWORD', ''))


def _get_local_settings_path():
    return os.path.join(_app_base(), 'data', 'local_settings.json')


def _load_local_settings():
    """Φορτώνει το data/local_settings.json. Επιστρέφει dict."""
    import json
    path = _get_local_settings_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


_SENSITIVE_KEYS = {'MYSCHOOL_USER', 'MYSCHOOL_PASS', 'FROM_PASSWORD'}
_KEYRING_SERVICE = 'MySchoolChecks'


def _save_config(updates):
    """
    Αποθηκεύει ρυθμίσεις:
      - Ευαίσθητα (MYSCHOOL_USER/PASS, FROM_PASSWORD) → Windows Credential Manager
      - Μη-ευαίσθητα → data/local_settings.json
    Ενημερώνει επίσης το live config object ώστε να ισχύουν άμεσα.
    """
    import json

    sensitive   = {k: v for k, v in updates.items() if k in _SENSITIVE_KEYS}
    nonsensitive = {k: v for k, v in updates.items() if k not in _SENSITIVE_KEYS}

    # ── Αποθήκευση ευαίσθητων στο keyring ───────────────────────────────────
    try:
        import keyring
        for key, val in sensitive.items():
            if val:
                keyring.set_password(_KEYRING_SERVICE, key, val)
    except Exception as e:
        # Αν το keyring αποτύχει, πέσε back στο JSON (δεν σπάμε τη ροή)
        nonsensitive.update(sensitive)
        print(f'[Προσοχή] keyring μη διαθέσιμο, credentials αποθηκεύονται στο JSON: {e}')

    # ── Αποθήκευση μη-ευαίσθητων στο JSON ───────────────────────────────────
    if nonsensitive:
        path = _get_local_settings_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        existing = _load_local_settings()
        # Βεβαιώσου ότι δεν ξαναμπαίνουν sensitive στο JSON
        for sk in _SENSITIVE_KEYS:
            existing.pop(sk, None)
        existing.update(nonsensitive)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)

    # ── Ενημέρωση live config object ────────────────────────────────────────
    for key, val in updates.items():
        setattr(config, key, val)


class Indicator(tk.Canvas):
    SIZE = 14

    def __init__(self, parent, **kw):
        bg = kw.pop('bg', C['norm_bg'])
        super().__init__(parent, width=self.SIZE, height=self.SIZE,
                         bg=bg, highlightthickness=0, **kw)
        self._circle = self.create_oval(1, 1, self.SIZE-1, self.SIZE-1,
                                        fill=C['ind_idle'],
                                        outline=C['ind_out'], width=1)

    def set_state(self, state):
        colors = {
            'idle'   : (C['ind_idle'], C['ind_out']),
            'running': (C['ind_run'],  '#E65100'),
            'ok'     : (C['ind_ok'],   '#2E7D32'),
            'error'  : (C['ind_err'],  '#B71C1C'),
        }
        fill, outline = colors.get(state, colors['idle'])
        self.itemconfig(self._circle, fill=fill, outline=outline)


class SettingsDialog(tk.Toplevel):
    """Παράθυρο ρυθμίσεων — με tabs: Σύνδεση / Email / Αρχεία."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Ρυθμίσεις')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._ady_new_path = tk.StringVar(value='')
        self._build()

        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{px}+{py}')

    # ── helpers ──────────────────────────────────────────────────────────────
    def _cfg(self, key, default=''):
        return getattr(config, key, default)

    def _labeled_entry(self, parent, row, label, var, width=28, show=''):
        """Βοηθητικό: label + entry σε grid."""
        tk.Label(parent, text=label, bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9), anchor='w').grid(
                 row=row, column=0, sticky='w', pady=(6, 2))
        e = tk.Entry(parent, textvariable=var, width=width,
                     font=('Arial', 10), relief='solid', bd=1, show=show)
        e.grid(row=row, column=1, sticky='w', padx=(10, 0), pady=(6, 2))
        return e

    def _pw_row(self, parent, row, label, var):
        """Label + entry κωδικού με κουμπί 👁."""
        tk.Label(parent, text=label, bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9), anchor='w').grid(
                 row=row, column=0, sticky='w', pady=(6, 2))
        frame = tk.Frame(parent, bg=C['bg'])
        frame.grid(row=row, column=1, sticky='w', padx=(10, 0), pady=(6, 2))
        show_var = tk.BooleanVar(value=False)
        entry = tk.Entry(frame, textvariable=var, show='•',
                         width=24, font=('Arial', 10), relief='solid', bd=1)
        entry.pack(side='left')
        tk.Button(frame, text='👁', bg=C['bg'], relief='flat',
                  font=('Arial', 11), cursor='hand2',
                  command=lambda: [show_var.set(not show_var.get()),
                                   entry.configure(show='' if show_var.get() else '•')]
                  ).pack(side='left', padx=(4, 0))
        return entry

    def _section_label(self, parent, row, text):
        tk.Label(parent, text=text, bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').grid(
                 row=row, column=0, columnspan=2, sticky='w', pady=(14, 4))

    # ── build ─────────────────────────────────────────────────────────────────
    def _build(self):
        from tkinter import ttk

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=C['hdr_bg'], pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='⚙  Ρυθμίσεις', bg=C['hdr_bg'], fg=C['hdr_fg'],
                 font=('Arial', 12, 'bold')).pack()

        # ── ttk.Notebook ─────────────────────────────────────────────────────
        # Σημ: στα Windows το theme override κάνει το foreground αόρατο σε σκούρο bg.
        # Λύση: επιλεγμένο tab = ανοιχτό μπλε (hdr_sub) με σκούρο κείμενο (hdr_bg).
        style = ttk.Style()
        style.configure('TNotebook',     background=C['bg'])
        style.configure('TNotebook.Tab', background=C['bg2'], foreground=C['desc'],
                        font=('Arial', 9, 'bold'), padding=(12, 5))
        style.map('TNotebook.Tab',
                  background=[('selected', C['hdr_sub']), ('active', C['sel_bg'])],
                  foreground=[('selected', C['hdr_bg']),  ('active', C['hdr_bg'])])

        nb = ttk.Notebook(self)
        nb.pack(fill='both', padx=16, pady=12)

        tab1 = tk.Frame(nb, bg=C['bg'], padx=20, pady=14)
        tab2 = tk.Frame(nb, bg=C['bg'], padx=20, pady=14)
        tab3 = tk.Frame(nb, bg=C['bg'], padx=20, pady=14)
        nb.add(tab1, text='  Σύνδεση  ')
        nb.add(tab2, text='  Email  ')
        nb.add(tab3, text='  Αρχεία  ')

        # ── Tab 1: Σύνδεση (MySchool + email password) ───────────────────────
        self._section_label(tab1, 0, 'MySchool (SSO):')
        self._ms_user_var = tk.StringVar(value=self._cfg('MYSCHOOL_USER'))
        self._labeled_entry(tab1, 1, 'Username:', self._ms_user_var)
        self._ms_pass_var = tk.StringVar(value=self._cfg('MYSCHOOL_PASS'))
        self._pw_row(tab1, 2, 'Κωδικός:', self._ms_pass_var)

        sep = tk.Frame(tab1, bg=C['border'], height=1)
        sep.grid(row=3, column=0, columnspan=2, sticky='ew', pady=(16, 4))

        # ── Επιλογή Browser ───────────────────────────────────────────────────
        self._section_label(tab1, 4, 'Browser για σύνδεση:')
        self._browser_var = tk.StringVar(value=self._cfg('BROWSER', 'chrome'))
        br_frame = tk.Frame(tab1, bg=C['bg'])
        br_frame.grid(row=5, column=0, columnspan=2, sticky='w', pady=(2, 4))
        tk.Radiobutton(br_frame, text='Chrome', variable=self._browser_var,
                       value='chrome', bg=C['bg'], fg=C['hdr_bg'],
                       font=('Arial', 9), activebackground=C['bg'],
                       selectcolor=C['bg2']).pack(side='left', padx=(0, 16))
        tk.Radiobutton(br_frame, text='Firefox', variable=self._browser_var,
                       value='firefox', bg=C['bg'], fg=C['hdr_bg'],
                       font=('Arial', 9), activebackground=C['bg'],
                       selectcolor=C['bg2']).pack(side='left')
        tk.Label(tab1, text='(και οι δύο πρέπει να είναι εγκατεστημένοι για να επιλεγούν)',
                 bg=C['bg'], fg=C['footer'], font=('Arial', 8)).grid(
                 row=6, column=0, columnspan=2, sticky='w')

        sep1b = tk.Frame(tab1, bg=C['border'], height=1)
        sep1b.grid(row=7, column=0, columnspan=2, sticky='ew', pady=(16, 4))

        self._section_label(tab1, 8, 'Λογαριασμός email:')
        self._pw_var = tk.StringVar(value=self._cfg('FROM_PASSWORD'))
        self._pw_row(tab1, 9, 'Κωδικός email:', self._pw_var)

        if not password_is_set():
            warn = tk.Frame(tab1, bg='#FFF3E0',
                            highlightbackground='#FFB74D', highlightthickness=1)
            warn.grid(row=10, column=0, columnspan=2, sticky='ew', pady=(8, 0))
            tk.Label(warn, text='⚠  Ο κωδικός email δεν έχει οριστεί.',
                     bg='#FFF3E0', fg=C['warn'],
                     font=('Arial', 8), padx=8, pady=4).pack()

        # ── Tab 2: Email (ταυτότητα αποστολέα) ───────────────────────────────
        self._section_label(tab2, 0, 'Στοιχεία αποστολέα:')
        self._from_name_var  = tk.StringVar(value=self._cfg('FROM_NAME'))
        self._from_email_var = tk.StringVar(value=self._cfg('FROM_EMAIL'))
        self._smtp_var       = tk.StringVar(value=self._cfg('SMTP_HOST'))

        self._labeled_entry(tab2, 1, 'Εμφανιζόμενο όνομα:', self._from_name_var,  width=30)
        self._labeled_entry(tab2, 2, 'Email αποστολής:',     self._from_email_var, width=30)
        tk.Label(tab2, text='(χρησιμοποιείται και για δοκιμαστική αποστολή)',
                 bg=C['bg'], fg=C['footer'], font=('Arial', 8)).grid(
                 row=3, column=1, sticky='w', padx=(10, 0))

        sep2 = tk.Frame(tab2, bg=C['border'], height=1)
        sep2.grid(row=4, column=0, columnspan=2, sticky='ew', pady=(14, 4))

        self._section_label(tab2, 5, 'Διακομιστής:')
        self._labeled_entry(tab2, 6, 'SMTP Host:', self._smtp_var, width=24)
        tk.Label(tab2, text='(π.χ. mail.sch.gr)', bg=C['bg'], fg=C['footer'],
                 font=('Arial', 8)).grid(row=7, column=1, sticky='w', padx=(10, 0))

        sep3 = tk.Frame(tab2, bg=C['border'], height=1)
        sep3.grid(row=8, column=0, columnspan=2, sticky='ew', pady=(14, 4))

        self._section_label(tab2, 9, 'Υπογραφή email:')
        self._sig_text = tk.Text(tab2, height=6, width=36,
                                  font=('Consolas', 9), relief='solid', bd=1,
                                  bg='white', fg='#1a1a2e',
                                  wrap='word', padx=6, pady=4)
        self._sig_text.grid(row=10, column=0, columnspan=2, sticky='ew', pady=(4, 2))
        sig_val = self._cfg('EMAIL_SIGNATURE')
        if sig_val:
            self._sig_text.insert('1.0', sig_val)
        tk.Label(tab2, text='(εμφανίζεται στο τέλος κάθε email)',
                 bg=C['bg'], fg=C['footer'], font=('Arial', 8)).grid(
                 row=11, column=0, columnspan=2, sticky='w')

        # ── Tab 3: Αρχεία ─────────────────────────────────────────────────────
        self._section_label(tab3, 0, 'Αρχείο Αδυνατούντων (υπό έγκριση):')

        current_ady = self._cfg('ADY_XORIS_EGKRISI_PATH')
        current_lbl = (os.path.basename(current_ady)
                       if current_ady and os.path.exists(current_ady)
                       else '— δεν έχει οριστεί —')

        ady_frame = tk.Frame(tab3, bg=C['bg'])
        ady_frame.grid(row=1, column=0, columnspan=2, sticky='w', pady=(4, 0))

        self._ady_lbl = tk.Label(ady_frame, text=current_lbl,
                                  bg=C['bg'], fg=C['desc'],
                                  font=('Arial', 8), anchor='w', wraplength=240)
        self._ady_lbl.pack(side='left', padx=(0, 10))

        tk.Button(ady_frame, text='Αλλαγή...',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 8), relief='flat', padx=8, pady=2,
                  cursor='hand2',
                  command=self._browse_ady).pack(side='left')

        tk.Label(tab3,
                 text='Χρησιμοποιείται στον έλεγχο Υπολοίπων.\n'
                      'Ανεβάστε νέο αρχείο μόνο αν άλλαξαν οι αδυνατούντες υπό έγκριση.',
                 bg=C['bg'], fg=C['footer'], font=('Arial', 8),
                 justify='left', anchor='w').grid(
                 row=2, column=0, columnspan=2, sticky='w', pady=(10, 0))

        # ── Κουμπιά ───────────────────────────────────────────────────────────
        btn_frame = tk.Frame(self, bg=C['bg2'], pady=12)
        btn_frame.pack(fill='x')

        tk.Button(btn_frame, text='Αποθήκευση',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 10, 'bold'),
                  relief='flat', padx=18, pady=6,
                  cursor='hand2',
                  command=self._save).pack(side='right', padx=16)

        tk.Button(btn_frame, text='Ακύρωση',
                  bg=C['bg2'], fg=C['desc'],
                  font=('Arial', 10),
                  relief='flat', padx=12, pady=6,
                  cursor='hand2',
                  command=self.destroy).pack(side='right', padx=4)

    def _browse_ady(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title='Αρχείο Αδυνατούντων (υπό έγκριση)',
            filetypes=[('CSV & Excel', '*.csv *.xlsx *.xls'), ('Όλα τα αρχεία', '*.*')],
            parent=self
        )
        if path and os.path.exists(path):
            self._ady_new_path.set(path)
            self._ady_lbl.configure(text=os.path.basename(path), fg=C['status_ok'])

    def _save(self):
        try:
            from_email = self._from_email_var.get().strip()
            updates = {
                'MYSCHOOL_USER'  : self._ms_user_var.get().strip(),
                'MYSCHOOL_PASS'  : self._ms_pass_var.get().strip(),
                'FROM_PASSWORD'  : self._pw_var.get().strip(),
                'FROM_NAME'      : self._from_name_var.get().strip(),
                'FROM_EMAIL'     : from_email,
                'TEST_EMAIL'     : from_email,   # ίδιο με FROM_EMAIL
                'SMTP_HOST'      : self._smtp_var.get().strip(),
                'EMAIL_SIGNATURE': self._sig_text.get('1.0', tk.END).strip(),
                'BROWSER'        : self._browser_var.get(),
            }

            if not updates['FROM_PASSWORD']:
                messagebox.showwarning('Προσοχή', 'Ο κωδικός email δεν μπορεί να είναι κενός.',
                                       parent=self)
                return

            # Αν επιλέχθηκε νέο αρχείο Αδυνατούντων
            new_ady = self._ady_new_path.get().strip()
            if new_ady and os.path.exists(new_ady):
                data_dir = os.path.join(_app_base(), 'data')
                os.makedirs(data_dir, exist_ok=True)
                ext  = os.path.splitext(new_ady)[1]
                dest = os.path.join(data_dir, f'adynatountes_ypo_egkrisi{ext}')
                import shutil as _sh
                _sh.copy2(new_ady, dest)
                updates['ADY_XORIS_EGKRISI_PATH'] = dest

            _save_config(updates)
            messagebox.showinfo('Επιτυχία', 'Οι ρυθμίσεις αποθηκεύτηκαν.', parent=self)
            self.destroy()
        except Exception as e:
            messagebox.showerror('Σφάλμα', f'Δεν ήταν δυνατή η αποθήκευση:\n{e}',
                                 parent=self)


class DownloadDialog(tk.Toplevel):
    """Παράθυρο κατεβάσματος δεδομένων από MySchool."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Λήψη Δεδομένων MySchool')
        self.configure(bg=C['bg'])
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)

        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(ico):
            try: self.iconbitmap(ico)
            except Exception: pass

        self._build()
        self.update_idletasks()
        px = parent.winfo_x() + (parent.winfo_width()  - self.winfo_width())  // 2
        py = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f'+{px}+{py}')

    def _build(self):
        from core.downloader import downloads_info, REPORTS, get_downloads_dir, FILE_PREFIX_MAP
        import glob as _glob
        base_dir = _docs_base()

        # Βρες αρχεία που υπάρχουν ήδη στον σημερινό φάκελο
        today_dir = get_downloads_dir(base_dir)
        already_have = set()
        if os.path.exists(today_dir):
            for rid, prefix in FILE_PREFIX_MAP.items():
                matches = [f for f in _glob.glob(os.path.join(today_dir, f'{prefix}*'))
                           if not f.endswith('.tmp') and not f.endswith('.crdownload')]
                if matches:
                    already_have.add(rid)

        # Header
        hdr = tk.Frame(self, bg=C['hdr_bg'], pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='⬇  Λήψη Δεδομένων MySchool',
                 bg=C['hdr_bg'], fg=C['hdr_fg'],
                 font=('Arial', 12, 'bold')).pack()

        body = tk.Frame(self, bg=C['bg'], padx=20, pady=14)
        body.pack(fill='both')

        # Έλεγχος credentials
        ms_user = getattr(config, 'MYSCHOOL_USER', '').strip()
        ms_pass = getattr(config, 'MYSCHOOL_PASS', '').strip()
        if not ms_user or not ms_pass:
            warn = tk.Frame(body, bg='#FFF3E0',
                            highlightbackground='#FFB74D',
                            highlightthickness=1)
            warn.pack(fill='x', pady=(0, 12))
            tk.Label(warn,
                     text='Τα στοιχεία σύνδεσης MySchool δεν εχουν οριστει. '
                          'Πηγαινε στις Ρυθμισεις για να τα συμπληρωσεις.',
                     bg='#FFF3E0', fg=C['warn'],
                     font=('Arial', 9), padx=10, pady=8,
                     justify='left').pack(anchor='w')

        # Σημερινός φάκελος / τελευταία λήψη
        if already_have:
            from datetime import datetime as _dt
            today_str = _dt.now().strftime('%d/%m/%Y')
            tk.Label(body,
                     text=f'Σήμερα ({today_str}): {len(already_have)}/{len(REPORTS)} αρχεία υπάρχουν ήδη.',
                     bg=C['bg'], fg=C['status_ok'], font=('Arial', 8)).pack(anchor='w', pady=(0,4))
        else:
            info = downloads_info(base_dir)
            if info:
                ts_str, found, age_min = info
                age_txt = f'{age_min} λεπτά' if age_min < 60 else f'{age_min//60}ω {age_min%60}λ'
                tk.Label(body,
                         text=f'Τελευταία λήψη: {ts_str}  ({age_txt} πριν)  —  {len(found)}/{len(REPORTS)} αρχεία',
                         bg=C['bg'], fg=C['desc'], font=('Arial', 8)).pack(anchor='w', pady=(0,4))
            else:
                tk.Label(body, text='Δεν υπάρχουν αποθηκευμένα δεδομένα.',
                         bg=C['bg'], fg=C['desc'], font=('Arial', 8)).pack(anchor='w', pady=(0,4))

        if already_have:
            tk.Label(body,
                     text='Τα ✓ αρχεία υπάρχουν ήδη και θα παραλειφθούν αυτόματα.',
                     bg=C['bg'], fg=C['desc'], font=('Arial', 8, 'italic')).pack(anchor='w', pady=(0,8))

        # Επιλογή αρχείων
        hdr_row = tk.Frame(body, bg=C['bg'])
        hdr_row.pack(fill='x', pady=(0,6))
        tk.Label(hdr_row, text='Επιλέξτε αρχεία για λήψη:',
                 bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold')).pack(side='left')
        tk.Button(hdr_row, text='Όλα',
                  bg=C['bg2'], fg=C['hdr_bg'],
                  font=('Arial', 8), relief='flat', padx=6, pady=1,
                  cursor='hand2',
                  command=lambda: [v.set(True) for v in self._report_vars.values()]
                  ).pack(side='left', padx=(10,0))

        self._report_vars = {}
        grid = tk.Frame(body, bg=C['bg'])
        grid.pack(fill='x')

        for i, (rid, label, *_) in enumerate(REPORTS):
            exists = rid in already_have
            var = tk.BooleanVar(value=False)
            self._report_vars[rid] = var
            row, col = divmod(i, 2)
            prefix   = '' if rid == 'ady' else f'{rid} — '
            lbl_text = f'✓ {prefix}{label}' if exists else f'{prefix}{label}'
            lbl_fg   = C['status_ok'] if exists else C['fg'] if 'fg' in C else '#000000'
            tk.Checkbutton(grid, text=lbl_text,
                           variable=var,
                           fg=lbl_fg,
                           bg=C['bg'], font=('Arial', 9),
                           activebackground=C['bg']).grid(
                           row=row, column=col, sticky='w', padx=(0,16), pady=2)

        # Progress label
        self._progress_var = tk.StringVar(value='')
        self._progress_lbl = tk.Label(body, textvariable=self._progress_var,
                                       bg=C['bg'], fg=C['status_run'],
                                       font=('Arial', 8), wraplength=380, justify='left')
        self._progress_lbl.pack(anchor='w', pady=(10,0))

        # Κουμπιά
        btn_frame = tk.Frame(self, bg=C['bg2'], pady=12)
        btn_frame.pack(fill='x')

        self._start_btn = tk.Button(btn_frame, text='⬇  Έναρξη Λήψης',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 10, 'bold'),
                  relief='flat', padx=18, pady=6,
                  cursor='hand2',
                  command=self._start).pack(side='right', padx=16)

        tk.Button(btn_frame, text='Κλείσιμο',
                  bg=C['bg2'], fg=C['desc'],
                  font=('Arial', 10),
                  relief='flat', padx=12, pady=6,
                  cursor='hand2',
                  command=self.destroy).pack(side='right', padx=4)

    def _start(self):
        from core.downloader import (MySchoolDownloader,
                                      get_downloads_dir,
                                      cleanup_old_downloads)

        ms_user = getattr(config, 'MYSCHOOL_USER', '').strip()
        ms_pass = getattr(config, 'MYSCHOOL_PASS', '').strip()

        if not ms_user or not ms_pass:
            messagebox.showwarning('Προσοχή',
                'Συμπλήρωσε username και κωδικό MySchool στις Ρυθμίσεις (⚙).',
                parent=self)
            return

        selected = [rid for rid, var in self._report_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning('Προσοχή', 'Επίλεξε τουλάχιστον ένα αρχείο.', parent=self)
            return

        base_dir = _docs_base()
        dest_dir = get_downloads_dir(base_dir)

        self._progress_var.set('Εκκίνηση...')
        self.update()

        def on_progress(msg):
            self.after(0, lambda m=msg: self._progress_var.set(m))

        def task():
            try:
                dl = MySchoolDownloader(
                    username=ms_user,
                    password=ms_pass,
                    dest_dir=dest_dir,
                    callback=on_progress,
                    reports=selected,
                    browser=getattr(config, 'BROWSER', 'chrome'),
                )
                results = dl.run()
                ok   = sum(1 for v in results.values() if v)
                fail = len(results) - ok

                # Κράτα μόνο τον τελευταίο φάκελο
                cleanup_old_downloads(base_dir, keep=1)

                msg = f'Ολοκληρώθηκε: {ok}/{len(results)} αρχεία κατεβήκαν.'
                if fail:
                    msg += f' Αποτυχίες: {fail}.'
                self.after(0, lambda m=msg: [
                    self._progress_var.set(m),
                    messagebox.showinfo('Λήψη', m, parent=self)
                ])
            except Exception as e:
                err = str(e)
                self.after(0, lambda m=err: [
                    self._progress_var.set(f'Σφάλμα: {m}'),
                    messagebox.showerror('Σφάλμα Λήψης', m, parent=self)
                ])

        threading.Thread(target=task, daemon=True).start()


class LauncherApp:

    def __init__(self, root, checks):
        self.root         = root
        self.checks       = checks
        self.indicators   = []
        self.check_frames = []
        self._status_q    = queue.Queue()

        root.title('MySchool Αυτοματισμοί')
        root.configure(bg=C['bg'])
        root.resizable(False, False)

        self._build_ui()

        # Παγώνουμε το μέγεθος μετά το render ώστε να μην αλλάζει
        root.update_idletasks()
        root.geometry(f'{root.winfo_width()}x{root.winfo_height()}')

        self._poll_status()
        # Σύνδεση stdout με status bar
        _gui_stream.set_callback(self._on_print)

        # Έλεγχος για νέα έκδοση (background, αθόρυβος — 3s μετά την εκκίνηση)
        def _on_update(new_ver, dl_url):
            def _show():
                self._set_status(
                    f'⬆  Διαθέσιμη νέα έκδοση v{new_ver} — κλικ για ενημέρωση',
                    C['status_run'])
                self.status_lbl.configure(cursor='hand2')
                self.status_lbl.bind('<Button-1>',
                    lambda e: _do_update(self.root, new_ver, dl_url))
            self.root.after(0, _show)
        root.after(3000, lambda: _check_for_update(_on_update))

        # Αν δεν έχει οριστεί κωδικός, άνοιξε αυτόματα τις ρυθμίσεις
        if not password_is_set():
            root.after(400, self._open_settings)

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self.root, bg=C['hdr_bg'], pady=14)
        hdr.pack(fill='x')

        # Μόνο ⚙ στο header
        btn_hdr = tk.Frame(hdr, bg=C['hdr_bg'])
        btn_hdr.place(relx=1.0, x=-6, y=2, anchor='ne')

        tk.Button(btn_hdr, text='⚙',
                  bg=C['hdr_bg'], fg=C['hdr_sub'],
                  font=('Arial', 13), relief='flat', cursor='hand2',
                  activebackground=C['hdr_bg'], activeforeground='white',
                  command=self._open_settings).pack(side='right', padx=(2, 0))

        tk.Button(btn_hdr, text='?',
                  bg=C['hdr_bg'], fg=C['hdr_sub'],
                  font=('Arial', 13, 'bold'), relief='flat', cursor='hand2',
                  activebackground=C['hdr_bg'], activeforeground='white',
                  command=self._open_help).pack(side='right', padx=(0, 2))

        tk.Label(hdr, text='MySchool Αυτοματισμοί',
                 bg=C['hdr_bg'], fg=C['hdr_fg'],
                 font=('Arial', 15, 'bold')).pack()
        tk.Label(hdr, text='Δ/νση Π.Ε. Ανατολικής Θεσσαλονίκης',
                 bg=C['hdr_bg'], fg=C['hdr_sub'],
                 font=('Arial', 9)).pack()
        tk.Label(hdr, text=f'v{config.APP_VERSION}',
                 bg=C['hdr_bg'], fg=C['hdr_sub'],
                 font=('Arial', 7)).pack()

        # Ένδειξη αν ο κωδικός λείπει
        if not password_is_set():
            warn_bar = tk.Frame(self.root, bg='#FFF3E0',
                                highlightbackground='#FFB74D',
                                highlightthickness=1)
            warn_bar.pack(fill='x')
            tk.Label(warn_bar,
                     text='⚠  Ο κωδικός email δεν έχει οριστεί — κλικ στο ⚙ για να τον ορίσεις',
                     bg='#FFF3E0', fg=C['warn'],
                     font=('Arial', 8), padx=10, pady=4).pack(side='left')

        # Toolbar
        toolbar = tk.Frame(self.root, bg=C['bg2'], pady=6)
        toolbar.pack(fill='x')
        tk.Button(toolbar, text='⬇  Λήψη Δεδομένων',
                  bg=C['bg2'], fg=C['hdr_bg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=14, pady=4, cursor='hand2',
                  activebackground=C['sel_bg'], activeforeground=C['hdr_bg'],
                  command=self._open_download).pack(side='left', padx=(6, 0))
        tk.Label(toolbar, text='|', bg=C['bg2'], fg=C['desc'],
                 font=('Arial', 9)).pack(side='left', padx=4)
        tk.Button(toolbar, text='📋  Εκπ/κοί ανά Ειδικότητα',
                  bg=C['bg2'], fg=C['hdr_bg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=14, pady=4, cursor='hand2',
                  activebackground=C['sel_bg'], activeforeground=C['hdr_bg'],
                  command=self._open_eidikotita_tool).pack(side='left', padx=(0, 0))
        tk.Button(toolbar, text='🏫  Σχολικές Μονάδες',
                  bg=C['bg2'], fg=C['hdr_bg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=14, pady=4, cursor='hand2',
                  activebackground=C['sel_bg'], activeforeground=C['hdr_bg'],
                  command=self._open_monada_tool).pack(side='left', padx=(0, 0))

        # Body
        body = tk.Frame(self.root, bg=C['bg'], padx=18, pady=14)
        body.pack(fill='both')

        lbl_row = tk.Frame(body, bg=C['bg'])
        lbl_row.pack(fill='x', pady=(0, 4))
        tk.Label(lbl_row, text='Επιλέξτε ένα ή περισσότερους ελέγχους:',
                 bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 10, 'bold'), anchor='w').pack(side='left')
        tk.Button(lbl_row, text='Όλοι',
                  bg=C['hdr_bg'], fg='white',
                  font=('Arial', 8, 'bold'), relief='flat',
                  padx=8, pady=2, cursor='hand2',
                  command=self._select_all).pack(side='right')

        self._check_vars = []

        for i, (title, desc, mod) in enumerate(self.checks):
            var = tk.BooleanVar(value=(i == 0))
            self._check_vars.append(var)

            f = tk.Frame(body, bg=C['norm_bg'],
                         highlightbackground=C['norm_bd'],
                         highlightthickness=1,
                         pady=6, padx=10)
            f.pack(fill='x', pady=3)
            self.check_frames.append(f)

            top = tk.Frame(f, bg=C['norm_bg'])
            top.pack(fill='x')

            ind = Indicator(top, bg=C['norm_bg'])
            ind.pack(side='left', padx=(0, 6))
            self.indicators.append(ind)

            # Κουμπί επεξεργασίας email (μόνο για ελέγχους με email)
            if getattr(mod, 'HAS_EMAIL', False):
                mod_name = mod.__name__.split('.')[-1]
                tk.Button(top, text='✏',
                          bg=C['norm_bg'], fg=C['hdr_bg'],
                          font=('Arial', 10), relief='flat', cursor='hand2',
                          activebackground=C['sel_bg'],
                          command=lambda m=mod, mn=mod_name: self._open_email_editor(mn, m)
                          ).pack(side='right', padx=(4, 0))

            cb = tk.Checkbutton(top, text=title,
                                variable=var,
                                bg=C['norm_bg'], selectcolor=C['sel_bg'],
                                activebackground=C['norm_bg'],
                                font=('Arial', 10), anchor='w',
                                command=self._refresh_highlights)
            cb.pack(fill='x', expand=True)

            if desc:
                tk.Label(f, text=desc, bg=C['norm_bg'], fg=C['desc'],
                         font=('Arial', 8), anchor='w',
                         wraplength=430, justify='left').pack(fill='x', padx=20)

        self._refresh_highlights()

        # Κουμπί εκκίνησης
        btn_frame = tk.Frame(self.root, bg=C['bg'], pady=10)
        btn_frame.pack()
        self.btn_run = tk.Button(btn_frame,
                                  text='▶  Εκκίνηση ελέγχου',
                                  font=('Arial', 11, 'bold'),
                                  bg=C['btn_bg'], fg=C['btn_fg'],
                                  activebackground=C['btn_act'],
                                  padx=26, pady=9,
                                  relief='flat',
                                  cursor='hand2',
                                  command=self._run)
        self.btn_run.pack()

        # Κουμπί ρυθμίσεων ⚙ στο header
        tk.Button(hdr, text='⚙', font=('Arial', 11),
                  bg=C['hdr_bg'], fg=C['hdr_sub'],
                  activebackground=C['btn_act'],
                  activeforeground=WHITE,
                  relief='flat', cursor='hand2',
                  bd=0, padx=6,
                  command=self._open_settings).place(relx=1.0, rely=0.0,
                                                      anchor='ne', x=-8, y=6)

        # Status bar
        self.status_var = tk.StringVar(value='Έτοιμο')
        status_bar = tk.Frame(self.root, bg=C['bg2'],
                              highlightbackground=C['border'],
                              highlightthickness=1)
        status_bar.pack(fill='x', side='bottom')
        self.status_lbl = tk.Label(status_bar,
                                    textvariable=self.status_var,
                                    bg=C['bg2'], fg=C['footer'],
                                    font=('Arial', 8), anchor='w',
                                    padx=10, pady=4)
        self.status_lbl.pack(side='left')
        tk.Label(status_bar,
                 text=f'{len(self.checks)} έλεγχοι  •  Μιχάλης Κατσιρντάκης  •  2310954145',
                 bg=C['bg2'], fg=C['footer'],
                 font=('Arial', 7), padx=10).pack(side='right')

    def _open_settings(self):
        SettingsDialog(self.root)

    def _open_download(self):
        DownloadDialog(self.root)

    def _open_eidikotita_tool(self):
        EidikotitaDialog(self.root)

    def _open_monada_tool(self):
        MonadaDialog(self.root)

    def _refresh_highlights(self):
        for i, (f, ind) in enumerate(zip(self.check_frames, self.indicators)):
            sel = self._check_vars[i].get()
            bg  = C['sel_bg']   if sel else C['norm_bg']
            bd  = C['sel_bd']   if sel else C['norm_bd']
            f.configure(bg=bg, highlightbackground=bd)
            for w in f.winfo_children():
                try:
                    w.configure(bg=bg)
                    for ww in w.winfo_children():
                        try: ww.configure(bg=bg)
                        except Exception: pass
                except Exception:
                    pass
            ind.configure(bg=bg)

    def _select_all(self):
        for var in self._check_vars:
            var.set(True)
        self._refresh_highlights()

    # ── Email template editor ────────────────────────────────────────────────

    def _get_default_email_body(self, module):
        """Επιστρέφει το default body text χωρίς υπογραφή."""
        body_t = getattr(module, 'EMAIL_BODY', '')
        try:
            full = body_t('') if callable(body_t) else body_t
            sig  = config.email_signature()
            if sig and full.endswith(sig):
                return full[:-len(sig)]
            return full
        except Exception:
            return ''

    def _open_email_editor(self, mod_name, module):
        """Dialog επεξεργασίας email template για συγκεκριμένο έλεγχο."""
        import json

        # Φόρτωσε τρέχον template (custom ή default)
        settings = _load_local_settings()
        templates = settings.get('email_templates', {})
        custom = templates.get(mod_name)

        if custom:
            cur_subject = custom.get('subject', '')
            cur_body    = custom.get('body', '')
        else:
            cur_subject = getattr(module, 'EMAIL_SUBJECT', '')
            cur_body    = self._get_default_email_body(module)

        title_str = getattr(module, 'CHECK_TITLE', mod_name)

        # Παράθυρο
        dlg = tk.Toplevel(self.root)
        dlg.title(f'Πρότυπο Email — {title_str}')
        dlg.configure(bg=C['bg'])
        dlg.resizable(True, False)
        dlg.grab_set()
        dlg.transient(self.root)

        pad = dict(padx=14, pady=5)

        tk.Label(dlg, text='Θέμα:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **pad)

        subj_var = tk.StringVar(value=cur_subject)
        tk.Entry(dlg, textvariable=subj_var, font=('Arial', 9),
                 width=60).pack(fill='x', padx=14, pady=(0, 8))

        tk.Label(dlg, text='Κείμενο email:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', **pad)

        txt = tk.Text(dlg, font=('Arial', 9), width=60, height=10,
                      wrap='word', relief='solid', bd=1)
        txt.pack(fill='x', padx=14, pady=(0, 4))
        txt.insert('1.0', cur_body)

        tk.Label(dlg, text='Η υπογραφή σας προστίθεται αυτόματα στο τέλος.',
                 bg=C['bg'], fg=C['desc'], font=('Arial', 8),
                 anchor='w').pack(fill='x', padx=14, pady=(0, 10))

        def _save():
            new_subj = subj_var.get().strip()
            new_body = txt.get('1.0', 'end-1c')
            s = _load_local_settings()
            s.setdefault('email_templates', {})[mod_name] = {
                'subject': new_subj,
                'body':    new_body,
            }
            path = _get_local_settings_path()
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(s, f, ensure_ascii=False, indent=2)
            dlg.destroy()
            messagebox.showinfo('Αποθήκευση', 'Το πρότυπο email αποθηκεύτηκε.',
                                parent=self.root)

        def _reset():
            if messagebox.askyesno('Επαναφορά', 'Να επανέλθει το προεπιλεγμένο κείμενο;',
                                   parent=dlg):
                s = _load_local_settings()
                s.get('email_templates', {}).pop(mod_name, None)
                path = _get_local_settings_path()
                os.makedirs(os.path.dirname(path), exist_ok=True)
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(s, f, ensure_ascii=False, indent=2)
                dlg.destroy()

        btn_row = tk.Frame(dlg, bg=C['bg'])
        btn_row.pack(pady=(0, 12))

        tk.Button(btn_row, text='Αποθήκευση',
                  bg=C['btn_bg'], fg=C['btn_fg'],
                  font=('Arial', 9, 'bold'), relief='flat',
                  padx=14, pady=5, cursor='hand2',
                  command=_save).pack(side='left', padx=4)

        tk.Button(btn_row, text='Επαναφορά προεπιλογής',
                  bg=C['bg2'], fg=C['hdr_bg'],
                  font=('Arial', 9), relief='flat',
                  padx=14, pady=5, cursor='hand2',
                  command=_reset).pack(side='left', padx=4)

        tk.Button(btn_row, text='Άκυρο',
                  bg=C['bg2'], fg=C['desc'],
                  font=('Arial', 9), relief='flat',
                  padx=14, pady=5, cursor='hand2',
                  command=dlg.destroy).pack(side='left', padx=4)

        dlg.update_idletasks()
        # Κεντράρισμα
        w, h = dlg.winfo_width(), dlg.winfo_height()
        x = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
        dlg.geometry(f'+{x}+{y}')

    def _open_help(self):
        """Παράθυρο βοήθειας — εμφανίζει το README.md."""
        _show_help(self.root)

    def _open_settings(self):
        """Παράθυρο ρυθμίσεων (email + MySchool credentials)."""
        SettingsDialog(self.root)

    def _run(self):
        selected = [i for i, v in enumerate(self._check_vars) if v.get()]
        if not selected:
            messagebox.showinfo('Επιλογή', 'Επιλέξτε τουλάχιστον έναν έλεγχο.', parent=self.root)
            return

        multi = len(selected) > 1
        self.btn_run.config(state='disabled', bg=C['btn_dis'], text='  Εκτελείται...')

        def task():
            import core.framework as _fw
            _fw._multi_run_mode    = multi
            _fw._multi_run_results = [] if multi else _fw._multi_run_results

            for pos, idx in enumerate(selected, 1):
                _, _, mod = self.checks[idx]
                self._status_q.put(('running', idx, pos, len(selected)))
                try:
                    if getattr(mod, 'CUSTOM_RUN', False):
                        mod.run(config)
                    else:
                        from core.framework import run_check
                        run_check(mod, config)
                    self._status_q.put(('chk_ok', idx))
                except SystemExit:
                    self._status_q.put(('chk_ok', idx))
                except Exception as e:
                    import traceback
                    self._status_q.put(('chk_err', idx, str(e), traceback.format_exc()))

            if multi:
                _fw._multi_run_mode = False
                results = list(_fw._multi_run_results)
            else:
                results = None

            self._status_q.put(('all_done', selected, results))

        threading.Thread(target=task, daemon=True).start()

    def _poll_status(self):
        try:
            while True:
                msg  = self._status_q.get_nowait()
                kind = msg[0]

                if kind == 'running':
                    _, idx, pos, total = msg
                    self.indicators[idx].set_state('running')
                    title = self.checks[idx][0]
                    suffix = f' [{pos}/{total}]' if total > 1 else ''
                    self._set_status(f'Εκτέλεση{suffix}: {title}', C['status_run'])
                    self.btn_run.config(text=f'  Εκτελείται {pos}/{total}...' if total > 1
                                             else '  Εκτελείται...')

                elif kind == 'chk_ok':
                    self.indicators[msg[1]].set_state('ok')

                elif kind == 'chk_err':
                    _, idx, err, tb = msg
                    self.indicators[idx].set_state('error')
                    self._set_status(f'Σφάλμα: {err}', C['status_err'])
                    messagebox.showerror('Σφάλμα', f'{err}\n\n{tb[-600:]}', parent=self.root)

                elif kind == 'all_done':
                    _, selected, results = msg
                    self.btn_run.config(state='normal', bg=C['btn_bg'],
                                        text='▶  Εκκίνηση ελέγχου')
                    if results is None:
                        # single run — popup εμφανίστηκε ήδη από το framework
                        self._set_status(
                            f'Ολοκληρώθηκε: {self.checks[selected[0]][0]}', C['status_ok'])
                    else:
                        # multi-run
                        n = len(selected)
                        self._set_status(f'Ολοκληρώθηκαν {n} έλεγχοι', C['status_ok'])
                        self.root.after(0, lambda r=results, n=n: self._ask_show_results(r, n))

        except queue.Empty:
            pass
        self.root.after(100, self._poll_status)

    def _ask_show_results(self, results, total_ran):
        answer = messagebox.askyesno(
            'Αποτελέσματα',
            f'Θέλεις να δεις τα αποτελέσματα;\n(Σύνολο: {total_ran} έλεγχοι που έτρεξαν)',
            parent=self.root
        )
        if answer and results:
            self._show_results_navigator(results)

    def _show_results_navigator(self, results):
        import tkinter as tk
        from tkinter import scrolledtext

        COLORS = {
            'ok':   ('#E8F5E9', '#2E7D32'),
            'warn': ('#FFF8E1', '#E65100'),
        }
        total   = len(results)
        cur     = [0]

        win = tk.Toplevel(self.root)
        win.resizable(True, True)
        win.grab_set()
        win.attributes('-topmost', True)
        _ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
        if os.path.exists(_ico):
            try: win.iconbitmap(_ico)
            except Exception: pass
        win.update_idletasks()
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        win.geometry(f'560x500+{sw//2-280}+{sh//2-250}')

        # Header
        hdr_frame = tk.Frame(win)
        hdr_frame.pack(fill='x')
        hdr_lbl = tk.Label(hdr_frame, fg='white', font=('Arial', 11, 'bold'), pady=8)
        hdr_lbl.pack(fill='x')

        # Μετρητής
        counter_lbl = tk.Label(win, font=('Arial', 9, 'bold'))
        counter_lbl.pack(pady=(4, 0))

        # Κουμπί κλεισίματος — pack πρώτα (bottom)
        nav_frame = tk.Frame(win)
        nav_frame.pack(side='bottom', pady=(4, 12))

        btn_prev = tk.Button(nav_frame, text='◀  Προηγούμενο',
                             font=('Arial', 9), relief='flat',
                             padx=12, pady=5, cursor='hand2',
                             command=lambda: navigate(-1))
        btn_prev.pack(side='left', padx=4)

        btn_next = tk.Button(nav_frame, text='Συνέχεια  ▶',
                             font=('Arial', 9, 'bold'), relief='flat',
                             padx=12, pady=5, cursor='hand2',
                             command=lambda: navigate(+1))
        btn_next.pack(side='left', padx=4)

        btn_close = tk.Button(nav_frame, text='Κλείσιμο',
                              font=('Arial', 9), relief='flat',
                              padx=12, pady=5, cursor='hand2',
                              command=win.destroy)
        btn_close.pack(side='left', padx=4)

        # Body text
        txt = scrolledtext.ScrolledText(
            win, wrap=tk.WORD, font=('Consolas', 9),
            relief='flat', bd=0, padx=14, pady=10)
        txt.pack(fill='both', expand=True, padx=10, pady=(6, 4))

        def show(i):
            title, body, rtype = results[i]
            bg, hdr_bg = COLORS.get(rtype, COLORS['warn'])
            icon = '✓' if rtype == 'ok' else '⚠'

            win.configure(bg=bg)
            hdr_frame.configure(bg=hdr_bg)
            hdr_lbl.configure(text=f'{icon}  {title}', bg=hdr_bg)
            counter_lbl.configure(text=f'{i + 1} / {total}', bg=bg, fg=hdr_bg)
            nav_frame.configure(bg=bg)

            txt.configure(state='normal', bg=bg, fg='#212121')
            txt.delete('1.0', tk.END)
            txt.insert('1.0', body)
            txt.configure(state='disabled')

            # Προηγούμενο: ορατό μόνο από το 2ο
            if i == 0:
                btn_prev.pack_forget()
            else:
                btn_prev.pack(side='left', padx=4)
                btn_prev.configure(bg=hdr_bg, fg='white')

            # Συνέχεια vs Κλείσιμο
            if i == total - 1:
                btn_next.pack_forget()
                btn_close.configure(bg=hdr_bg, fg='white', font=('Arial', 9, 'bold'))
            else:
                btn_next.pack(side='left', padx=4)
                btn_next.configure(bg=hdr_bg, fg='white')
                btn_close.configure(bg=bg, fg='#777', font=('Arial', 9))

        def navigate(direction):
            cur[0] = max(0, min(total - 1, cur[0] + direction))
            show(cur[0])

        show(0)
        win.wait_window()

    def _on_print(self, text):
        """Λαμβάνει μηνύματα από sys.stdout και τα εμφανίζει στο status bar."""
        self.root.after(0, lambda t=text: self._set_status(t))

    def _set_status(self, text, color=None):
        self.status_var.set(text)
        if color:
            self.status_lbl.configure(fg=color)


def _check_for_update(on_update_available):
    """Ελέγχει αν υπάρχει νεότερη έκδοση στο GitHub. Τρέχει σε background thread.
    Αν βρεθεί νεότερη, καλεί on_update_available(new_version, download_url)."""
    def _task():
        try:
            import urllib.request, json as _json
            api_url = 'https://api.github.com/repos/MichalisKat/myschool-checks/releases/latest'
            req = urllib.request.Request(api_url, headers={'User-Agent': 'MySchoolChecks'})
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = _json.loads(resp.read().decode())
            latest = data.get('tag_name', '').lstrip('v')
            if not latest:
                return
            current = getattr(config, 'APP_VERSION', '0.0.0')
            def _ver(s):
                try: return tuple(int(x) for x in s.split('.'))
                except: return (0,)
            if _ver(latest) > _ver(current):
                # Βρες το .exe asset
                assets = data.get('assets', [])
                dl_url = next(
                    (a['browser_download_url'] for a in assets
                     if a['name'].endswith('.exe')),
                    None)
                if dl_url:
                    on_update_available(latest, dl_url)
        except Exception:
            pass  # Αθόρυβη αποτυχία — δεν επηρεάζει τη λειτουργία
    threading.Thread(target=_task, daemon=True).start()


def _do_update(parent, new_ver, dl_url):
    """Κατεβάζει το νέο setup.exe και το εκτελεί. Εμφανίζει progress dialog."""
    import urllib.request, tempfile, subprocess as _sub

    # Ερώτηση επιβεβαίωσης
    answer = messagebox.askyesno(
        'Διαθέσιμη ενημέρωση',
        f'Διαθέσιμη νέα έκδοση v{new_ver}!\n\n'
        f'Θέλεις να κατεβάσεις και να εγκαταστήσεις τώρα;\n\n'
        f'Η εφαρμογή θα κλείσει αυτόματα για την εγκατάσταση.',
        parent=parent)
    if not answer:
        return

    # Progress dialog
    dlg = tk.Toplevel(parent)
    dlg.title('Λήψη ενημέρωσης')
    dlg.configure(bg=C['bg'])
    dlg.resizable(False, False)
    dlg.grab_set()
    dlg.transient(parent)

    tk.Label(dlg, text=f'Λήψη MySchool Checks v{new_ver}...',
             bg=C['bg'], fg=C['hdr_bg'],
             font=('Arial', 10, 'bold')).pack(padx=24, pady=(18, 6))

    from tkinter import ttk as _ttk
    pb = _ttk.Progressbar(dlg, length=320, mode='determinate')
    pb.pack(padx=24, pady=(0, 6))

    status_var = tk.StringVar(value='Σύνδεση...')
    tk.Label(dlg, textvariable=status_var,
             bg=C['bg'], fg=C['footer'],
             font=('Arial', 8)).pack(padx=24, pady=(0, 18))

    dlg.update_idletasks()
    px = parent.winfo_x() + (parent.winfo_width()  - dlg.winfo_width())  // 2
    py = parent.winfo_y() + (parent.winfo_height() - dlg.winfo_height()) // 2
    dlg.geometry(f'+{px}+{py}')

    def _download():
        try:
            tmp_dir  = tempfile.mkdtemp()
            fname    = f'myschool-checks-{new_ver}-setup.exe'
            tmp_path = os.path.join(tmp_dir, fname)

            def _reporthook(count, block_size, total_size):
                if total_size > 0:
                    pct = min(100, int(count * block_size * 100 / total_size))
                    mb_done = count * block_size / 1_048_576
                    mb_total = total_size / 1_048_576
                    dlg.after(0, lambda p=pct, d=mb_done, t=mb_total: [
                        pb.configure(value=p),
                        status_var.set(f'{d:.1f} / {t:.1f} MB  ({p}%)')])

            urllib.request.urlretrieve(dl_url, tmp_path, _reporthook)

            # Κατεβάστηκε — εκτέλεση
            dlg.after(0, lambda: [
                status_var.set('Εκκίνηση εγκατάστασης...'),
                dlg.update()])
            import time as _t; _t.sleep(0.8)

            import ctypes
            ctypes.windll.shell32.ShellExecuteW(None, 'runas', tmp_path, None, None, 1)
            dlg.after(0, lambda: [dlg.destroy(), parent.destroy()])

        except Exception as e:
            dlg.after(0, lambda err=str(e): [
                dlg.destroy(),
                messagebox.showerror('Σφάλμα λήψης',
                    f'Δεν ήταν δυνατή η λήψη:\n{err}\n\n'
                    'Κατέβασέ την χειροκίνητα από:\n'
                    'github.com/MichalisKat/myschool-checks/releases',
                    parent=parent)])

    threading.Thread(target=_download, daemon=True).start()


class EidikotitaDialog(tk.Toplevel):
    """Εργαλείο εξαγωγής εκπαιδευτικών ανά ειδικότητα."""

    _SETTINGS_KEY   = 'eidikotita_tool'
    _DEFAULT_BODY   = (
        'Αποτύπωση Myschool {date}.\n\n'
        'Καλημέρα σας,\n\n'
        'Επισυνάπτω πίνακα excel με τους εκπαιδευτικούς ειδικότητας {specialty} '
        'που υπηρετούν στην Δ/νση Α/θμιας Αν. Θεσ/κης σύμφωνα με τα καταχωρημένα '
        'στοιχεία στο myschool.\n\n\n'
        'Στη διάθεσή σας για οποιαδήποτε πληροφορία'
    )
    _DEFAULT_SUBJECT = 'Στοιχεία τοποθετήσεων εκπ/κών "{specialty}" σε σχολικές μονάδες'

    # Σταθερές στήλες εξόδου
    _OUT_COLS = [
        'ΑΜ',
        'Επώνυμο', 'Όνομα', 'Κύρια Ειδικ.',
        'Email στο ΠΣΔ', 'Email', 'Κινητό',
        'Σχέση εργασίας', 'Σχέση τοποθέτησης',
        'Κατάσταση',
        'Φορέας τοποθέτησης',
        'Τηλέφωνο', 'e-mail',
        'ΑΠΟΥΣΙΑ', 'Έως',
    ]

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Εκπ/κοί ανά Ειδικότητα')
        self.configure(bg=C['bg'])
        self.resizable(True, True)
        self.grab_set()
        self.transient(parent)
        self._parent = parent

        s = _load_local_settings().get(self._SETTINGS_KEY, {})
        self._saved_subject = s.get('subject',       self._DEFAULT_SUBJECT)
        self._saved_body    = s.get('body',          self._DEFAULT_BODY)
        self._saved_email   = s.get('advisor_email', '')

        # Αυτόματη εύρεση αρχείων από downloads
        self._topoth_path = self._auto_find('Topothetiseis')
        self._grid_path   = self._auto_find('gridResults')
        self._stat_path   = self._auto_find('stat4_16')
        self._stat41_path = self._auto_find('stat4_1')
        self._stat42_path = self._auto_find('stat4_2')

        self._build_form()
        self.update_idletasks()
        w, h = 620, 500
        x = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        y = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f'{w}x{h}+{x}+{y}')

    # ── Auto-find ────────────────────────────────────────────────────────────

    @staticmethod
    def _auto_find(prefix):
        """Ψάχνει το αρχείο στους φακέλους downloads (νεότερος πρώτα), μετά ~/Downloads."""
        import glob as _glob
        dl_base = os.path.join(_docs_base(), 'downloads')
        if os.path.isdir(dl_base):
            folders = sorted([
                os.path.join(dl_base, d)
                for d in os.listdir(dl_base)
                if os.path.isdir(os.path.join(dl_base, d))
            ], reverse=True)
            for folder in folders:
                matches = [f for f in _glob.glob(os.path.join(folder, f'{prefix}*'))
                           if not f.endswith('.tmp') and not f.endswith('.crdownload')]
                if matches:
                    return sorted(matches)[-1]
        dl_user = os.path.join(os.path.expanduser('~'), 'Downloads')
        matches = [f for f in _glob.glob(os.path.join(dl_user, f'*{prefix}*'))
                   if not f.endswith('.tmp') and not f.endswith('.crdownload')]
        return sorted(matches)[-1] if matches else ''

    # ── Κύρια φόρμα ──────────────────────────────────────────────────────────

    def _build_form(self):
        self._clear()

        tk.Label(self, text='Εκπαιδευτικοί ανά Ειδικότητα',
                 bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 11, 'bold')).pack(anchor='w', padx=18, pady=(14, 4))

        # Προειδοποίηση αν λείπουν κρίσιμα αρχεία
        missing = []
        if not self._topoth_path: missing.append('Τοποθετήσεις')
        if not self._grid_path:   missing.append('Κατάλογος σχολείων (2.1)')
        if missing:
            warn = tk.Label(self,
                text=f'⚠  Δεν βρέθηκαν: {", ".join(missing)}. Κατέβασε τα πρώτα από «Λήψη Δεδομένων».',
                bg='#FFF3E0', fg='#E65100', font=('Arial', 8), anchor='w', padx=10, pady=5,
                wraplength=560, justify='left')
            warn.pack(fill='x', padx=18, pady=(0, 6))

        # ── Ειδικότητα ───────────────────────────────────────────────────────
        tk.Label(self, text='Ειδικότητα:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))

        spec_row = tk.Frame(self, bg=C['bg'])
        spec_row.pack(fill='x', padx=18, pady=(2, 6))

        self._spec_var = tk.StringVar()
        from tkinter import ttk as _ttk
        self._spec_combo = _ttk.Combobox(spec_row, textvariable=self._spec_var,
                                          width=46, state='readonly')
        self._spec_combo.pack(side='left')

        self._spec_lbl = tk.Label(spec_row, text='Φόρτωση…', bg=C['bg'],
                                   fg=C['desc'], font=('Arial', 8))
        self._spec_lbl.pack(side='left', padx=(10, 0))

        # Όταν αλλάζει η ειδικότητα → ενημέρωσε το θέμα
        self._spec_var.trace_add('write', self._on_spec_change)

        # ── Στήλες εξόδου ────────────────────────────────────────────────────
        tk.Label(self, text='Στήλες εξόδου:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(6, 2))

        col_frame = tk.Frame(self, bg=C['bg'])
        col_frame.pack(fill='x', padx=18, pady=(0, 6))

        self._col_vars = {}
        for col_name in ('Email στο ΠΣΔ', 'Email', 'Κινητό'):
            var = tk.BooleanVar(value=True)
            self._col_vars[col_name] = var
            tk.Checkbutton(col_frame, text=col_name, variable=var,
                           bg=C['bg'], font=('Arial', 9),
                           activebackground=C['bg']).pack(side='left', padx=(0, 12))

        # ── Email ─────────────────────────────────────────────────────────────
        pad = dict(padx=18, pady=2)

        tk.Label(self, text='Προς (email συμβούλου):',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._to_var = tk.StringVar(value=self._saved_email)
        tk.Entry(self, textvariable=self._to_var,
                 font=('Arial', 9)).pack(fill='x', padx=18, pady=(0, 6))

        tk.Label(self, text='Θέμα:',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._subj_var = tk.StringVar(value=self._saved_subject)
        tk.Entry(self, textvariable=self._subj_var,
                 font=('Arial', 9)).pack(fill='x', padx=18, pady=(0, 6))

        tk.Label(self, text='Κείμενο email:',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._body_txt = tk.Text(self, font=('Arial', 9), height=6,
                                  wrap='word', relief='solid', bd=1)
        self._body_txt.pack(fill='x', padx=18, pady=(0, 6))
        from datetime import datetime as _dt
        self._body_txt.insert('1.0',
            self._saved_body.replace('{date}', _dt.today().strftime('%d/%m/%Y'))
                             .replace('{specialty}', self._spec_var.get()))

        btn_row = tk.Frame(self, bg=C['bg'])
        btn_row.pack(side='bottom', pady=10)
        tk.Button(btn_row, text='Μόνο Excel (χωρίς email)',
                  bg=C['bg2'], fg=C['hdr_bg'], relief='flat',
                  font=('Arial', 9), padx=10, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=False)).pack(side='left', padx=4)
        tk.Button(btn_row, text='▶  Δημιουργία & Αποστολή',
                  bg=C['btn_bg'], fg=C['btn_fg'], relief='flat',
                  font=('Arial', 9, 'bold'), padx=14, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=True)).pack(side='left', padx=4)

        self.after(100, self._load_specialties)

    def _on_spec_change(self, *_):
        """Ενημερώνει το θέμα όταν αλλάζει η ειδικότητα."""
        spec = self._spec_var.get()
        self._subj_var.set(self._saved_subject.replace('{specialty}', spec))

    def _load_specialties(self):
        """Φορτώνει τις ειδικότητες από το Topothetiseis αρχείο."""
        if not self._topoth_path:
            self._spec_lbl.config(text='Δεν βρέθηκε αρχείο Τοποθετήσεων.', fg='#CC0000')
            return
        try:
            import pandas as pd
            df = pd.read_excel(self._topoth_path, header=0)
            spec_col = self._fc(df, 'κλάδ', 'ειδικ') or df.columns[4]
            self._topoth_spec_col = spec_col
            specialties = sorted(df[spec_col].dropna().astype(str).unique())
            self._spec_combo.config(values=specialties)
            if specialties:
                self._spec_var.set(specialties[0])
                # Ενημέρωσε το body με την πρώτη ειδικότητα
                if hasattr(self, '_body_txt'):
                    from datetime import datetime as _dt
                    self._body_txt.delete('1.0', 'end')
                    self._body_txt.insert('1.0',
                        self._saved_body
                            .replace('{date}', _dt.today().strftime('%d/%m/%Y'))
                            .replace('{specialty}', specialties[0]))
            self._spec_lbl.config(
                text=f'{len(specialties)} ειδικότητες | {os.path.basename(self._topoth_path)}',
                fg=C['desc'])
        except Exception as e:
            self._spec_lbl.config(text=f'Σφάλμα φόρτωσης: {e}', fg='#CC0000')

    # ── Βοηθητικά ───────────────────────────────────────────────────────────

    @staticmethod
    def _fc(df, *keywords):
        """Επιστρέφει το όνομα της πρώτης στήλης που ταιριάζει με κάποια λέξη-κλειδί."""
        for kw in keywords:
            kw = kw.lower()
            for col in df.columns:
                if kw in str(col).lower():
                    return col
        return None

    @staticmethod
    def _norm_code(series):
        """Κανονικοποίηση κωδικού σχολείου: αριθμός ή string → stripped string."""
        return series.fillna('').astype(str).str.strip().str.lstrip('0') \
                     .str.replace(r'\.0$', '', regex=True)

    @staticmethod
    def _clean_afm(val):
        """Καθαρισμός ΑΦΜ από CSV format =\"152159882\" → '152159882'."""
        return str(val).strip().strip('"').lstrip('=').strip('"').strip()

    # ── Εκτέλεση ────────────────────────────────────────────────────────────

    def _execute(self, send=True):
        import json, pandas as pd
        from datetime import datetime, date
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        specialty = self._spec_var.get()
        to_email  = self._to_var.get().strip()
        subject   = self._subj_var.get().strip()
        body_text = self._body_txt.get('1.0', 'end-1c')
        full_body = body_text + '\n\n' + config.email_signature()

        if send and not to_email:
            messagebox.showwarning('Email', 'Εισάγετε email παραλήπτη.', parent=self)
            return

        if not specialty:
            messagebox.showwarning('Ειδικότητα', 'Επίλεξε ειδικότητα.', parent=self)
            return
        if not self._topoth_path or not self._grid_path:
            messagebox.showwarning('Αρχεία', 'Δεν βρέθηκαν τα αρχεία Τοποθετήσεων / Καταλόγου.\nΚατέβασέ τα πρώτα από «Λήψη Δεδομένων».', parent=self)
            return

        try:
            # ── 1. Τοποθετήσεις ──────────────────────────────────────────────
            df_t = pd.read_excel(self._topoth_path, header=0)

            # Στήλες στο Topothetiseis (βάσει dump: col4=ειδικ, col5=σχέση εργ,
            #   col6=σχέση τοποθ, col7=κωδικός, col8=φορέας, col16=έως)
            spec_col      = self._topoth_spec_col
            code_col      = self._fc(df_t, 'κωδικ')        or df_t.columns[7]
            eos_col       = self._fc(df_t, 'έως', 'εως')   or df_t.columns[16]
            eponym_col    = self._fc(df_t, 'επώνυμ')       or df_t.columns[2]
            org_col       = self._fc(df_t, 'σχέση εργ', 'οργαν') or df_t.columns[5]
            topoth_col    = self._fc(df_t, 'σχέση τοποθ')  or df_t.columns[6]
            school_name_col = self._fc(df_t, 'φορέας τοποθ', 'φορέας') or df_t.columns[8]

            # ΑΦΜ (Α.Φ.Μ.) — join key, υπάρχει σε όλους (μόνιμοι + αναπληρωτές)
            # Α.Φ.Μ. (ΑΦΜ) — join key για stat4_1/4_2 (9 ψηφία)
            afm_col = self._fc(df_t, 'α.φ.μ', 'αφμ') or df_t.columns[1]
            # Α.Μ. — join key για stat4_16 (6 ψηφία) + εμφάνιση στην έξοδο
            am_col = None
            for col in df_t.columns:
                c = str(col).lower().strip()
                if 'α.μ' in c and 'φ' not in c:
                    am_col = col; break
            if am_col is None:
                am_col = df_t.columns[0]

            # Βρες Όνομα εκπ/κού — εξαιρούμε "Ονομασία σχολείου"
            onoma_col = None
            for col in df_t.columns:
                c = str(col).lower()
                if ('όνομ' in c or 'ονομ' in c) and 'ονομασ' not in c and 'σχολ' not in c:
                    onoma_col = col; break

            # ── Φίλτρα καθαρισμού Τοποθετήσεων ──────────────────────────────
            # 1. Αφαίρεση εγγραφών με Κατάσταση = ΠΑΡΗΛΘΕ
            status_col = self._fc(df_t, 'κατάσταση', 'κατασταση') or df_t.columns[17]
            df_t = df_t[df_t[status_col].fillna('').astype(str).str.strip() != 'ΠΑΡΗΛΘΕ'].copy()

            # 2. Αφαίρεση συγκεκριμένων τύπων σχέσης εργασίας
            # Χρήση contains γιατί οι τιμές μπορεί να έχουν παρενθέσεις
            # π.χ. "Ιδιωτικού Δικαίου Αορίστου Χρόνου (Ι.Δ.Α.Χ.)"
            _EXCLUDE_ORG_PAT = (
                r'Με άδεια διδασκαλίας για Ξένο Σχολείο'
                r'|Αναπληρωτής Ιδιωτικής Εκπαίδευσης'
                r'|Ιδιωτικού Δικαίου Αορίστου Χρόνου'
            )
            df_t = df_t[~df_t[org_col].fillna('').astype(str).str.strip()
                        .str.contains(_EXCLUDE_ORG_PAT, regex=True, na=False)].copy()

            # 3. Κράτησε μόνο Περιοχή Μετάθεσης Φορέα = Α΄ ΘΕΣΣΑΛΟΝΙΚΗΣ (Π.Ε.)
            area_mt_col = self._fc(df_t, 'περιοχή μετάθεσης φορέα', 'μετάθεσης φορέα') \
                          or df_t.columns[19]
            df_t = df_t[
                df_t[area_mt_col].fillna('').astype(str)
                    .str.contains(r'Α.{0,2}\s*ΘΕΣΣΑΛΟΝΙΚΗΣ.*Π\.Ε', regex=True, na=False)
            ].copy()

            # 4. Αφαίρεση συγκεκριμένων τύπων σχέσης τοποθέτησης
            # Χρήση contains γιατί οι τιμές μπορεί να έχουν παρενθέσεις
            # π.χ. "Μερική Διάθεση (αναπληρωτές εκπαιδευτικοί)"
            _EXCLUDE_TOPOTH_PAT = r'Υπερωριακά|Μερική Διάθεση|Τοποθέτηση Διοικητικού'
            df_t = df_t[~df_t[topoth_col].fillna('').astype(str).str.strip()
                        .str.contains(_EXCLUDE_TOPOTH_PAT, regex=True, na=False)].copy()


            # Κανονικοποίηση κωδικού, ΑΦΜ και Α.Μ.
            df_t['_code'] = self._norm_code(df_t[code_col])
            df_t['_afm']  = df_t[afm_col].fillna('').astype(str).str.strip() \
                                         .str.replace(r'\.0$', '', regex=True) \
                                         .str.zfill(9)
            df_t['_am']   = df_t[am_col].fillna('').astype(str).str.strip() \
                                        .str.replace(r'\.0$', '', regex=True)

            # ── 2. gridResults ────────────────────────────────────────────────
            df_g = pd.read_excel(self._grid_path, header=0)

            gc_code  = self._fc(df_g, 'κωδικός', 'κωδ')    or df_g.columns[11]
            gc_name  = self._fc(df_g, 'ονομασ')             or df_g.columns[1]
            gc_phone = self._fc(df_g, 'τηλ')                or df_g.columns[15]
            gc_email = self._fc(df_g, 'e-mail', 'email')    or df_g.columns[17]
            gc_area  = self._fc(df_g, 'περιοχ', 'τοποθεσ') or df_g.columns[18]

            # Φίλτρα gridResults
            gc_eidos = self._fc(df_g, 'είδος', 'ειδος')
            if gc_eidos:
                df_g = df_g[df_g[gc_eidos].fillna('').astype(str).str.strip() != 'Ιδιωτικά Σχολεία'].copy()

            df_g['_code'] = self._norm_code(df_g[gc_code])
            df_g_lu = df_g[['_code', gc_name, gc_phone, gc_email, gc_area]] \
                          .drop_duplicates('_code').copy()
            df_g_lu.columns = ['_code', '_school_name', '_phone', '_school_email', '_area']
            df_g_lu['_phone'] = df_g_lu['_phone'].fillna('').astype(str) \
                                    .str.replace(r'\.0$', '', regex=True).str.strip()

            valid_codes = set(df_g_lu['_code'])

            # ── 3. Φιλτράρισμα: μόνο σχολεία Δ/νσης + ειδικότητα ───────────
            df_t = df_t[df_t['_code'].isin(valid_codes)].copy()
            df_t = df_t[df_t[spec_col].astype(str) == specialty].copy()

            if df_t.empty:
                messagebox.showwarning('Αποτέλεσμα',
                    f'Δεν βρέθηκαν εκπαιδευτικοί ειδικότητας "{specialty}".', parent=self)
                return

            # ── 4. stat4_16 (απόντες — αιτιολόγηση απουσίας) ────────────────
            def _read_csv_enc(path):
                if not path: return pd.DataFrame()
                import zipfile as _zf, io as _io
                if path.endswith('.zip'):
                    try:
                        with _zf.ZipFile(path) as z:
                            csvname = [n for n in z.namelist() if n.endswith('.csv')][0]
                            data = z.read(csvname)
                    except Exception:
                        return pd.DataFrame()
                    for enc in ['utf-8-sig', 'utf-8', 'iso-8859-7', 'cp1253']:
                        try:
                            return pd.read_csv(_io.BytesIO(data), sep=None, engine='python',
                                               encoding=enc, header=0, dtype=str)
                        except Exception:
                            continue
                    return pd.DataFrame()
                for enc in ['utf-8-sig', 'utf-8', 'iso-8859-7', 'cp1253']:
                    try:
                        return pd.read_csv(path, sep=None, engine='python',
                                           encoding=enc, header=0, dtype=str)
                    except Exception:
                        continue
                return pd.DataFrame()

            df_s16 = _read_csv_enc(self._stat_path)
            if not df_s16.empty:
                # stat4_16: col16 labeled "Α.Μ." αλλά έχει ΑΦΜ δεδομένα (9 ψηφία, ="..." format)
                # → join με ΑΦΜ (ίδιο key με Topothetiseis col1 και stat4_1/4_2 col0)
                s16_afm_col = None
                for col in df_s16.columns:
                    c = str(col).lower().strip()
                    if 'α.μ' in c and 'φ' not in c:
                        s16_afm_col = col; break
                if s16_afm_col is None:
                    s16_afm_col = df_s16.columns[16]
                # col shift: header[i] περιγράφει data[i-1]
                # header[45]='Αιτιολόγηση Απουσίας' → data[44]
                # header[48]='Έως'                  → data[47]
                s16_abs_col = df_s16.columns[44] if len(df_s16.columns) > 44 else df_s16.columns[45]
                s16_eos_col = df_s16.columns[47] if len(df_s16.columns) > 47 else None

                df_s16['_afm'] = df_s16[s16_afm_col].apply(self._clean_afm).str.zfill(9)
                keep16 = ['_afm', s16_abs_col]
                if s16_eos_col: keep16.append(s16_eos_col)
                df_s16_lu = df_s16[keep16].drop_duplicates('_afm').copy()
                rename16 = {'_afm': '_afm', s16_abs_col: '_apoysia'}
                if s16_eos_col: rename16[s16_eos_col] = '_eos'
                df_s16_lu = df_s16_lu.rename(columns=rename16)
                if '_eos' not in df_s16_lu: df_s16_lu['_eos'] = ''
                absent_afms = set(df_s16_lu['_afm'])
            else:
                df_s16_lu = pd.DataFrame(columns=['_afm', '_apoysia', '_eos'])
                absent_afms = set()

            # ── 5. stat4_1 & stat4_2 (Email ΠΣΔ, Κινητό) — join με ΑΦΜ ─────
            # stat4_1 col0 labeled "Α.Μ." αλλά έχει ΑΦΜ δεδομένα (9 ψηφία, ="..." format)
            frames_41_42 = []
            for path in [self._stat41_path, self._stat42_path]:
                df_tmp = _read_csv_enc(path)
                if not df_tmp.empty:
                    frames_41_42.append(df_tmp)

            if frames_41_42:
                df_41_42 = pd.concat(frames_41_42, ignore_index=True)
                # stat4_1/4_2 έχουν 1-column shift στα headers.
                # col0  (1-based:  1) → ΑΦΜ data (="..." format)
                # col9  (1-based: 10) → Κινητό data
                # col11 (1-based: 12) → Email προσωπικό data
                # col12 (1-based: 13) → Email ΠΣΔ (sch.gr) data
                s41_afm_col      = df_41_42.columns[0]
                s41_psd_col      = df_41_42.columns[12] if len(df_41_42.columns) > 12 else None
                s41_email_col    = df_41_42.columns[11] if len(df_41_42.columns) > 11 else None
                s41_mobile_col   = df_41_42.columns[9]  if len(df_41_42.columns) > 9  else None

                df_41_42['_afm'] = df_41_42[s41_afm_col].apply(self._clean_afm).str.zfill(9)
                keep = ['_afm']
                if s41_psd_col:    keep.append(s41_psd_col)
                if s41_email_col:  keep.append(s41_email_col)
                if s41_mobile_col: keep.append(s41_mobile_col)
                df_41_lu = df_41_42[keep].drop_duplicates('_afm').copy()
                rename = {}
                if s41_psd_col:    rename[s41_psd_col]    = '_email_psd'
                if s41_email_col:  rename[s41_email_col]  = '_email_personal'
                if s41_mobile_col: rename[s41_mobile_col] = '_kinito'
                df_41_lu = df_41_lu.rename(columns=rename)
                if '_email_psd'      not in df_41_lu: df_41_lu['_email_psd']      = ''
                if '_email_personal' not in df_41_lu: df_41_lu['_email_personal'] = ''
                if '_kinito'         not in df_41_lu: df_41_lu['_kinito']         = ''
            else:
                df_41_lu = pd.DataFrame(columns=['_afm', '_email_psd', '_email_personal', '_kinito'])

            # ── 6. Join ───────────────────────────────────────────────────────
            df_t = df_t.merge(df_g_lu,   on='_code', how='left')
            df_t = df_t.merge(df_s16_lu, on='_afm',  how='left')   # ΑΦΜ join (stat4_16)
            df_t = df_t.merge(df_41_lu,  on='_afm',  how='left')   # ΑΦΜ join (stat4_1/4_2)
            df_t['_absent'] = df_t[status_col].fillna('').astype(str).str.strip() == 'ΑΠΟΥΣΙΑ'

            # ── 7. Χτίσε dataframe εξόδου ─────────────────────────────────────
            def gcol(col):
                if col is not None and col in df_t.columns:
                    return df_t[col].fillna('').astype(str)
                return pd.Series([''] * len(df_t), index=df_t.index)

            out = pd.DataFrame(index=df_t.index)
            out['ΑΜ']                = df_t['_am'].fillna('')
            out['Επώνυμο']            = gcol(eponym_col)
            out['Όνομα']              = gcol(onoma_col)
            out['Κύρια Ειδικ.']      = gcol(spec_col)
            out['Email στο ΠΣΔ']     = df_t['_email_psd'].fillna('')
            out['Email']             = df_t['_email_personal'].fillna('') \
                                        if '_email_personal' in df_t.columns else ''
            out['Κινητό']            = df_t['_kinito'].fillna('')
            out['Σχέση εργασίας']    = gcol(org_col)
            out['Σχέση τοποθέτησης'] = gcol(topoth_col)
            out['Κατάσταση']         = gcol(status_col)
            out['Φορέας τοποθέτησης']= gcol(school_name_col)
            out['Τηλέφωνο']          = df_t['_phone'].fillna('')
            out['e-mail']            = df_t['_school_email'].fillna('')
            out['ΑΠΟΥΣΙΑ']           = df_t['_apoysia'].fillna('')
            out['Έως']               = df_t['_eos'].fillna('') if '_eos' in df_t.columns else ''
            out['_absent']           = df_t['_absent']
            # Αιτιολόγηση + ημ/νία επιστροφής μόνο για απόντες
            out.loc[~out['_absent'], 'ΑΠΟΥΣΙΑ'] = ''
            out.loc[~out['_absent'], 'Έως']     = ''

            out = out.sort_values('Επώνυμο', na_position='last').reset_index(drop=True)

            # ── 7. Δημιουργία Excel ──────────────────────────────────────────
            today_str = datetime.today().strftime('%Y%m%d')
            out_dir   = os.path.join(_docs_base(), f'results_{today_str}')
            os.makedirs(out_dir, exist_ok=True)
            spec_safe = specialty.replace('/', '_').replace('\\', '_')
            out_path  = os.path.join(out_dir, f'Εκπαιδευτικοί_{spec_safe}_{today_str}.xlsx')

            # Φιλτράρισμα στηλών βάσει επιλογής χρήστη
            _disabled = {c for c, v in getattr(self, '_col_vars', {}).items() if not v.get()}
            active_cols = [c for c in self._OUT_COLS if c not in _disabled]

            wb = Workbook()
            ws = wb.active
            ws.title = specialty[:31]

            RED = 'FF0000'
            thin   = Side(style='thin', color='CCCCCC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row_align = Alignment(horizontal='left',   vertical='center')

            # Επικεφαλίδες — κόκκινο φόντο, λευκό bold
            for ci, col in enumerate(active_cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
                cell.fill      = PatternFill('solid', start_color=RED)
                cell.alignment = hdr_align
                cell.border    = border

            # Δεδομένα
            alt_fill = PatternFill('solid', start_color='FFF0F0')
            for ri, row in out.iterrows():
                is_absent = bool(row.get('_absent', False))
                for ci, col in enumerate(active_cols, 1):
                    val = row.get(col, '')
                    if pd.isna(val):
                        val = ''
                    cell = ws.cell(row=ri + 2, column=ci, value=str(val) if val != '' else '')
                    if is_absent:
                        cell.font = Font(name='Arial', size=9, color=RED, bold=True)
                    else:
                        cell.font = Font(name='Arial', size=9, color='000000')
                        if ri % 2 == 1:
                            cell.fill = alt_fill
                    cell.alignment = row_align
                    cell.border    = border

            # Πλάτη στηλών
            for ci, col in enumerate(active_cols, 1):
                vals = [str(out.iloc[r][col]) for r in range(min(len(out), 50))
                        if col in out.columns and not pd.isna(out.iloc[r][col])]
                w = max([len(col)] + [len(v) for v in vals]) if vals else len(col)
                ws.column_dimensions[get_column_letter(ci)].width = min(w + 3, 42)

            ws.row_dimensions[1].height = 30
            ws.freeze_panes = 'A2'
            wb.save(out_path)

        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror('Σφάλμα', str(e), parent=self)
            return

        # ── 8. Αποθήκευση ρυθμίσεων ──────────────────────────────────────────
        s = _load_local_settings()
        s[self._SETTINGS_KEY] = {
            'subject':       self._saved_subject,
            'body':          self._saved_body,
            'advisor_email': to_email,
        }
        path_s = _get_local_settings_path()
        os.makedirs(os.path.dirname(path_s), exist_ok=True)
        with open(path_s, 'w', encoding='utf-8') as f:
            json.dump(s, f, ensure_ascii=False, indent=2)

        absent_count = int(out['_absent'].sum())
        total_count  = len(out)

        if not send:
            messagebox.showinfo('Έτοιμο',
                f'Αρχείο αποθηκεύτηκε:\n{out_path}\n\n'
                f'Σύνολο: {total_count} εκπ/κοί  |  Απόντες (κόκκινο): {absent_count}',
                parent=self)
            try:
                import subprocess; subprocess.Popen(['explorer', out_dir])
            except Exception:
                pass
            self.destroy()
            return

        # Αποστολή email
        try:
            from core.framework import send_email
            send_email(config, to_email, subject, full_body, out_path)
            messagebox.showinfo('Αποστολή',
                f'Email στάλθηκε: {to_email}\n\n'
                f'Αρχείο: {out_path}\n'
                f'Σύνολο: {total_count} εκπ/κοί  |  Απόντες: {absent_count}',
                parent=self)
            try:
                import subprocess; subprocess.Popen(['explorer', out_dir])
            except Exception:
                pass
            self.destroy()
        except Exception as e:
            messagebox.showerror('Σφάλμα αποστολής', str(e), parent=self)

    # ── Βοηθητικά ───────────────────────────────────────────────────────────

    def _clear(self):
        for w in self.winfo_children():
            w.destroy()


class MonadaDialog(tk.Toplevel):
    """Εργαλείο εξαγωγής στοιχείων σχολικών μονάδων ανά Δήμο."""

    _SETTINGS_KEY   = 'monada_tool'
    _DEFAULT_BODY   = (
        'Αποτύπωση Myschool {date}.\n\n'
        'Καλημέρα σας,\n\n'
        'Επισυνάπτω πίνακα excel με τα στοιχεία των σχολικών μονάδων '
        'Δήμου {dimos} σύμφωνα με τα καταχωρημένα στοιχεία στο myschool.\n\n\n'
        'Στη διάθεσή σας για οποιαδήποτε πληροφορία'
    )
    _DEFAULT_SUBJECT = 'Στοιχεία σχολικών μονάδων Δήμου {dimos}'

    # Ταξινόμηση τάξεων (Νηπιαγωγείο + Δημοτικό)
    _CLASS_ORDER = ['ΠΡΟΝΗΠΙΑ', 'ΝΗΠΙΑ', 'ΠΡΟΝΗΠΙΑ-ΝΗΠΙΑ', 'Α', 'Β', 'Γ', 'Δ', 'Ε', 'ΣΤ']

    def __init__(self, parent):
        super().__init__(parent)
        self.title('Στοιχεία Σχολικών Μονάδων')
        self.configure(bg=C['bg'])
        self.resizable(True, True)
        self.grab_set()
        self.transient(parent)
        self._parent = parent

        s = _load_local_settings().get(self._SETTINGS_KEY, {})
        self._saved_subject = s.get('subject',     self._DEFAULT_SUBJECT)
        self._saved_body    = s.get('body',         self._DEFAULT_BODY)
        self._saved_email   = s.get('dimos_email',  '')

        # Αυτόματη εύρεση αρχείων zip
        # CSV_* = χειροκίνητο download, stat2_2* = μέσω app downloader (2.2), gridResults* = 2.1 fallback
        self._csv_path    = self._auto_find_zip('CSV_', 'stat2_2', 'gridResults')
        self._stat31_path = self._auto_find_zip('stat3_1')

        self._build_form()
        self.update_idletasks()
        w, h = 630, 540
        x = parent.winfo_x() + (parent.winfo_width()  - w) // 2
        y = parent.winfo_y() + (parent.winfo_height() - h) // 2
        self.geometry(f'{w}x{h}+{x}+{y}')

    # ── Auto-find ────────────────────────────────────────────────────────────

    @staticmethod
    def _auto_find_zip(*prefixes):
        """Ψάχνει αρχείο (zip/csv/xlsx) με δοθέν prefix — downloads app πρώτα, μετά ~/Downloads.
        Δέχεται πολλαπλά prefixes (πρώτο εύρημα κερδίζει).
        """
        import glob as _glob
        folders = []
        dl_base = os.path.join(_docs_base(), 'downloads')
        if os.path.isdir(dl_base):
            folders += sorted([
                os.path.join(dl_base, d)
                for d in os.listdir(dl_base)
                if os.path.isdir(os.path.join(dl_base, d))
            ], reverse=True)
        folders.append(os.path.join(os.path.expanduser('~'), 'Downloads'))
        for folder in folders:
            for prefix in prefixes:
                # .zip πρώτα (χειροκίνητο κατέβασμα), μετά .csv/.xlsx (μέσω app downloader)
                for pattern in (f'{prefix}*.zip', f'{prefix}*.csv', f'{prefix}*.xlsx'):
                    matches = [f for f in _glob.glob(os.path.join(folder, pattern))
                               if not f.endswith('.tmp') and not f.endswith('.crdownload')]
                    if matches:
                        return sorted(matches)[-1]
        return ''

    # ── Βοηθητικά ────────────────────────────────────────────────────────────

    @staticmethod
    def _clean_code(val):
        """Αφαιρεί =\"XXXXX\" format και επιστρέφει τα ψηφία (lstrip 0)."""
        import re
        s = str(val).strip().strip('"').lstrip('=').strip('"').strip()
        s = re.sub(r'\.0$', '', s)
        return s.lstrip('0') or s  # lstrip('0') αλλά όχι αν το αποτέλεσμα είναι κενό

    @staticmethod
    def _s(val):
        """Επιστρέφει string από τιμή — NaN/nan → κενό, .0 stripped."""
        import re
        if val is None: return ''
        s = str(val).strip()
        if s.lower() in ('nan', 'none', ''): return ''
        return re.sub(r'\.0$', '', s)

    @staticmethod
    def _read_zip_csv(path, encoding='cp1253', strip_trailing_sep=False):
        """Διαβάζει CSV/XLSX — από zip, plain CSV, ή xlsx.
        strip_trailing_sep: True για αρχεία με trailing ';' (π.χ. stat3_1).
        """
        import zipfile, io, pandas as pd
        lower = path.lower()
        if lower.endswith('.xlsx'):
            return pd.read_excel(path, dtype=str)
        if lower.endswith('.zip'):
            with zipfile.ZipFile(path) as z:
                raw = z.read(z.namelist()[0])
        else:
            # Plain CSV (κατεβασμένο μέσω app downloader)
            with open(path, 'rb') as f:
                raw = f.read()
        text = raw.decode(encoding)
        if strip_trailing_sep:
            text = '\n'.join(l.rstrip(';') for l in text.splitlines())
        return pd.read_csv(io.StringIO(text), sep=';', dtype=str)

    # ── Κύρια φόρμα ──────────────────────────────────────────────────────────

    def _build_form(self):
        self._clear()

        tk.Label(self, text='Στοιχεία Σχολικών Μονάδων',
                 bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 11, 'bold')).pack(anchor='w', padx=18, pady=(14, 4))

        # Προειδοποίηση αν λείπουν αρχεία
        missing = []
        if not self._csv_path:    missing.append('Κατάλογος Μονάδων (CSV_...zip)')
        if not self._stat31_path: missing.append('Στατιστικό 3.1 (stat3_1...zip)')
        if missing:
            tk.Label(self,
                text=f'⚠  Δεν βρέθηκαν: {", ".join(missing)}. Κατέβασέ τα από MySchool.',
                bg='#FFF3E0', fg='#E65100', font=('Arial', 8),
                anchor='w', padx=10, pady=5, wraplength=570, justify='left',
            ).pack(fill='x', padx=18, pady=(0, 6))

        # ── Δήμος ─────────────────────────────────────────────────────────────
        tk.Label(self, text='Δήμος:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))

        dimos_row = tk.Frame(self, bg=C['bg'])
        dimos_row.pack(fill='x', padx=18, pady=(2, 6))
        self._dimos_var = tk.StringVar()
        from tkinter import ttk as _ttk
        self._dimos_combo = _ttk.Combobox(dimos_row, textvariable=self._dimos_var,
                                           width=40, state='readonly')
        self._dimos_combo.pack(side='left')
        self._dimos_lbl = tk.Label(dimos_row, text='Φόρτωση…',
                                    bg=C['bg'], fg=C['desc'], font=('Arial', 8))
        self._dimos_lbl.pack(side='left', padx=(10, 0))
        self._dimos_var.trace_add('write', self._on_dimos_change)

        # ── Εμφάνιση ──────────────────────────────────────────────────────────
        tk.Label(self, text='Εμφάνιση:', bg=C['bg'], fg=C['hdr_bg'],
                 font=('Arial', 9, 'bold'), anchor='w').pack(fill='x', padx=18, pady=(4, 0))
        mode_row = tk.Frame(self, bg=C['bg'])
        mode_row.pack(fill='x', padx=18, pady=(2, 6))
        self._mode_var = tk.StringVar(value='monada')
        tk.Radiobutton(mode_row, text='Ανά Σχολική Μονάδα', variable=self._mode_var,
                       value='monada', bg=C['bg'], font=('Arial', 9),
                       activebackground=C['bg']).pack(side='left', padx=(0, 18))
        tk.Radiobutton(mode_row, text='Ανά Τάξη', variable=self._mode_var,
                       value='taxh', bg=C['bg'], font=('Arial', 9),
                       activebackground=C['bg']).pack(side='left')

        # ── Email ─────────────────────────────────────────────────────────────
        pad = dict(padx=18, pady=2)
        tk.Label(self, text='Προς (email δήμου):',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._to_var = tk.StringVar(value=self._saved_email)
        tk.Entry(self, textvariable=self._to_var,
                 font=('Arial', 9)).pack(fill='x', padx=18, pady=(0, 6))

        tk.Label(self, text='Θέμα:',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._subj_var = tk.StringVar(value=self._saved_subject)
        tk.Entry(self, textvariable=self._subj_var,
                 font=('Arial', 9)).pack(fill='x', padx=18, pady=(0, 6))

        tk.Label(self, text='Κείμενο email:',
                 bg=C['bg'], fg=C['hdr_bg'], font=('Arial', 9, 'bold'),
                 anchor='w').pack(fill='x', **pad)
        self._body_txt = tk.Text(self, font=('Arial', 9), height=5,
                                  wrap='word', relief='solid', bd=1)
        self._body_txt.pack(fill='x', padx=18, pady=(0, 6))
        from datetime import datetime as _dt
        self._body_txt.insert('1.0',
            self._saved_body.replace('{date}', _dt.today().strftime('%d/%m/%Y'))
                             .replace('{dimos}', self._dimos_var.get()))

        btn_row = tk.Frame(self, bg=C['bg'])
        btn_row.pack(side='bottom', pady=10)
        tk.Button(btn_row, text='Μόνο Excel (χωρίς email)',
                  bg=C['bg2'], fg=C['hdr_bg'], relief='flat',
                  font=('Arial', 9), padx=10, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=False)).pack(side='left', padx=4)
        tk.Button(btn_row, text='▶  Δημιουργία & Αποστολή',
                  bg=C['btn_bg'], fg=C['btn_fg'], relief='flat',
                  font=('Arial', 9, 'bold'), padx=14, pady=5, cursor='hand2',
                  command=lambda: self._execute(send=True)).pack(side='left', padx=4)

        self.after(100, self._load_dimos)

    def _on_dimos_change(self, *_):
        dimos = self._dimos_var.get()
        self._subj_var.set(self._saved_subject.replace('{dimos}', dimos))
        # Ενημέρωση body: αντικατάσταση παλιού δήμου με νέο
        from datetime import datetime as _dt
        today = _dt.today().strftime('%d/%m/%Y')
        self._body_txt.delete('1.0', 'end')
        self._body_txt.insert('1.0',
            self._saved_body.replace('{date}', today).replace('{dimos}', dimos))

    def _load_dimos(self):
        """Φορτώνει τους Δήμους από το stat3_1 zip (col7) ή fallback CSV (col6)."""
        src_path = self._stat31_path or self._csv_path
        if not src_path:
            self._dimos_lbl.config(text='Δεν βρέθηκε αρχείο.', fg='#CC0000')
            return
        try:
            import pandas as pd
            use_31 = bool(self._stat31_path)
            df = self._read_zip_csv(src_path, strip_trailing_sep=use_31)
            col_idx   = 7 if use_31 else 6
            dimos_col = df.columns[col_idx]
            dimos_list = sorted(df[dimos_col].dropna().astype(str).str.strip().unique())
            self._dimos_combo.config(values=dimos_list)
            if dimos_list:
                self._dimos_var.set(dimos_list[0])
            self._dimos_lbl.config(
                text=f'{len(dimos_list)} δήμοι | {os.path.basename(src_path)}',
                fg=C['desc'])
        except Exception as e:
            self._dimos_lbl.config(text=f'Σφάλμα: {e}', fg='#CC0000')

    # ── Εκτέλεση ─────────────────────────────────────────────────────────────

    def _execute(self, send=True):
        import json, pandas as pd
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        dimos     = self._dimos_var.get().strip()
        mode      = self._mode_var.get()
        to_email  = self._to_var.get().strip()
        subject   = self._subj_var.get().strip()
        body_text = self._body_txt.get('1.0', 'end-1c')
        full_body = body_text + '\n\n' + config.email_signature()

        if send and not to_email:
            messagebox.showwarning('Email', 'Εισάγετε email παραλήπτη.', parent=self)
            return
        if not dimos:
            messagebox.showwarning('Δήμος', 'Επίλεξε Δήμο.', parent=self)
            return
        if not self._stat31_path:
            messagebox.showwarning('Αρχεία',
                'Δεν βρέθηκε το αρχείο στατιστικού 3.1.\n'
                'Κατέβασέ το από MySchool (Στατιστικά 3.1).', parent=self)
            return
        if not self._csv_path:
            messagebox.showwarning('Αρχεία',
                'Δεν βρέθηκε το αρχείο CSV σχολικών μονάδων.\n'
                'Κατέβασέ το από MySchool (Στατιστικά → Κατάλογος Μονάδων).', parent=self)
            return

        try:
            # ── 1. CSV → lookup dict (δευτερεύουσα πηγή, στοιχεία επικοινωνίας) ──
            # Επιβεβαιωμένα offsets λόγω 1-column shift στα headers από col11:
            #   col10 = Είδος (τύπος σχολείου)
            #   col11 = Κωδ. ΥΠΠΘ (αριθμητικός κωδικός)
            #   col12 = Ονομασία
            #   col16 = Τηλέφωνο
            #   col18 = e-mail σχολείου
            #   col20 = Ταχ. Διεύθυνση
            #   col48 = Αναστολή (NAI/OXI)
            #   col55 = Ονομ/μο Διευθυντή
            #   col58 = Κινητό Διευθυντή
            #   col59 = Email Διευθυντή
            #   col60 = Email ΠΣΔ Διευθυντή
            csv_df = self._read_zip_csv(self._csv_path)

            c_eidos    = csv_df.columns[10]
            c_code_csv = csv_df.columns[11]
            c_onoma    = csv_df.columns[12]
            c_phone    = csv_df.columns[16]
            c_email    = csv_df.columns[18]
            c_address  = csv_df.columns[20]
            c_anast    = csv_df.columns[48]   # Αναστολή (NAI = κλειστό)
            c_dir_name = csv_df.columns[55]
            c_dir_mob  = csv_df.columns[58]
            c_dir_mail = csv_df.columns[59]
            c_dir_psd  = csv_df.columns[60]

            # Φίλτρο τύπου: Δημοτικά + Νηπιαγωγεία, όχι Ιδιωτικά / Ξένα
            eidos_ser = csv_df[c_eidos].fillna('').astype(str)
            mask_type = (
                (eidos_ser.str.contains('Δημοτικό',    na=False) |
                 eidos_ser.str.contains('Νηπιαγωγείο', na=False)) &
                ~eidos_ser.str.contains('Ιδιωτικό', na=False) &
                ~eidos_ser.str.contains('Ξένο',     na=False)
            )
            csv_df = csv_df[mask_type].copy()

            # Φίλτρο Αναστολής: εξαίρεση σχολείων με Αναστολή = NAI
            csv_df = csv_df[
                csv_df[c_anast].fillna('').astype(str).str.strip().str.upper() != 'NAI'
            ].copy()

            # Κατασκευή lookup dict: {clean_code → στοιχεία}
            csv_df['_code'] = csv_df[c_code_csv].apply(self._clean_code)
            csv_lookup = {}
            for _, row in csv_df.iterrows():
                code = row['_code']
                if not code:
                    continue
                eidos_val   = self._s(row[c_eidos])
                is_dim      = 'Δημοτικό' in eidos_val
                eidos_short = (eidos_val
                               .replace('Ενιαίου Τύπου Ολοήμερο ', '')
                               .replace('Ολοήμερο ', '')
                               .strip())
                csv_lookup[code] = {
                    'eidos':    eidos_short,
                    'is_dim':   is_dim,
                    'onoma':    self._s(row[c_onoma]),
                    'phone':    self._s(row[c_phone]),
                    'email':    self._s(row[c_email]),
                    'address':  self._s(row[c_address]),
                    'dir_name': self._s(row[c_dir_name]),
                    'dir_mob':  self._s(row[c_dir_mob]),
                    'dir_mail': self._s(row[c_dir_mail]),
                    'dir_psd':  self._s(row[c_dir_psd]),
                }

            # ── 2. stat3_1 — ΚΥΡΙΑ πηγή (κατανομή ανά τάξη & φύλο) ──────────
            df31 = self._read_zip_csv(self._stat31_path, strip_trailing_sep=True)

            kwd31  = df31.columns[4]   # Κωδικός σχολείου
            dim31  = df31.columns[7]   # Δήμος
            taxh31 = df31.columns[10]  # Τάξη
            tmhm31 = df31.columns[11]  # Αριθμός Τμημάτων
            ag31   = df31.columns[12]  # Αγόρια
            ko31   = df31.columns[13]  # Κορίτσια
            sy31   = df31.columns[14]  # Σύνολο

            df31['_code'] = df31[kwd31].apply(self._clean_code)

            # Φίλτρο Δήμου
            df31 = df31[df31[dim31].fillna('').astype(str).str.strip() == dimos].copy()

            # Κράτα μόνο σχολεία που υπάρχουν στο CSV lookup
            # (αποκλείονται αυτόματα Ιδιωτικά, Ξένα, Αναστολή)
            df31 = df31[df31['_code'].isin(csv_lookup)].copy()

            if df31.empty:
                messagebox.showwarning('Αποτέλεσμα',
                    f'Δεν βρέθηκαν σχολεία στον Δήμο "{dimos}".', parent=self)
                return

            # Αριθμητικές στήλες
            for col in [tmhm31, ag31, ko31, sy31]:
                df31[col] = pd.to_numeric(df31[col], errors='coerce').fillna(0).astype(int)

            # Sorted list κωδικών: Νηπιαγωγεία πρώτα, μετά Δημοτικά, αλφαβητικά
            unique_codes = df31['_code'].unique()
            sorted_codes = sorted(unique_codes,
                key=lambda c: (1 if csv_lookup.get(c, {}).get('is_dim') else 0,
                               csv_lookup.get(c, {}).get('onoma', '')))

            # ── 3. Δημιουργία Excel ───────────────────────────────────────────
            today_str  = datetime.today().strftime('%Y%m%d')
            out_dir    = os.path.join(_docs_base(), f'results_{today_str}')
            os.makedirs(out_dir, exist_ok=True)
            dimos_safe = dimos.replace('/', '_').replace('\\', '_')
            mode_sfx   = 'ανά_τάξη' if mode == 'taxh' else 'ανά_μονάδα'
            out_path   = os.path.join(
                out_dir, f'Σχολικές_Μονάδες_{dimos_safe}_{mode_sfx}_{today_str}.xlsx')

            wb  = Workbook()
            ws  = wb.active
            ws.title = dimos[:31]

            RED        = 'FF0000'
            LIGHT_BLUE = 'DCE6F1'
            LIGHT_RED  = 'FCE4EC'
            thin   = Side(style='thin', color='CCCCCC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_al = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ctr_al = Alignment(horizontal='center', vertical='center')
            lft_al = Alignment(horizontal='left',   vertical='center')

            def _hdr_cell(ws, row, col, value):
                c = ws.cell(row=row, column=col, value=value)
                c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
                c.fill      = PatternFill('solid', start_color=RED)
                c.alignment = hdr_al
                c.border    = border

            CLASS_RANK = {c: i for i, c in enumerate(self._CLASS_ORDER)}

            if mode == 'taxh':
                # ─── Ανά Τάξη ─────────────────────────────────────────────────
                all_cols = [
                    'Είδος', 'Ονομασία', 'Τάξη', 'Τμήματα',
                    'Αγόρια', 'Κορίτσια', 'Σύνολο',
                    'Τηλέφωνο', 'e-mail σχολείου',
                    'Ονομ/μο Διευθυντή', 'Κινητό Διευθυντή',
                    'Email Διευθυντή', 'Email ΠΣΔ Διευθυντή',
                ]

                for ci, col in enumerate(all_cols, 1):
                    _hdr_cell(ws, 1, ci, col)

                subtot_fill = PatternFill('solid', start_color=LIGHT_BLUE)
                grand_fill  = PatternFill('solid', start_color=LIGHT_RED)
                alt_fill    = PatternFill('solid', start_color='F7F7F7')
                er = 2
                tot_ag_g = tot_ko_g = tot_sy_g = 0

                for code in sorted_codes:
                    info   = csv_lookup.get(code, {})
                    is_dim = info.get('is_dim', False)
                    sc_df  = df31[df31['_code'] == code].copy()
                    if sc_df.empty:
                        continue

                    sc_df['_rank'] = sc_df[taxh31].apply(
                        lambda t: CLASS_RANK.get(str(t).strip(), 99))
                    sc_df = sc_df.sort_values('_rank').reset_index(drop=True)

                    onoma = info.get('onoma', '')
                    sc_ag = sc_ko = sc_sy = sc_tm = 0
                    for row_i, (_, crow) in enumerate(sc_df.iterrows()):
                        taxh = self._s(crow[taxh31])
                        tm   = int(crow[tmhm31] or 0)
                        ag   = int(crow[ag31]   or 0)
                        ko   = int(crow[ko31]   or 0)
                        sy   = int(crow[sy31]   or 0)
                        sc_ag += ag; sc_ko += ko; sc_sy += sy; sc_tm += tm

                        # Στοιχεία επικοινωνίας μόνο στην 1η γραμμή
                        first = row_i == 0
                        vals = [
                            info.get('eidos',    '') if first else '',
                            onoma                    if first else '',
                            taxh, tm, ag, ko, sy,
                            info.get('phone',    '') if first else '',
                            info.get('email',    '') if first else '',
                            info.get('dir_name', '') if first else '',
                            info.get('dir_mob',  '') if first else '',
                            info.get('dir_mail', '') if first else '',
                            info.get('dir_psd',  '') if first else '',
                        ]

                        row_fill = alt_fill if row_i % 2 == 1 else None
                        for ci, val in enumerate(vals, 1):
                            cell = ws.cell(row=er, column=ci, value=val)
                            cell.font      = Font(name='Arial', size=9)
                            cell.alignment = ctr_al if 4 <= ci <= 7 else lft_al
                            cell.border    = border
                            if row_fill: cell.fill = row_fill
                        er += 1

                    # Subtotal μόνο για Δημοτικά (τα Νηπιαγωγεία έχουν 1-2 γραμμές, δεν χρειάζεται)
                    tot_ag_g += sc_ag; tot_ko_g += sc_ko; tot_sy_g += sc_sy
                    if is_dim:
                        for ci in range(1, len(all_cols) + 1):
                            cell = ws.cell(row=er, column=ci)
                            cell.font      = Font(name='Arial', size=9, bold=True)
                            cell.fill      = subtot_fill
                            cell.border    = border
                            cell.alignment = ctr_al if 4 <= ci <= 7 else lft_al
                        ws.cell(row=er, column=2, value=f'Σύνολο {onoma}')
                        ws.cell(row=er, column=4, value=sc_tm)
                        ws.cell(row=er, column=5, value=sc_ag)
                        ws.cell(row=er, column=6, value=sc_ko)
                        ws.cell(row=er, column=7, value=sc_sy)
                        er += 1

                # Grand total
                for ci in range(1, len(all_cols) + 1):
                    cell = ws.cell(row=er, column=ci)
                    cell.font      = Font(name='Arial', size=9, bold=True)
                    cell.fill      = grand_fill
                    cell.border    = border
                    cell.alignment = ctr_al if 4 <= ci <= 7 else lft_al
                ws.cell(row=er, column=2, value='ΓΕΝΙΚΟ ΣΥΝΟΛΟ')
                ws.cell(row=er, column=5, value=tot_ag_g)
                ws.cell(row=er, column=6, value=tot_ko_g)
                ws.cell(row=er, column=7, value=tot_sy_g)

                col_widths = [26, 40, 20, 10, 10, 12, 10, 16, 30, 28, 18, 32, 26]
                for ci, w in enumerate(col_widths[:len(all_cols)], 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w

            else:
                # ─── Ανά Σχολική Μονάδα ──────────────────────────────────────
                # Ομαδοποίηση stat3_1 ανά σχολείο + επικοινωνία από CSV
                base_cols = [
                    'Είδος', 'Ονομασία',
                    'Τμήματα', 'Αγόρια', 'Κορίτσια', 'Σύνολο',
                    'Τηλέφωνο', 'e-mail σχολείου', 'Ταχ. Διεύθυνση',
                    'Ονομ/μο Διευθυντή',
                ]
                all_cols = base_cols + [
                    'Κινητό Διευθυντή', 'Email Διευθυντή', 'Email ΠΣΔ Διευθυντή',
                ]

                for ci, col in enumerate(all_cols, 1):
                    _hdr_cell(ws, 1, ci, col)

                alt_fill   = PatternFill('solid', start_color='FFF0F0')
                tot_ag = tot_ko = tot_sy = tot_tm = 0

                for ri, code in enumerate(sorted_codes):
                    info  = csv_lookup.get(code, {})
                    sc_df = df31[df31['_code'] == code]
                    tm = int(sc_df[tmhm31].sum())
                    ag = int(sc_df[ag31].sum())
                    ko = int(sc_df[ko31].sum())
                    sy = int(sc_df[sy31].sum())
                    tot_tm += tm; tot_ag += ag; tot_ko += ko; tot_sy += sy

                    vals = [
                        info.get('eidos',    ''),
                        info.get('onoma',    ''),
                        tm, ag, ko, sy,
                        info.get('phone',    ''),
                        info.get('email',    ''),
                        info.get('address',  ''),
                        info.get('dir_name', ''),
                        info.get('dir_mob',  ''),
                        info.get('dir_mail', ''),
                        info.get('dir_psd',  ''),
                    ]

                    fill = alt_fill if ri % 2 == 1 else None
                    er   = ri + 2
                    for ci, val in enumerate(vals, 1):
                        cell = ws.cell(row=er, column=ci, value=val)
                        cell.font      = Font(name='Arial', size=9)
                        cell.alignment = ctr_al if 3 <= ci <= 6 else lft_al
                        cell.border    = border
                        if fill: cell.fill = fill

                # Σειρά ΣΥΝΟΛΟ
                tot_row = len(sorted_codes) + 2
                for ci in range(1, len(all_cols) + 1):
                    cell = ws.cell(row=tot_row, column=ci)
                    cell.font      = Font(name='Arial', size=9, bold=True)
                    cell.fill      = PatternFill('solid', start_color=LIGHT_BLUE)
                    cell.border    = border
                    cell.alignment = ctr_al if 3 <= ci <= 6 else lft_al
                ws.cell(row=tot_row, column=2, value='ΣΥΝΟΛΟ')
                ws.cell(row=tot_row, column=3, value=tot_tm)
                ws.cell(row=tot_row, column=4, value=tot_ag)
                ws.cell(row=tot_row, column=5, value=tot_ko)
                ws.cell(row=tot_row, column=6, value=tot_sy)

                col_widths = [26, 40, 10, 10, 12, 10, 16, 30, 30, 28, 18, 32, 26]
                for ci, w in enumerate(col_widths[:len(all_cols)], 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w

            ws.row_dimensions[1].height = 28
            ws.freeze_panes = 'A2'
            wb.save(out_path)
            school_count = len(sorted_codes)

        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror('Σφάλμα', str(e), parent=self)
            return

        # ── Αποθήκευση ρυθμίσεων ─────────────────────────────────────────────
        s = _load_local_settings()
        s[self._SETTINGS_KEY] = {
            'subject':     self._saved_subject,
            'body':        self._saved_body,
            'dimos_email': to_email,
        }
        path_s = _get_local_settings_path()
        os.makedirs(os.path.dirname(path_s), exist_ok=True)
        with open(path_s, 'w', encoding='utf-8') as f:
            json.dump(s, f, ensure_ascii=False, indent=2)
        if not send:
            messagebox.showinfo('Έτοιμο',
                f'Αρχείο αποθηκεύτηκε:\n{out_path}\n\nΣχολεία: {school_count}',
                parent=self)
            try:
                import subprocess; subprocess.Popen(['explorer', out_dir])
            except Exception:
                pass
            self.destroy()
            return

        try:
            from core.framework import send_email
            send_email(config, to_email, subject, full_body, out_path)
            messagebox.showinfo('Αποστολή',
                f'Email στάλθηκε: {to_email}\n\nΑρχείο: {out_path}\nΣχολεία: {school_count}',
                parent=self)
            try:
                import subprocess; subprocess.Popen(['explorer', out_dir])
            except Exception:
                pass
            self.destroy()
        except Exception as e:
            messagebox.showerror('Σφάλμα αποστολής', str(e), parent=self)

    def _clear(self):
        for w in self.winfo_children():
            w.destroy()


def _show_help(parent):
    """Ανοίγει τον οδηγό PDF με τον προεπιλεγμένο viewer των Windows."""
    import glob
    if getattr(sys, 'frozen', False):
        # Frozen exe: ψάχνε δίπλα στο .exe
        base_dir = os.path.dirname(sys.executable)
        candidates = glob.glob(os.path.join(base_dir, '*.pdf'))
        if not candidates:
            # ή μέσα στο bundle (αν συμπεριληφθεί με --add-data)
            candidates = glob.glob(os.path.join(sys._MEIPASS, '*.pdf'))
    else:
        base = os.path.dirname(os.path.abspath(__file__))
        candidates = glob.glob(os.path.join(base, '..', '*.pdf'))
    pdf_path = candidates[0] if candidates else None

    if pdf_path and os.path.exists(pdf_path):
        os.startfile(os.path.normpath(pdf_path))
    else:
        from tkinter import messagebox
        messagebox.showinfo('Βοήθεια',
                            'Δεν βρέθηκε αρχείο οδηγού (PDF) στον φάκελο της εφαρμογής.',
                            parent=parent)


def _show_splash(root):
    splash = tk.Toplevel(root)
    splash.overrideredirect(True)
    splash.configure(bg=C['hdr_bg'])

    w, h = 460, 320
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    x  = (sw - w) // 2
    y  = (sh - h) // 2
    splash.geometry(f'{w}x{h}+{x}+{y}')
    splash.lift()
    splash.attributes('-topmost', True)

    tk.Label(splash, text='MySchool Αυτοματισμοί',
             bg=C['hdr_bg'], fg=C['hdr_fg'],
             font=('Arial', 17, 'bold')).pack(pady=(24, 2))
    tk.Label(splash, text='Δ/νση Π.Ε. Ανατολικής Θεσσαλονίκης',
             bg=C['hdr_bg'], fg=C['hdr_sub'],
             font=('Arial', 9)).pack()

    from tkinter import ttk as _ttk
    _style = _ttk.Style(splash)
    _style.theme_use('default')
    _style.configure('Splash.Horizontal.TProgressbar',
                     troughcolor='#163D60', background='#E53935',
                     bordercolor=C['hdr_bg'], lightcolor='#E53935',
                     darkcolor='#C62828')
    _style.configure('SplashDone.Horizontal.TProgressbar',
                     troughcolor='#163D60', background='#4CA870',
                     bordercolor=C['hdr_bg'], lightcolor='#4CA870',
                     darkcolor='#4CA870')
    pb = _ttk.Progressbar(splash, mode='indeterminate', length=360,
                          style='Splash.Horizontal.TProgressbar')
    pb.pack(pady=(14, 8))
    pb.start(10)

    log_frame = tk.Frame(splash, bg='#163D60', padx=16, pady=8)
    log_frame.pack(fill='x', padx=30, pady=(0, 8))
    log_txt = tk.Text(log_frame, height=4,
                      bg='#163D60', fg=C['hdr_sub'],
                      font=('Consolas', 8), relief='flat',
                      state='disabled', cursor='arrow', wrap='word',
                      selectbackground='#163D60',
                      insertbackground='#163D60')
    log_txt.pack(fill='x')

    ready_btn = tk.Button(splash, text='▶   Είσοδος',
                          bg='#2E7D32', fg='white',
                          font=('Arial', 10, 'bold'),
                          relief='flat', padx=24, pady=6,
                          cursor='hand2',
                          activebackground='#43A047',
                          activeforeground='white')

    tk.Label(splash, text='Μιχάλης Κατσιρντάκης  •  2310 954145',
             bg=C['hdr_bg'], fg='#4A7FAF',
             font=('Arial', 8)).pack(side='bottom', pady=(0, 10))

    splash.update()
    return splash, pb, log_txt, ready_btn


def _splash_log(log_txt, msg):
    def _do():
        log_txt.configure(state='normal')
        log_txt.insert(tk.END, msg + '\n')
        log_txt.see(tk.END)
        log_txt.configure(state='disabled')
    try:
        log_txt.after(0, _do)
    except Exception:
        pass


def _play_startup_sound(path, on_finished):
    """Παίζει MP3 μέσω MCI (blocking). Καλεί on_finished() μόνο αν το άνοιγμα
    του αρχείου πέτυχε — ώστε αποτυχία MCI να μην προκαλεί πρόωρο launch."""
    _opened = False
    try:
        import ctypes
        mci = ctypes.windll.winmm.mciSendStringW
        if mci(f'open "{path}" type mpegvideo alias splash_snd', None, 0, None) == 0:
            _opened = True
            mci('play splash_snd wait', None, 0, None)
    except Exception:
        pass
    if _opened:
        on_finished()


def _stop_startup_sound():
    try:
        import ctypes
        mci = ctypes.windll.winmm.mciSendStringW
        mci('stop splash_snd', None, 0, None)
        mci('close splash_snd', None, 0, None)
    except Exception:
        pass



def main():
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    root = tk.Tk()
    root.withdraw()

    ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.ico')
    if os.path.exists(ico):
        try: root.iconbitmap(ico)
        except Exception: pass

    splash, pb, log_txt, ready_btn = _show_splash(root)

    checks_result = []
    done_flag     = threading.Event()
    launched      = [False]

    def _do_launch():
        if not launched[0]:
            launched[0] = True
            _stop_startup_sound()
            # Περιμένουμε το done_flag (20s) πριν ανοίξουμε — σε περίπτωση
            # που η μουσική τελειώσει νωρίτερα από τον χρονομετρητή
            def _wait_ready():
                if not done_flag.is_set():
                    root.after(100, _wait_ready)
                    return
                checks = checks_result[0] if checks_result else []
                _launch(root, checks, splash, pb)
            root.after(0, _wait_ready)

    _snd_path  = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'startup.mp3')
    _has_sound = os.path.exists(_snd_path)
    if _has_sound:
        threading.Thread(target=_play_startup_sound,
                         args=(_snd_path, _do_launch), daemon=True).start()

    def _startup():
        import subprocess as _sub
        import traceback as _tb
        _log_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'crash.log')

        def _log(msg):
            try:
                with open(_log_path, 'a', encoding='utf-8') as _f:
                    _f.write(msg + '\n')
            except Exception:
                pass

        try:
            _start_time = time.time()

            _splash_log(log_txt, f'✓ Python {sys.version.split()[0]}')
            time.sleep(0.3)

            if getattr(sys, 'frozen', False):
                # ── Frozen exe: βιβλιοθήκες είναι ήδη bundled ──────────
                # ΔΕΝ τρέχουμε pip — sys.executable είναι το .exe, όχι Python
                _splash_log(log_txt, '✓ Βιβλιοθήκες εγκατεστημένες')
                time.sleep(0.2)
            else:
                # ── Development mode: έλεγχος & εγκατάσταση βιβλιοθηκών ─
                _base    = os.path.dirname(os.path.abspath(__file__))
                _libs_ok = os.path.join(_base, '.libs_ok')
                _reqs    = [('pandas','pandas'), ('openpyxl','openpyxl'),
                            ('selenium','selenium'), ('xlrd','xlrd'),
                            ('html2text','html2text')]

                if os.path.exists(_libs_ok):
                    _splash_log(log_txt, '✓ Βιβλιοθήκες εγκατεστημένες')
                    time.sleep(0.2)
                else:
                    _splash_log(log_txt, 'Έλεγχος βιβλιοθηκών...')
                    for pkg, imp in _reqs:
                        try:
                            __import__(imp)
                            _splash_log(log_txt, f'  ✓ {pkg}')
                        except ImportError:
                            _splash_log(log_txt, f'  ⬇ Εγκατάσταση {pkg}...')
                            _sub.run([sys.executable, '-m', 'pip', 'install', pkg,
                                      '--disable-pip-version-check', '-q'],
                                     capture_output=True)
                            _splash_log(log_txt, f'  ✓ {pkg} εγκαταστάθηκε')
                        time.sleep(0.15)
                    open(_libs_ok, 'w').close()

            _splash_log(log_txt, 'Φόρτωση ελέγχων...')
            checks = load_checks()
            checks_result.append(checks)
            _splash_log(log_txt, f'✓ {len(checks)} έλεγχοι φορτώθηκαν')
            time.sleep(0.4)

            elapsed   = time.time() - _start_time
            remaining = 20 - elapsed
            if remaining > 0:
                time.sleep(remaining)

        except Exception as _e:
            checks_result.append([])
        finally:
            done_flag.set()

    threading.Thread(target=_startup, daemon=True).start()

    def _poll_checks():
        if not done_flag.is_set():
            root.after(100, _poll_checks)
            return
        checks = checks_result[0] if checks_result else []
        if not checks:
            pb.stop()
            from tkinter import messagebox
            _log_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'crash.log')
            messagebox.showerror('Σφάλμα',
                f'Δεν φορτώθηκαν έλεγχοι!\n\nΔες το αρχείο:\n{_log_path}')
            splash.destroy()
            sys.exit(1)
        pb.stop()
        pb.configure(style='SplashDone.Horizontal.TProgressbar',
                     mode='determinate', value=100)
        _splash_log(log_txt, '✓ Έτοιμο!')
        ready_btn.configure(command=_do_launch)
        ready_btn.pack(pady=(6, 4))
        splash.update()

    root.after(100, _poll_checks)
    root.mainloop()


def _launch(root, checks, splash, pb):
    pb.stop()
    splash.destroy()
    root.deiconify()
    LauncherApp(root, checks)


if __name__ == '__main__':
    try:
        main()
    except Exception:
        import traceback
        _log = os.path.join(os.path.expanduser('~'), 'Desktop', 'crash.log')
        with open(_log, 'w', encoding='utf-8') as _f:
            _f.write(traceback.format_exc())
        raise

"""
checks/dioikitiko_ergo.py
═════════════════════════
Έλεγχος καταχωρήσεων διοικητικού έργου (4.12 + Αδυνατούντες).

Φύλλο 1 — Περίπτωση ΠΔΕ απόφασης:
  Εγγραφές 4.12 με μη-μηδενική "Γραμματειακή Υποστήριξη" ΚΑΙ
  Παρατηρήσεις της μορφής "ΠΔΕ αριθμός/ημερομηνία".
  Ομαδοποίηση ανά ειδικότητα → σύγκριση με αρχείο Αδυνατούντων.
  Εμφανίζει αποκλίσεις (ειδικότητες που λείπουν ή έχουν διαφορετικό πλήθος).

Φύλλο 2 — Περίπτωση χωρίς έγκυρη απόφαση:
  Εγγραφές 4.12 με μη-μηδενική "Γραμματειακή Υποστήριξη" ΚΑΙ
  Παρατηρήσεις που ΔΕΝ είναι της μορφής ΠΔΕ.
  Ταξινόμηση αύξουσα βάσει "Γραμματειακή Υποστήριξη".
  Χρωματισμός γραμμής όταν Γραμματειακή == Ώρες Φορέα (πλήρες διοικητικό).
"""

import re, csv, os, smtplib, ssl
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from email.header import Header
from email.utils import formataddr

from core.framework import ask_file, get_downloaded_file, ask_date_yyyymmdd, yes_no, read_csv_fixed, _show_results_popup, ENCODING, SEP

# ── Μεταδεδομένα ────────────────────────────────────────────────────────────
CHECK_TITLE       = 'Έλεγχος καταχωρήσεων διοικητικού έργου'
CHECK_DESCRIPTION = 'Έλεγχος Γραμματειακής Υποστήριξης 4.12 vs Αδυνατούντες'
RESULTS_FOLDER    = 'dioikitiko_ergo'
HAS_EMAIL         = True
CUSTOM_RUN        = True
TEST_ONLY         = True   # Μόνο test mode — δεν αφορά σχολεία

# ── Στήλες 4.12 που χρησιμοποιούμε ─────────────────────────────────────────
COL_GRAM   = 'Γραμματειακή Υποστήριξη'
COL_ORES   = 'Ώρες Υποχ. Διδακτικού Ωραρίου Υπηρέτησης στο Φορέα'
COL_PARAT  = 'Παρατηρήσεις'
COL_EID    = 'Κωδικός Κύριας Ειδικότητας'
COL_AM     = 'ΑΜ'
COL_EPWN   = 'Επώνυμο'
COL_ONOM   = 'Όνομα'
COL_SXOL     = 'Ονομασία Φορέα'
COL_SXOL_ALT = 'Ονομασία Σχολείου'
COL_KWD      = 'Κωδικός Φορέα'
COL_EIDOS    = 'Είδος Σχολείου'

# ── Regex για ΠΔΕ απόφαση ────────────────────────────────────────────────────
# Π.χ. "ΠΔΕ 1234/15-03-2026" ή "ΠΔΕ 1234/2026"
RE_PDE = re.compile(r'ΠΔΕ\s+\d+/[\d\-]+', re.IGNORECASE)

# ── Χρώματα Excel ────────────────────────────────────────────────────────────
COLOR_HEADER    = '1F4E79'
COLOR_SUB       = 'D6E4F0'
COLOR_ALT       = 'EBF3FB'
COLOR_FULL      = 'FFE2CC'   # πλήρες διοικητικό (Γραμ == Ώρες Φορέα)
COLOR_FULL_ALT  = 'FFD0A8'
COLOR_OK        = 'E8F5E9'   # καμία απόκλιση στο φύλλο 1
COLOR_DIFF      = 'FFEEEE'   # απόκλιση στο φύλλο 1
COLOR_DIFF_ALT  = 'FFD6D6'
COLOR_HDR_GRN   = '375623'   # header φύλλου 1 (πράσινο)
COLOR_HDR_AMB   = 'BF8F00'   # header φύλλου 2 (κεχριμπαρί)

EMAIL_SUBJECT = 'Αποτελέσματα ελέγχου καταχωρήσεων διοικητικού έργου'


# ═══════════════════════════════════════════════════════════════════
# ΦΟΡΤΩΣΗ
# ═══════════════════════════════════════════════════════════════════

def _to_num(series):
    """Καθαρισμός ="..." και μετατροπή σε αριθμό."""
    return pd.to_numeric(
        series.astype(str)
              .str.replace('="', '', regex=False)
              .str.replace('"', '', regex=False)
              .str.replace(',', '.', regex=False)
              .str.strip(),
        errors='coerce'
    ).fillna(0)

def load_412(path):
    df = read_csv_fixed(path)
    df.columns = [c.strip() for c in df.columns]
    # Καθαρισμός βασικών στηλών
    for col in [COL_AM, COL_EID, COL_KWD]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace('="','').str.replace('"','').str.strip()
    df[COL_GRAM] = _to_num(df[COL_GRAM])
    df[COL_ORES] = _to_num(df[COL_ORES])
    df[COL_PARAT] = df[COL_PARAT].astype(str).str.strip() if COL_PARAT in df.columns else ''
    return df

def load_adynatountes(path):
    """
    Φορτώνει το αρχείο αδυνατούντων ανά ειδικότητα.
    Αναμένεται: στήλη "Κωδικός" + στήλη πλήθους.
    Επιστρέφει dict {κωδικός_ειδικότητας: πλήθος}.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        df = read_csv_fixed(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip() for c in df.columns]

    # Στήλη κωδικού: ακριβώς "Κωδικός", αλλιώς πρώτη στήλη
    eid_col = 'Κωδικός' if 'Κωδικός' in df.columns else df.columns[0]

    # Στήλη πλήθους: οτιδήποτε εκτός από "Κωδικός" και κείμενο ειδικότητας
    cnt_col = next((c for c in df.columns
                    if c != eid_col and any(w in c.lower()
                    for w in ['πλήθος','αριθμ','σύνολ','count','total','σύν'])),
                   None)
    if not cnt_col:
        # Fallback: δεύτερη αριθμητική στήλη
        cnt_col = next((c for c in df.columns if c != eid_col), df.columns[1])

    print(f'  Αδυνατούντες — στήλη κωδικού: "{eid_col}", στήλη πλήθους: "{cnt_col}"')

    result = {}
    for _, row in df.iterrows():
        eid = str(row[eid_col]).strip().replace('="','').replace('"','')
        try:
            cnt = int(float(str(row[cnt_col]).replace(',', '.')))
        except (ValueError, TypeError):
            continue
        if eid and eid not in ('nan', 'None', ''):
            result[eid] = cnt
    return result


# ═══════════════════════════════════════════════════════════════════
# ΕΠΕΞΕΡΓΑΣΙΑ
# ═══════════════════════════════════════════════════════════════════

def is_pde(parat):
    return bool(RE_PDE.search(str(parat)))

def process(df412, adynatountes):
    """
    Επιστρέφει (df_pde, df_no_pde, summary1, summary2).
    df_pde    : DataFrame Περίπτωσης 1 (αποκλίσεις)
    df_no_pde : DataFrame Περίπτωσης 2 (χωρίς έγκυρη απόφαση)
    summary1  : str — περίληψη φύλλου 1 για email body
    summary2  : str — περίληψη φύλλου 2 για email body
    """
    # Φίλτρο: μόνο μη-μηδενική Γραμματειακή
    df = df412[df412[COL_GRAM] > 0].copy()

    # Διαχωρισμός
    mask_pde   = df[COL_PARAT].apply(is_pde)
    df_pde_raw = df[mask_pde].copy()
    df_nopde   = df[~mask_pde].copy()

    # ── Περίπτωση 1: ομαδοποίηση ανά ειδικότητα & σύγκριση ──────────
    grp = df_pde_raw.groupby(COL_EID).size().to_dict()   # {ειδικότητα: πλήθος 4.12}

    rows_p1 = []
    all_eids = sorted(set(list(grp.keys()) + list(adynatountes.keys())))
    has_diff = False

    for eid in all_eids:
        cnt_412  = grp.get(eid, 0)
        cnt_ady  = adynatountes.get(eid, 0)
        diff     = cnt_412 - cnt_ady
        status   = 'ΟΚ' if diff == 0 else ('ΠΛΕΟΝ ' + str(abs(diff)) if diff > 0 else 'ΕΛΛΕΙΠ ' + str(abs(diff)))
        if diff != 0:
            has_diff = True
        rows_p1.append({
            'Κωδικός Ειδικότητας': eid,
            'Πλήθος 4.12':         cnt_412,
            'Πλήθος Αδυνατούντων': cnt_ady,
            'Διαφορά':             diff,
            'Κατάσταση':           status,
        })

    df_p1 = pd.DataFrame(rows_p1)

    # Summary φύλλου 1
    if not has_diff:
        summary1 = 'Φύλλο 1 (ΠΔΕ αποφάσεις): Καμία απόκλιση μεταξύ 4.12 και αρχείου Αδυνατούντων.'
    else:
        diffs = df_p1[df_p1['Διαφορά'] != 0]
        lines = []
        for _, r in diffs.iterrows():
            lines.append(f"  - {r['Κωδικός Ειδικότητας']}: 4.12={r['Πλήθος 4.12']}, "
                         f"Αδυνατούντες={r['Πλήθος Αδυνατούντων']}, "
                         f"Διαφορά={r['Διαφορά']:+d}")
        summary1 = ('Φύλλο 1 (ΠΔΕ αποφάσεις): Αποκλίσεις σε '
                    f'{len(diffs)} ειδικότητες:\n' + '\n'.join(lines))

    # ── Περίπτωση 2: εξαίρεση Ιδιωτικών + ταξινόμηση + σήμανση πλήρους ──
    if COL_EIDOS in df_nopde.columns:
        before = len(df_nopde)
        df_nopde = df_nopde[df_nopde[COL_EIDOS].str.strip() != 'Ιδιωτικά Σχολεία'].copy()
        print(f'  Εξαιρέθηκαν {before - len(df_nopde)} εγγραφές Ιδιωτικών Σχολείων')
    df_nopde = df_nopde.sort_values(COL_GRAM)

    # Στήλη σήμανσης
    df_nopde['_full'] = df_nopde[COL_GRAM] == df_nopde[COL_ORES]

    cnt_total = len(df_nopde)
    cnt_full  = df_nopde['_full'].sum()

    summary2 = (f'Φύλλο 2 (χωρίς έγκυρη απόφαση): {cnt_total} εγγραφές '
                f'εκ των οποίων {cnt_full} με πλήρες διοικητικό '
                f'(Γραμματειακή = Ώρες Φορέα).')

    return df_p1, df_nopde, summary1, summary2


# ═══════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════

def _brd():
    t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=t)

def _cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border
    return c

def build_workbook(df_p1, df_p2, today, out_path):
    wb  = Workbook()
    brd = _brd()
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ══════════════════════════════════════════════════
    # ΦΥΛΛΟ 1 — Σύγκριση ΠΔΕ vs Αδυνατούντες
    # ══════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'ΠΔΕ vs Αδυνατούντες'

    cols1 = [
        ('Κωδικός Ειδικότητας', 20),
        ('Πλήθος 4.12',         14),
        ('Πλήθος Αδυνατούντων', 20),
        ('Διαφορά',             12),
        ('Κατάσταση',           16),
    ]
    ncols1 = len(cols1)

    # Τίτλος
    ws1.merge_cells(f'A1:{get_column_letter(ncols1)}1')
    ws1['A1'] = f'{CHECK_TITLE}  —  Σύγκριση ΠΔΕ vs Αδυνατούντες  —  {today.strftime("%d/%m/%Y")}'
    ws1['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws1['A1'].fill      = PatternFill('solid', start_color=COLOR_HDR_GRN)
    ws1['A1'].alignment = ctr
    ws1.row_dimensions[1].height = 24

    ws1.merge_cells(f'A2:{get_column_letter(ncols1)}2')
    has_diff = (df_p1['Διαφορά'] != 0).any()
    ws1['A2'] = (f'Σύνολο ειδικοτήτων: {len(df_p1)}  |  '
                 f'Αποκλίσεις: {(df_p1["Διαφορά"] != 0).sum()}')
    ws1['A2'].font      = Font(name='Arial', italic=True, size=9)
    ws1['A2'].fill      = PatternFill('solid', start_color=COLOR_SUB)
    ws1['A2'].alignment = ctr
    ws1.row_dimensions[2].height = 16

    for ci, (name, width) in enumerate(cols1, 1):
        c = ws1.cell(row=3, column=ci, value=name)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill      = PatternFill('solid', start_color=COLOR_HDR_GRN)
        c.alignment = ctr
        c.border    = brd
        ws1.column_dimensions[get_column_letter(ci)].width = width
    ws1.row_dimensions[3].height = 28

    for ri, (_, row) in enumerate(df_p1.iterrows(), start=4):
        is_diff = row['Διαφορά'] != 0
        fill_c  = COLOR_DIFF if is_diff else COLOR_OK
        if ri % 2 == 0 and is_diff:
            fill_c = COLOR_DIFF_ALT
        fill = PatternFill('solid', start_color=fill_c)
        vals = [row['Κωδικός Ειδικότητας'], row['Πλήθος 4.12'],
                row['Πλήθος Αδυνατούντων'], row['Διαφορά'], row['Κατάσταση']]
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font      = Font(name='Arial', size=10,
                               bold=(ci == 4 and is_diff),
                               color='8B0000' if (ci == 4 and is_diff) else '000000')
            c.fill      = fill
            c.alignment = lft if ci == 1 else ctr
            c.border    = brd
        ws1.row_dimensions[ri].height = 18

    ws1.freeze_panes = 'A4'
    ws1.auto_filter.ref = f'A3:{get_column_letter(ncols1)}3'

    # ══════════════════════════════════════════════════
    # ΦΥΛΛΟ 2 — Χωρίς έγκυρη απόφαση
    # ══════════════════════════════════════════════════
    ws2 = wb.create_sheet('Χωρίς Απόφαση ΠΔΕ')

    # Επιλογή ονόματος στήλης σχολείου (ό,τι υπάρχει στο df)
    sxol_col = COL_SXOL_ALT if COL_SXOL_ALT in df_p2.columns else COL_SXOL

    # Δυναμικές στήλες — παίρνουμε ό,τι υπάρχει στο df
    base_cols = [
        (COL_KWD,   14, 'Κωδικός Σχολείου'),
        (sxol_col,  42, 'Ονομασία Σχολείου'),
        (COL_AM,    11, 'ΑΜ'),
        (COL_EPWN,  18, 'Επώνυμο'),
        (COL_ONOM,  14, 'Όνομα'),
        (COL_EID,   14, 'Ειδικότητα'),
        (COL_ORES,  14, 'Ώρες Φορέα'),
        (COL_GRAM,  16, 'Γραμματειακή Υποστήριξη'),
        (COL_PARAT, 40, 'Παρατηρήσεις'),
    ]
    # Κρατάμε μόνο στήλες που υπάρχουν στο df
    cols2 = [(src, w, lbl) for src, w, lbl in base_cols if src in df_p2.columns or src == sxol_col]
    ncols2 = len(cols2)

    ws2.merge_cells(f'A1:{get_column_letter(ncols2)}1')
    ws2['A1'] = f'{CHECK_TITLE}  —  Χωρίς Απόφαση ΠΔΕ  —  {today.strftime("%d/%m/%Y")}'
    ws2['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws2['A1'].fill      = PatternFill('solid', start_color=COLOR_HDR_AMB)
    ws2['A1'].alignment = ctr
    ws2.row_dimensions[1].height = 24

    cnt_full = df_p2['_full'].sum()
    ws2.merge_cells(f'A2:{get_column_letter(ncols2)}2')
    ws2['A2'] = (f'Σύνολο εγγραφών: {len(df_p2)}  |  '
                 f'Πλήρες διοικητικό (Γραμματειακή = Ώρες Φορέα): {cnt_full}  '
                 f'— χρωματισμένες γραμμές')
    ws2['A2'].font      = Font(name='Arial', italic=True, size=9)
    ws2['A2'].fill      = PatternFill('solid', start_color='FFF2CC')
    ws2['A2'].alignment = ctr
    ws2.row_dimensions[2].height = 16

    for ci, (_, width, label) in enumerate(cols2, 1):
        c = ws2.cell(row=3, column=ci, value=label)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill      = PatternFill('solid', start_color=COLOR_HDR_AMB)
        c.alignment = ctr
        c.border    = brd
        ws2.column_dimensions[get_column_letter(ci)].width = width
    ws2.row_dimensions[3].height = 28

    for ri, (_, row) in enumerate(df_p2.iterrows(), start=4):
        is_full  = bool(row.get('_full', False))
        if is_full:
            fill_color = COLOR_FULL if ri % 2 == 0 else COLOR_FULL_ALT
        else:
            fill_color = COLOR_ALT if ri % 2 == 0 else 'FFFFFF'
        fill = PatternFill('solid', start_color=fill_color)

        for ci, (src_col, _, _) in enumerate(cols2, 1):
            val = row.get(src_col, '')
            if hasattr(val, 'item'):
                val = val.item()
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font      = Font(name='Arial', size=9,
                               bold=is_full and src_col == COL_GRAM)
            c.fill      = fill
            c.alignment = ctr if src_col in (COL_AM, COL_EID, COL_ORES, COL_GRAM, COL_KWD) else lft
            c.border    = brd
        ws2.row_dimensions[ri].height = 16

    ws2.freeze_panes = 'A4'
    ws2.auto_filter.ref = f'A3:{get_column_letter(ncols2)}3'

    wb.save(out_path)


# ═══════════════════════════════════════════════════════════════════
# EMAIL
# ═══════════════════════════════════════════════════════════════════

def _send(cfg, to_list, subject, body, attachment_path):
    msg = MIMEMultipart()
    from_name = getattr(cfg, 'FROM_NAME', cfg.FROM_EMAIL)
    msg['From']    = formataddr((from_name, cfg.FROM_EMAIL))
    msg['To']      = to_list[0]
    msg['Subject'] = Header(subject, 'utf-8')
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    fname = os.path.basename(attachment_path)
    with open(attachment_path, 'rb') as f:
        part = MIMEBase('application', 'vnd.ms-excel')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename=fname)
    part.add_header('Content-Type', 'application/vnd.ms-excel', name=fname)
    msg.attach(part)

    def _lenient_ctx():
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode    = ssl.CERT_NONE
        return ctx

    recipients = list(dict.fromkeys(to_list))  # dedup
    msg_str = msg.as_string()
    sent = False
    try:
        with smtplib.SMTP_SSL(cfg.SMTP_HOST, 465,
                               context=ssl.create_default_context()) as s:
            s.login(cfg.FROM_EMAIL, cfg.FROM_PASSWORD)
            s.sendmail(cfg.FROM_EMAIL, recipients, msg_str)
            sent = True
    except Exception:
        pass
    if not sent:
        try:
            with smtplib.SMTP_SSL(cfg.SMTP_HOST, 465,
                                   context=_lenient_ctx()) as s:
                s.login(cfg.FROM_EMAIL, cfg.FROM_PASSWORD)
                s.sendmail(cfg.FROM_EMAIL, recipients, msg_str)
                sent = True
        except Exception:
            pass
    if not sent:
        with smtplib.SMTP(cfg.SMTP_HOST, 587) as s:
            s.starttls(context=_lenient_ctx())
            s.login(cfg.FROM_EMAIL, cfg.FROM_PASSWORD)
            s.sendmail(cfg.FROM_EMAIL, recipients, msg_str)


# ═══════════════════════════════════════════════════════════════════
# CUSTOM RUN
# ═══════════════════════════════════════════════════════════════════

def run(config):
    import core.framework as _fw
    _fw._current_check_title = CHECK_TITLE

    print('=' * 65)
    print(f'  {CHECK_TITLE}')
    print('=' * 65)

    path_412 = get_downloaded_file('4.12', 'Αρχείο 4.12 [csv]:', csv_only=True)
    path_ady = get_downloaded_file('ady', 'Αρχείο Αδυνατούντων ανά ειδικότητα [csv / xlsx]:')
    today    = ask_date_yyyymmdd()
    from core.framework import _ask_send_options_gui
    test_mode, do_send = _ask_send_options_gui(test_only=True)
    # Για αυτόν τον έλεγχο μόνο test mode έχει νόημα

    print(f'\n  Ημερομηνία : {today.strftime("%d/%m/%Y")}')
    print(f'  Λειτουργία : {"TEST MODE" if test_mode else "Χωρίς αποστολή"}')
    print('-' * 65)

    # Φόρτωση
    print('\nΦόρτωση αρχείων...')
    df412 = load_412(path_412)
    adyn  = load_adynatountes(path_ady)
    print(f'  ✓ 4.12           : {len(df412)} εγγραφές')
    print(f'  ✓ Αδυνατούντες   : {len(adyn)} ειδικότητες')

    # Επεξεργασία
    print('\nΕπεξεργασία...')
    df_p1, df_p2, summary1, summary2 = process(df412, adyn)
    print(f'  ✓ Φύλλο 1 (ΠΔΕ)            : {len(df_p1)} ειδικότητες')
    print(f'  ✓ Φύλλο 2 (χωρίς απόφαση) : {len(df_p2)} εγγραφές')

    # Αποθήκευση
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_dir  = os.path.join(base_dir, f'results_{today.strftime("%Y%m%d")}', 'dioikitiko')
    os.makedirs(out_dir, exist_ok=True)

    out_path = os.path.join(out_dir, f'{today.strftime("%Y%m%d")}_DIOIKITIKO.xlsx')
    build_workbook(df_p1, df_p2, today, out_path)
    print(f'\n  ✓ Αποθηκεύτηκε: {os.path.basename(out_path)}')

    # Σύνοψη αποτελεσμάτων (χτίζεται πάντα για popup ή email)
    body = (
        f'Σύνοψη ελέγχου καταχωρήσεων διοικητικού έργου — {today.strftime("%d/%m/%Y")}\n'
        f'{"─"*50}\n\n'
        f'{summary1}\n\n'
        f'{summary2}\n\n'
        f'Λεπτομερειες στο επισυναπτομενο αρχειο.'
    )

    # Email
    if do_send:
        cc_extra  = getattr(config, 'TEST_EMAIL_CC', None)
        to_list   = [config.TEST_EMAIL] + ([cc_extra] if cc_extra else [])
        subject   = f'[TEST] {EMAIL_SUBJECT} — {today.strftime("%d/%m/%Y")}'

        print(f'\n  Προεπισκόπηση body:\n{"─"*40}')
        print(body)
        print('─' * 40)
        try:
            _send(config, to_list, subject, body, out_path)
            print(f'  ✓ Εστάλη στο {", ".join(to_list)}')
        except Exception as e:
            print(f'  ✗ Σφάλμα: {e}')

    popup_body = body + (
        f'\n\n{"─"*40}\nΑποτελέσματα αποθηκεύτηκαν στο φάκελο:\n{out_dir}\n\n'
        f'📋 Για περισσότερες πληροφορίες δες τα αρχεία\nστο φάκελο αποτελεσμάτων.'
    )
    _show_results_popup(CHECK_TITLE, popup_body, result_type='warn')

    print('─' * 65)
    print(f'\n✓ Ολοκληρώθηκε! Αρχείο στο φάκελο:\n  {out_dir}')


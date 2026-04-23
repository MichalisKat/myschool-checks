"""
checks/ypoloipa.py
══════════════════
Υπόλοιπα Ωραρίου Εκπαιδευτικών (4.8).

Λόγω της σύνθετης φύσης του (pivot sheets, κατώφλι, 4.11/4.12, αδυνατούντες),
αυτό το module έχει δικό του run() που παρακάμπτει το κοινό framework.
Εμφανίζεται κανονικά στο GUI launcher.
"""

import csv, os, smtplib, ssl
import pandas as pd
import config
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from email.header import Header
from email.utils import formataddr

from core.framework import ask_file, get_downloaded_file, get_ady_xoris_egkrisi, ask_date_yyyymmdd, yes_no, _show_results_popup, ENCODING, SEP

# ── Μεταδεδομένα ────────────────────────────────────────────────────────────
CHECK_TITLE       = 'Υπόλοιπα Ωραρίου'
CHECK_DESCRIPTION = 'Υπόλοιπα ωραρίου εκπαιδευτικών ανά σχολείο και ειδικότητα'
RESULTS_FOLDER    = 'ypoloipa_wrariou'
HAS_EMAIL         = True
CUSTOM_RUN        = True   # Σημαία: το framework θα καλέσει run() αντί για run_check()

EIDI_SXOLEION  = ['Δημοτικά Σχολεία', 'Νηπιαγωγεία']
ADYN_FILTER    = 'ΑΔΥΝΑΤΟΥΝΤΕΣ'
JOIN_KEY_48    = 'Α.Μ.'
JOIN_KEY_ADY   = 'ΑΜΥ '
JOIN_KEY_412   = 'ΑΜ'
DETAIL_COL_START = 39
DETAIL_COL_END   = 65
DETAIL_411_START = 39
DETAIL_411_END   = 58
SCHOOL_COLUMN  = 'Ονομασία Σχολείου'
COL_YPOLOIPO   = 'Υπόλοιπο Υποχρεωτικού Διδακτικού Ωραρίου'
COL_SYMPL      = 'Συμπλήρωση Ωραρίου'
COL_SYMPL_DET  = '_Λεπτομέρειες Συμπλήρωσης'
COL_MEIWSI_DET = '_Λεπτομέρειες Μείωσης'
HIGHLIGHT_COL  = COL_YPOLOIPO

COLOR_MAIN    = '1F4E79'
COLOR_SUB     = 'D6E4F0'
COLOR_ALT     = 'EBF3FB'
COLOR_PIVOT   = '375623'
COLOR_ALT_P   = 'E8F5E9'
COLOR_HEAD2P  = 'C8E6C9'
COLOR_HL_HEADER = 'F4B942'
COLOR_HL_EVEN   = 'FFF3CD'
COLOR_HL_ODD    = 'FFF8E1'

COLUMNS = [
    ('Κωδικός Σχολείου',                                          14, None),
    ('Ονομασία Σχολείου',                                         42, None),
    ('Email',                                                     28, 'Email Σχολείου'),
    ('Α.Μ.',                                                      11, None),
    ('Επώνυμο',                                                   18, None),
    ('Όνομα',                                                     14, None),
    ('Κωδικός Κύριας Ειδικότητας',                                12, None),
    ('Διευθυντής Σχολείου',                                       10, None),
    ('Υποδιευθυντής Σχολείου',                                    10, None),
    ('Ώρες Υποχ. Διδακτικού Ωραρίου Υπηρέτησης στο Φορέα',       14, 'Ώρες στο Φορέα'),
    ('Μείωση Ωραρίου',                                            10, None),
    (COL_MEIWSI_DET,                                              38, 'Λεπτομέρειες Μείωσης'),
    ('Α Ανάθεση Συνολικά',                                        10, 'Α Ανάθεση'),
    ('Β Ανάθεση Συνολικά',                                        10, 'Β Ανάθεση'),
    ('Προσθ. Τμημ. Συνολικά',                                     10, 'Προσθ. Τμημ.'),
    ('Συμπλήρωση Ωραρίου',                                        12, None),
    (COL_SYMPL_DET,                                               38, 'Λεπτομέρειες Συμπλήρωσης'),
    ('Υπόλοιπο Υποχρεωτικού Διδακτικού Ωραρίου',                 14, 'Υπόλοιπο Ωραρίου'),
]

CENTER_COLS = {
    'Α.Μ.', 'Α.Φ.Μ.', 'Κωδικός Σχολείου', 'Κωδικός Κύριας Ειδικότητας',
    'Διευθυντής Σχολείου', 'Υποδιευθυντής Σχολείου',
    'Ώρες Υποχ. Διδακτικού Ωραρίου Υπηρέτησης στο Φορέα',
    'Μείωση Ωραρίου', 'Α Ανάθεση Συνολικά', 'Β Ανάθεση Συνολικά',
    'Γ Ανάθεση Συνολικά', 'Προσθ. Τμημ. Συνολικά',
    "'λλες Αναθέσεις Συνολικά", 'Συμπλήρωση Ωραρίου',
    'Υπόλοιπο Υποχρεωτικού Διδακτικού Ωραρίου',
}

EMAIL_SUBJECT = 'Υπόλοιπα ωραρίου-Αναθέσεις στο Myschool'
EMAIL_BODY    = lambda school='': (
    'Καλημέρα,\n\n'
    'Επισυνάπτονται σε αρχείο excel τα υπόλοιπα ωρών στο Myschool.\n'
    'Παρακαλούμε για άμεσες ενέργειες (αναθέσεις) !\n\n'
    + config.email_signature()
)


# ═══════════════════════════════════════════════════════════════════
# ΦΟΡΤΩΣΗ
# ═══════════════════════════════════════════════════════════════════

def _read_csv(path):
    rows = []
    with open(path, encoding=ENCODING) as f:
        reader = csv.reader(f, delimiter=SEP)
        headers = next(reader)
        n = len(headers)
        for row in reader:
            rows.append(row[:n])
    return pd.DataFrame(rows, columns=headers)

def _clean(series):
    return series.astype(str).str.replace('="', '').str.replace('"', '').str.strip()

def _to_int(val):
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return 0

def load_48(path):
    df = _read_csv(path)
    for col in ['Α.Μ.', 'Α.Φ.Μ.', 'Κωδικός Σχολείου']:
        if col in df.columns:
            df[col] = _clean(df[col])
    return df[df['Είδος Σχολείου'].isin(EIDI_SXOLEION)].copy()

def load_412(path):
    rows = []
    with open(path, encoding=ENCODING) as f:
        reader = csv.reader(f, delimiter=SEP)
        headers = next(reader)
        n = len(headers)
        for row in reader:
            rows.append(row[:n])
    df = pd.DataFrame(rows, columns=headers)
    df[JOIN_KEY_412] = df[JOIN_KEY_412].astype(str).str.strip()
    detail_cols = headers[DETAIL_COL_START:DETAIL_COL_END]
    lookup = {}
    for _, row in df.iterrows():
        am = row[JOIN_KEY_412]
        if not am:
            continue
        active = [col for col in detail_cols
                  if str(row.get(col, '')).strip() not in ('', '0')]
        if active:
            existing = lookup.get(am, [])
            for a in active:
                if a not in existing:
                    existing.append(a)
            lookup[am] = existing
    return lookup

def load_411(path):
    rows = []
    with open(path, encoding=ENCODING) as f:
        reader = csv.reader(f, delimiter=SEP)
        headers = next(reader)
        n = len(headers)
        for row in reader:
            rows.append(row[:n])
    df = pd.DataFrame(rows, columns=headers)
    df[JOIN_KEY_412] = df[JOIN_KEY_412].astype(str).str.strip()
    detail_cols = headers[DETAIL_411_START:DETAIL_411_END]
    lookup = {}
    for _, row in df.iterrows():
        am = row[JOIN_KEY_412]
        if not am:
            continue
        active = [col for col in detail_cols
                  if str(row.get(col, '')).strip() not in ('', '0')]
        if active:
            existing = lookup.get(am, [])
            for a in active:
                if a not in existing:
                    existing.append(a)
            lookup[am] = existing
    return lookup

def load_adynatoyntes(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    rows = [r for r in data[1:] if r[0] is not None]
    df = pd.DataFrame(rows, columns=data[0])
    df = df[df['ΠΑΡΑΤΗΡΗΣΕΙΣ'] == ADYN_FILTER].copy()
    df['AMY_str'] = df[JOIN_KEY_ADY].apply(
        lambda x: str(int(x)) if pd.notna(x) and x != '' else ''
    )
    return df


# ═══════════════════════════════════════════════════════════════════
# ΕΠΕΞΕΡΓΑΣΙΑ
# ═══════════════════════════════════════════════════════════════════

def process_data(df48, df_ady, threshold, lookup_412=None, lookup_411=None):
    if df_ady is not None and not df_ady.empty:
        ady_set = set(df_ady[df_ady['AMY_str'] != '']['AMY_str'])
    else:
        ady_set = set()
    df = df48[~df48[JOIN_KEY_48].isin(ady_set)].copy()
    df['_yp'] = df[COL_YPOLOIPO].apply(_to_int)
    df = df[df['_yp'] >= threshold].copy()

    def spec_sort_key(row):
        eid   = str(row['Κωδικός Κύριας Ειδικότητας']).strip()
        eidos = str(row['Είδος Σχολείου']).strip()
        if eidos == 'Δημοτικά Σχολεία' and eid == 'ΠΕ70':
            return '0_' + eid
        if eidos == 'Νηπιαγωγεία' and eid == 'ΠΕ60':
            return '0_' + eid
        return '1_' + eid

    df['_spec_sort'] = df.apply(spec_sort_key, axis=1)
    df['_yp_desc']   = -df['_yp']
    df = df.sort_values([SCHOOL_COLUMN, '_spec_sort', '_yp_desc'])
    df.drop(columns=['_yp', '_spec_sort', '_yp_desc'], inplace=True)

    if lookup_412 is not None:
        def get_sympl(row):
            if _to_int(row.get(COL_SYMPL, 0)) == 0:
                return ''
            am = str(row.get(JOIN_KEY_48, '')).strip()
            return ', '.join(lookup_412.get(am, []))
        df[COL_SYMPL_DET] = df.apply(get_sympl, axis=1)
    else:
        df[COL_SYMPL_DET] = ''

    if lookup_411 is not None:
        def get_meiwsi(row):
            if _to_int(row.get('Μείωση Ωραρίου', 0)) == 0:
                return ''
            am = str(row.get(JOIN_KEY_48, '')).strip()
            return ', '.join(lookup_411.get(am, []))
        df[COL_MEIWSI_DET] = df.apply(get_meiwsi, axis=1)
    else:
        df[COL_MEIWSI_DET] = ''

    return df


# ═══════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════

def _brd():
    t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=t)

def _brdH():
    t = Side(style='thin',   color='CCCCCC')
    m = Side(style='medium', color='F4B942')
    return Border(left=m, right=m, top=t, bottom=t)

def build_sheet(ws, df_sheet, today, col_defs,
                center_cols=None, hdr_color=COLOR_MAIN,
                alt_color=COLOR_ALT, sub_color=COLOR_SUB,
                title_extra='', highlight_key=None):
    if center_cols is None:
        center_cols = CENTER_COLS
    brd  = _brd()
    brdh = _brdH()
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    labels = [c[2] if c[2] else c[0] for c in col_defs]
    keys   = [c[0] for c in col_defs]
    widths = [c[1] for c in col_defs]
    ncols  = len(col_defs)
    hl_idx = keys.index(highlight_key) + 1 if highlight_key and highlight_key in keys else None

    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    t = CHECK_TITLE + (f'  —  {title_extra}' if title_extra else '') + f'  —  {today.strftime("%d/%m/%Y")}'
    ws['A1'] = t
    ws['A1'].font      = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws['A1'].fill      = PatternFill('solid', start_color=hdr_color)
    ws['A1'].alignment = ctr
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
    ws['A2'] = f'Σύνολο εγγραφών: {len(df_sheet)}'
    ws['A2'].font      = Font(name='Arial', italic=True, size=9)
    ws['A2'].fill      = PatternFill('solid', start_color=sub_color)
    ws['A2'].alignment = ctr
    ws.row_dimensions[2].height = 16

    for ci, (label, width) in enumerate(zip(labels, widths), 1):
        is_hl = (ci == hl_idx)
        c = ws.cell(row=3, column=ci, value=label)
        c.font      = Font(name='Arial', bold=True,
                           color='1F4E79' if is_hl else 'FFFFFF', size=10 if is_hl else 10)
        c.fill      = PatternFill('solid', start_color=COLOR_HL_HEADER if is_hl else hdr_color)
        c.border    = brdh if is_hl else brd
        c.alignment = ctr
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 30

    fill_alt      = PatternFill('solid', start_color=alt_color)
    fill_hl_even  = PatternFill('solid', start_color=COLOR_HL_EVEN)
    fill_hl_odd   = PatternFill('solid', start_color=COLOR_HL_ODD)
    for ri, (_, row) in enumerate(df_sheet.iterrows(), start=4):
        is_even  = (ri % 2 == 0)
        row_fill = fill_alt if is_even else PatternFill()
        for ci, key in enumerate(keys, 1):
            is_hl = (ci == hl_idx)
            val   = row.get(key, '')
            c     = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name='Arial', bold=is_hl, size=9,
                               color='1F4E79' if is_hl else '000000')
            c.fill      = (fill_hl_even if is_even else fill_hl_odd) if is_hl else row_fill
            c.border    = brdh if is_hl else brd
            c.alignment = ctr if key in center_cols else lft
        ws.row_dimensions[ri].height = 16
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(ncols)}3'

def save_main_workbook(df_sheet, today, output_path, school_name=''):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Αποτελέσματα'
    build_sheet(ws, df_sheet, today, COLUMNS,
                title_extra=school_name, highlight_key=HIGHLIGHT_COL)
    wb.save(output_path)

def save_pivot_workbook(df, today, output_path):
    df2    = df.copy()
    df2['_yp'] = df2[COL_YPOLOIPO].apply(_to_int)
    df_nz  = df2[df2['_yp'] > 0].copy()
    brd    = _brd()
    brdh   = _brdH()
    ctr    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    wb     = Workbook()
    alt_p  = PatternFill('solid', start_color=COLOR_ALT_P)

    def _hdr(ws, text, ncols, hdr_color):
        ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
        ws['A1'] = f'{text}  —  {today.strftime("%d/%m/%Y")}'
        ws['A1'].font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', start_color=hdr_color)
        ws['A1'].alignment = ctr
        ws.row_dimensions[1].height = 24

    def _sub(ws, text, ncols, sub_color):
        ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
        ws['A2'] = text
        ws['A2'].font = Font(name='Arial', italic=True, size=9)
        ws['A2'].fill = PatternFill('solid', start_color=sub_color)
        ws['A2'].alignment = ctr
        ws.row_dimensions[2].height = 16

    def _total_row(ws, tr, vals, hl_ci=None):
        for ci, val in enumerate(vals, 1):
            is_hl = (ci == hl_ci)
            c = ws.cell(row=tr, column=ci, value=val)
            c.font   = Font(name='Arial', bold=True, size=10, color='1F4E79')
            c.fill   = PatternFill('solid', start_color=COLOR_HL_HEADER if is_hl else 'FFF3CD')
            c.border = brdh if is_hl else brd
            c.alignment = lft if ci == 2 else ctr
        ws.row_dimensions[tr].height = 20

    def _detail_sheet(ws, df_s, col_defs, lbl_defs, hdr_color, sub_color, alt_color, title):
        nc      = len(col_defs)
        hl_key  = COL_YPOLOIPO
        ctr_keys = {'Α.Μ.', 'Κωδικός Κύριας Ειδικότητας', COL_YPOLOIPO,
                    'Ώρες Υποχ. Διδακτικού Ωραρίου Υπηρέτησης στο Φορέα',
                    'Α Ανάθεση Συνολικά', "'λλες Αναθέσεις Συνολικά", 'Συμπλήρωση Ωραρίου'}
        _hdr(ws, title, nc, hdr_color)
        _sub(ws, f'Σύνολο: {len(df_s)} εγγραφές', nc, sub_color)
        for ci, ((key, w), lbl) in enumerate(zip(col_defs, lbl_defs), 1):
            is_hl = (key == hl_key)
            c = ws.cell(row=3, column=ci, value=lbl)
            c.font   = Font(name='Arial', bold=True, color='1F4E79' if is_hl else 'FFFFFF', size=10)
            c.fill   = PatternFill('solid', start_color=COLOR_HL_HEADER if is_hl else hdr_color)
            c.alignment = ctr; c.border = brdh if is_hl else brd
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[3].height = 28
        alt = PatternFill('solid', start_color=alt_color)
        for ri, (_, row) in enumerate(df_s.iterrows(), 4):
            fill = alt if ri % 2 == 0 else PatternFill()
            for ci, (key, _) in enumerate(col_defs, 1):
                is_hl = (key == hl_key)
                hl_f  = PatternFill('solid', start_color=COLOR_HL_EVEN if ri%2==0 else COLOR_HL_ODD)
                c = ws.cell(row=ri, column=ci, value=row.get(key, ''))
                c.font   = Font(name='Arial', bold=is_hl, size=9, color='1F4E79' if is_hl else '000000')
                c.fill   = hl_f if is_hl else fill
                c.border = brdh if is_hl else brd
                c.alignment = ctr if key in ctr_keys else lft
            ws.row_dimensions[ri].height = 16
        ws.freeze_panes = 'A4'

    # Φύλλο 1: Ανά Ειδικότητα
    grp_eid = (df_nz.groupby(['Κωδικός Κύριας Ειδικότητας', 'Κύρια Ειδικότητα'])
               .agg(Πλήθος=('Α.Μ.', 'count'), Υπόλοιπο=('_yp', 'sum'))
               .reset_index().sort_values('Υπόλοιπο', ascending=False))
    ws1 = wb.active
    ws1.title = 'ΥΠΟΛΟΙΠΑ ΑΝΑ ΕΙΔΙΚΟΤΗΤΑ'
    _hdr(ws1, 'Υπόλοιπα ανά Ειδικότητα', 4, COLOR_PIVOT)
    _sub(ws1, f'Σύνολο ειδικοτήτων: {len(grp_eid)}', 4, COLOR_HEAD2P)
    for ci, (lbl, w) in enumerate([('Κωδικός', 14), ('Ειδικότητα', 34),
                                    ('Πλήθος', 10), ('Σύνολο Υπολοίπου', 16)], 1):
        is_hl = (lbl == 'Σύνολο Υπολοίπου')
        c = ws1.cell(row=3, column=ci, value=lbl)
        c.font   = Font(name='Arial', bold=True, color='1F4E79' if is_hl else 'FFFFFF', size=10)
        c.fill   = PatternFill('solid', start_color=COLOR_HL_HEADER if is_hl else COLOR_PIVOT)
        c.alignment = ctr; c.border = brdh if is_hl else brd
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[3].height = 28
    for ri, (_, row) in enumerate(grp_eid.iterrows(), 4):
        fill = alt_p if ri % 2 == 0 else PatternFill()
        vals = [row['Κωδικός Κύριας Ειδικότητας'], row['Κύρια Ειδικότητα'],
                row['Πλήθος'], row['Υπόλοιπο']]
        for ci, val in enumerate(vals, 1):
            is_hl = (ci == 4)
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font   = Font(name='Arial', bold=is_hl, size=9, color='1F4E79' if is_hl else '000000')
            c.fill   = PatternFill('solid', start_color=COLOR_HL_EVEN if ri%2==0 else COLOR_HL_ODD) if is_hl else fill
            c.border = brdh if is_hl else brd
            c.alignment = lft if ci == 2 else ctr
        ws1.row_dimensions[ri].height = 16
    _total_row(ws1, len(grp_eid) + 4,
               [' ', 'ΣΥΝΟΛΟ', grp_eid['Πλήθος'].sum(), grp_eid['Υπόλοιπο'].sum()], hl_ci=4)
    ws1.freeze_panes = 'A4'

    # Φύλλο 2: Cross-table ανά Σχολείο
    specs   = sorted(df_nz['Κωδικός Κύριας Ειδικότητας'].unique())
    pivot   = (df_nz.pivot_table(index=SCHOOL_COLUMN, columns='Κωδικός Κύριας Ειδικότητας',
                                  values='_yp', aggfunc='sum', fill_value=0)
               .reindex(columns=specs, fill_value=0))
    pivot['ΣΥΝΟΛΟ'] = pivot.sum(axis=1)
    pivot   = pivot.reset_index().sort_values(SCHOOL_COLUMN)
    ncols_p = len(specs) + 2
    ci_sum  = len(specs) + 2
    ws2     = wb.create_sheet('ΥΠΟΛΟΙΠΑ ΑΝΑ ΣΧΟΛΕΙΟ')
    _hdr(ws2, 'Υπόλοιπα ανά Σχολείο', ncols_p, COLOR_PIVOT)
    _sub(ws2, f'Σύνολο σχολείων: {len(pivot)}', ncols_p, COLOR_HEAD2P)
    c = ws2.cell(row=3, column=1, value='Ονομασία Σχολείου')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', start_color=COLOR_PIVOT)
    c.alignment = ctr; c.border = brd
    ws2.column_dimensions['A'].width = 44
    for ci, sp in enumerate(specs, 2):
        c = ws2.cell(row=3, column=ci, value=sp)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill = PatternFill('solid', start_color='2E75B6')
        c.alignment = ctr; c.border = brd
        ws2.column_dimensions[get_column_letter(ci)].width = 7
    c = ws2.cell(row=3, column=ci_sum, value='ΣΥΝΟΛΟ')
    c.font = Font(name='Arial', bold=True, color='1F4E79', size=10)
    c.fill = PatternFill('solid', start_color=COLOR_HL_HEADER)
    c.alignment = ctr; c.border = brdh
    ws2.column_dimensions[get_column_letter(ci_sum)].width = 10
    ws2.row_dimensions[3].height = 28
    for ri, (_, row) in enumerate(pivot.iterrows(), 4):
        fill = alt_p if ri % 2 == 0 else PatternFill()
        c = ws2.cell(row=ri, column=1, value=row[SCHOOL_COLUMN])
        c.font = Font(name='Arial', size=9); c.fill = fill; c.border = brd; c.alignment = lft
        for ci, sp in enumerate(specs, 2):
            val = int(row[sp]) if row[sp] != 0 else None
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = Font(name='Arial', size=9); c.fill = fill; c.border = brd; c.alignment = ctr
        hl_f = PatternFill('solid', start_color=COLOR_HL_EVEN if ri%2==0 else COLOR_HL_ODD)
        c = ws2.cell(row=ri, column=ci_sum, value=int(row['ΣΥΝΟΛΟ']))
        c.font = Font(name='Arial', bold=True, size=9, color='1F4E79')
        c.fill = hl_f; c.border = brdh; c.alignment = ctr
        ws2.row_dimensions[ri].height = 15
    tr2 = len(pivot) + 4
    c = ws2.cell(row=tr2, column=1, value='ΓΕΝΙΚΟ ΣΥΝΟΛΟ')
    c.font = Font(name='Arial', bold=True, size=10, color='1F4E79')
    c.fill = PatternFill('solid', start_color='FFF3CD')
    c.border = brd; c.alignment = lft
    for ci, sp in enumerate(specs, 2):
        val = int(pivot[sp].sum()) or None
        c = ws2.cell(row=tr2, column=ci, value=val)
        c.font = Font(name='Arial', bold=True, size=9, color='1F4E79')
        c.fill = PatternFill('solid', start_color='FFF3CD'); c.border = brd; c.alignment = ctr
    c = ws2.cell(row=tr2, column=ci_sum, value=int(pivot['ΣΥΝΟΛΟ'].sum()))
    c.font = Font(name='Arial', bold=True, size=11, color='1F4E79')
    c.fill = PatternFill('solid', start_color=COLOR_HL_HEADER)
    c.border = brdh; c.alignment = ctr
    ws2.row_dimensions[tr2].height = 20
    ws2.freeze_panes = 'B4'

    # Φύλλα 3-5: Διευθυντές / Υποδιευθυντές / Αναπληρωτές
    cols_d = [(SCHOOL_COLUMN, 44), ('Δήμος', 18), ('Α.Μ.', 11),
              ('Επώνυμο', 18), ('Όνομα', 14),
              ('Κωδικός Κύριας Ειδικότητας', 12), (COL_YPOLOIPO, 14)]
    lbls_d = ['Ονομασία Σχολείου', 'Δήμος', 'Α.Μ.', 'Επώνυμο', 'Όνομα', 'Ειδικότητα', 'Υπόλοιπο']

    df_dir  = df_nz[df_nz['Διευθυντής Σχολείου'] == 'Ναι'].sort_values([SCHOOL_COLUMN, 'Επώνυμο']).copy()
    ws3 = wb.create_sheet('ΥΠΟΛΟΙΠΑ ΔΙΕΥΘΥΝΤΩΝ')
    _detail_sheet(ws3, df_dir, cols_d, lbls_d, COLOR_PIVOT, COLOR_HEAD2P, COLOR_ALT_P, 'Υπόλοιπα Διευθυντών')

    df_vdir = df_nz[df_nz['Υποδιευθυντής Σχολείου'] == 'Ναι'].sort_values([SCHOOL_COLUMN, 'Επώνυμο']).copy()
    ws4 = wb.create_sheet('ΥΠΟΛΟΙΠΑ ΥΠΟΔΙΕΥΘΥΝΤΩΝ')
    _detail_sheet(ws4, df_vdir, cols_d, lbls_d, COLOR_PIVOT, COLOR_HEAD2P, COLOR_ALT_P, 'Υπόλοιπα Υποδιευθυντών')

    cols_a = [
        (SCHOOL_COLUMN, 44), ('Δήμος', 18), ('Email', 28),
        ('Α.Μ.', 11), ('Επώνυμο', 18), ('Όνομα', 14),
        ('Κωδικός Κύριας Ειδικότητας', 12), ('Σχέση Εργασίας', 28),
        ('Σχέση Τοποθέτησης', 30),
        ('Ώρες Υποχ. Διδακτικού Ωραρίου Υπηρέτησης στο Φορέα', 12),
        ('Α Ανάθεση Συνολικά', 10), ("'λλες Αναθέσεις Συνολικά", 12),
        ('Συμπλήρωση Ωραρίου', 12), (COL_YPOLOIPO, 14),
    ]
    lbls_a = ['Ονομασία Σχολείου', 'Δήμος', 'Email', 'Α.Μ.', 'Επώνυμο', 'Όνομα',
              'Ειδικότητα', 'Σχέση Εργασίας', 'Σχέση Τοποθέτησης',
              'Ώρες\nΦορέα', 'Α Ανάθεση', 'Άλλες\nΑναθέσεις', 'Συμπλήρωση', 'Υπόλοιπο']
    df_an = df_nz[~df_nz['Σχέση Εργασίας'].str.contains('Μόνιμος', na=False)].sort_values([SCHOOL_COLUMN, 'Επώνυμο']).copy()
    ws5 = wb.create_sheet('ΥΠΟΛΟΙΠΑ ΑΝΑΠΛΗΡΩΤΩΝ')
    _detail_sheet(ws5, df_an, cols_a, lbls_a, COLOR_MAIN, COLOR_SUB, COLOR_ALT, 'Υπόλοιπα Αναπληρωτών')

    wb.save(output_path)


# ═══════════════════════════════════════════════════════════════════
# EMAIL (multi-attachment)
# ═══════════════════════════════════════════════════════════════════

def _send_email(cfg, to_addr, subject, body, attachment_paths):
    if isinstance(attachment_paths, str):
        attachment_paths = [attachment_paths]
    if isinstance(to_addr, list):
        recipients = to_addr
        to_header  = recipients[0]
    else:
        recipients = [to_addr]
        to_header  = to_addr
    msg = MIMEMultipart()
    from_name = getattr(cfg, 'FROM_NAME', cfg.FROM_EMAIL)
    msg['From']    = formataddr((from_name, cfg.FROM_EMAIL))
    msg['To']      = to_header
    msg['Subject'] = Header(subject, 'utf-8')
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    for path in attachment_paths:
        fname = os.path.basename(path)
        with open(path, 'rb') as f:
            part = MIMEBase('application', 'vnd.ms-excel')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=fname)
        part.add_header('Content-Type', 'application/vnd.ms-excel', name=fname)
        msg.attach(part)
    def _lenient_ctx():
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode    = ssl.CERT_NONE
        return ctx

    recipients = list(dict.fromkeys(recipients))  # dedup
    msg_str = msg.as_string()
    sent = False
    try:
        with smtplib.SMTP_SSL(cfg.SMTP_HOST, 465,
                               context=ssl.create_default_context()) as s:
            s.login(cfg.FROM_EMAIL, cfg.FROM_PASSWORD)
            s.sendmail(cfg.FROM_EMAIL, recipients, msg_str)
            sent = True
    except Exception:
        pass
    if not sent:
        try:
            with smtplib.SMTP_SSL(cfg.SMTP_HOST, 465,
                                   context=_lenient_ctx()) as s:
                s.login(cfg.FROM_EMAIL, cfg.FROM_PASSWORD)
                s.sendmail(cfg.FROM_EMAIL, recipients, msg_str)
                sent = True
        except Exception:
            pass
    if not sent:
        with smtplib.SMTP(cfg.SMTP_HOST, 587) as s:
            s.starttls(context=_lenient_ctx())
            s.login(cfg.FROM_EMAIL, cfg.FROM_PASSWORD)
            s.sendmail(cfg.FROM_EMAIL, recipients, msg_str)


# ═══════════════════════════════════════════════════════════════════
# CUSTOM RUN  (καλείται από main.py αντί για framework.run_check)
# ═══════════════════════════════════════════════════════════════════

def _ask_threshold():
    """Κατώφλι ωρών μέσω GUI popup."""
    import tkinter as tk
    result = [8]
    win = tk.Toplevel()
    win.title('Κατώφλι Ωρών')
    win.configure(bg='#EEF4F0')
    win.resizable(False, False)
    win.grab_set()
    win.attributes('-topmost', True)
    _ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '8ball.ico')
    if os.path.exists(_ico):
        try: win.iconbitmap(_ico)
        except Exception: pass
    win.update_idletasks()
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    win.geometry(f'300x150+{sw//2-150}+{sh//2-75}')

    tk.Label(win, text='Κατώφλι υπολοίπου (ώρες):', bg='#EEF4F0',
             fg='#1F4E79', font=('Arial',10,'bold'), pady=12).pack()

    val_var = tk.StringVar(value='8')
    sb = tk.Spinbox(win, from_=0, to=100, textvariable=val_var,
                    width=6, font=('Arial',12), justify='center')
    sb.pack()

    msg_var = tk.StringVar()
    tk.Label(win, textvariable=msg_var, bg='#EEF4F0', fg='#C62828',
             font=('Arial',8)).pack()

    def confirm():
        try:
            v = int(val_var.get())
            if v >= 0:
                result[0] = v
                win.destroy()
            else:
                msg_var.set('Δώσε μη αρνητικό αριθμό.')
        except ValueError:
            msg_var.set('Δώσε ακέραιο αριθμό.')

    tk.Button(win, text='OK', font=('Arial',10,'bold'),
              bg='#1F4E79', fg='white', relief='flat',
              padx=20, pady=6, cursor='hand2',
              command=confirm).pack(pady=10)

    win.wait_window()
    return result[0]

def run(config):
    """Custom run — παρακάμπτει το κοινό framework."""
    import core.framework as _fw
    _fw._current_check_title = CHECK_TITLE

    print('=' * 65)
    print(f'  {CHECK_TITLE}')
    print('=' * 65)

    path_48  = get_downloaded_file('4.8',  'Αρχείο 4.8 [csv]:',  csv_only=True)
    path_412 = get_downloaded_file('4.12', 'Αρχείο 4.12 [csv]:', csv_only=True)
    path_411 = get_downloaded_file('4.11', 'Αρχείο 4.11 [csv]:', csv_only=True)
    path_ady = get_ady_xoris_egkrisi('Αρχείο Αδυνατούντων (υπό έγκριση) [csv / xlsx]:')
    today    = ask_date_yyyymmdd()
    threshold= _ask_threshold()

    from core.framework import _ask_send_options_gui
    test_mode, do_send = _ask_send_options_gui()

    print(f'\n  Ημερομηνία : {today.strftime("%d/%m/%Y")}')
    print(f'  Κατώφλι    : >= {threshold} ώρες')
    print(f'  Λειτουργία : {"🧪 TEST MODE" if test_mode else "🚀 ΚΑΝΟΝΙΚΗ"}')
    print(f'  Αποστολή   : {"ΝΑΙ" if do_send else "ΟΧΙ"}')
    print('-' * 65)

    print('\nΦόρτωση αρχείων...')
    df48       = load_48(path_48)
    lookup_412 = load_412(path_412)
    lookup_411 = load_411(path_411)
    if path_ady:
        df_ady = load_adynatoyntes(path_ady)
        print(f'  ✓ Αδυνατούντες : {len(df_ady)} εγγραφές')
    else:
        df_ady = None
        print(f'  ℹ Αδυνατούντες : δεν έχει οριστεί — αγνοούνται')
    print(f'  ✓ 4.8          : {len(df48)} εγγραφές')
    print(f'  ✓ 4.12         : {len(lookup_412)} εκπαιδευτικοί')
    print(f'  ✓ 4.11         : {len(lookup_411)} εκπαιδευτικοί')

    print('\nΕπεξεργασία...')
    df_out  = process_data(df48, df_ady, threshold, lookup_412, lookup_411)
    schools = sorted(df_out[SCHOOL_COLUMN].unique())
    print(f'  → {len(df_out)} εκπαιδευτικοί με υπόλοιπο >= {threshold} ώρες')
    print(f'  → {len(schools)} σχολεία')

    if df_out.empty:
        print(f'\n✓ Κανένας εκπαιδευτικός δεν έχει υπόλοιπο >= {threshold} ώρες.')
        _show_results_popup(
            CHECK_TITLE,
            f'Ημερομηνία ελέγχου: {today.strftime("%d/%m/%Y")}\n'
            f'Κατώφλι: >= {threshold} ώρες\n\n'
            f'✓  Κανένας εκπαιδευτικός δεν έχει υπόλοιπο >= {threshold} ώρες.\n\n'
            f'Ο έλεγχος ολοκληρώθηκε χωρίς θέματα.',
            result_type='ok'
        )
        return

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_dir  = os.path.join(base_dir, f'results_{today.strftime("%Y%m%d")}', 'ypoloipa')
    os.makedirs(out_dir, exist_ok=True)
    print(f'\nΑποθήκευση → {out_dir}')
    print('-' * 65)

    path_all   = os.path.join(out_dir, f'{today.strftime("%Y%m%d")}_ΣΥΝΟΛΟ.xlsx')
    path_pivot = os.path.join(out_dir, f'{today.strftime("%Y%m%d")}_ΑΝΑΦΟΡΑ_PIVOT.xlsx')
    save_main_workbook(df_out, today, path_all)
    save_pivot_workbook(df_out, today, path_pivot)
    print(f'  ✓ Συνολικό      : {os.path.basename(path_all)}')
    print(f'  ✓ Pivot αναφορά : {os.path.basename(path_pivot)}')

    school_codes = sorted(df_out['Κωδικός Σχολείου'].unique())
    school_files = {}
    if not do_send or test_mode:
        print(f'\n  Χωρίς αποστολή: τα {len(school_codes)} αρχεία ανά σχολείο παραλείπονται.')
        for code in school_codes:
            df_s    = df_out[df_out['Κωδικός Σχολείου'] == code].copy()
            school  = df_s[SCHOOL_COLUMN].iloc[0]
            email_s = str(df_s['Email'].iloc[0]).strip() if 'Email' in df_s.columns else ''
            school_files[code] = ('', email_s, school)
    else:
        print(f'\n  Δημιουργία {len(school_codes)} αρχείων ανά σχολείο...')
        for code in school_codes:
            df_s      = df_out[df_out['Κωδικός Σχολείου'] == code].copy()
            school    = df_s[SCHOOL_COLUMN].iloc[0]
            email_s   = str(df_s['Email'].iloc[0]).strip() if 'Email' in df_s.columns else ''
            safe_name = ''.join(c for c in school if c not in r'\/:*?"<>|').strip()[:55]
            path_s    = os.path.join(out_dir, f'{today.strftime("%Y%m%d")}_{code}_{safe_name}.xlsx')
            save_main_workbook(df_s, today, path_s, school_name=school)
            school_files[code] = (path_s, email_s, school)
            print(f'  ✓ [{code}] {safe_name[:50]}  ({len(df_s)} εγγ.)')

    # Σύνοψη αποτελεσμάτων (χτίζεται πάντα)
    total_yp = sum(
        int(str(v).replace('="','').replace('"','').strip() or 0)
        for v in df_out[COL_YPOLOIPO]
        if str(v).replace('="','').replace('"','').strip().lstrip('-').isdigit()
    )
    avg_yp = round(total_yp / len(df_out), 1) if len(df_out) else 0
    summary_body = (
        f'Σύνοψη ελέγχου υπολοίπων ωραρίου — {today.strftime("%d/%m/%Y")}\n'
        f'{"─"*50}\n'
        f'Καταφλι: >= {threshold} ωρες\n'
        f'Εκπαιδευτικοι με υπολοιπο >= {threshold}: {len(df_out)} σε {len(school_codes)} σχολεια\n'
        f'Συνολο υπολοιπων: {total_yp} ωρες\n'
        f'Μεσος ορος: {avg_yp} ωρες/εκπαιδευτικο\n'
    )

    if do_send:
        print(f'\n{"─"*65}')
        if test_mode:
            test_body = summary_body
            print(f'🧪 TEST MODE → {config.TEST_EMAIL}')
            print(f'   Θέμα: [TEST] {EMAIL_SUBJECT}')
            print(f'\n   Body:\n{"─"*40}\n{test_body}\n{"─"*40}')
            cc_extra   = getattr(config, 'TEST_EMAIL_CC', None)
            recipients = [config.TEST_EMAIL] + ([cc_extra] if cc_extra else [])
            try:
                _send_email(config, recipients, f'[TEST] {EMAIL_SUBJECT}',
                            test_body, [path_all, path_pivot])
                print(f'  ✓ Εστάλη στο {config.TEST_EMAIL}' +
                      (f' + {cc_extra}' if cc_extra else ''))
            except Exception as e:
                print(f'  ✗ Σφάλμα: {e}')
        else:
            print(f'🚀 ΚΑΝΟΝΙΚΗ ΑΠΟΣΤΟΛΗ — {len(school_codes)} σχολεία')
            no_email = []
            for code, (path_s, email_s, school) in school_files.items():
                if not email_s or email_s in ('', 'nan', 'None'):
                    print(f'   ⚠  [{code}] {school[:48]} — ΔΕΝ ΥΠΑΡΧΕΙ EMAIL')
                    no_email.append(code)
                else:
                    print(f'   →  [{code}] {school[:43]} → {email_s}')
            if no_email:
                print(f'\n   ⚠  {len(no_email)} σχολεία χωρίς email — θα παραλειφθούν.')
            ok = fail = 0
            for code, (path_s, email_s, school) in school_files.items():
                if not email_s or email_s in ('', 'nan', 'None'):
                    fail += 1; continue
                try:
                    _send_email(config, [email_s], EMAIL_SUBJECT, EMAIL_BODY, path_s)
                    print(f'  ✓ [{code}] {school[:43]} → {email_s}')
                    ok += 1
                except Exception as e:
                    print(f'  ✗ [{code}] {school[:43]} → {e}')
                    fail += 1
            print(f'\n  Αποστολές: {ok} επιτυχείς, {fail} αποτυχίες')

    popup_text = summary_body + (
        f'\n\n{"─"*40}\nΑποτελέσματα αποθηκεύτηκαν στο φάκελο:\n{out_dir}\n\n'
        f'📋 Για περισσότερες πληροφορίες δες τα αρχεία\nστο φάκελο αποτελεσμάτων.'
    )
    _show_results_popup(CHECK_TITLE, popup_text, result_type='warn')

    print('─' * 65)
    print(f'\n✓ Ολοκληρώθηκε!  {2 + len(school_codes)} αρχεία στο φάκελο:')
    print(f'  {out_dir}')


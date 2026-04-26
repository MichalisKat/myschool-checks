import os, glob, zipfile, io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

SPECIALTIES = ['ΠΕ06', 'ΠΕ11']
dl = os.path.join(os.path.expanduser('~'), 'Downloads')

def auto_find(prefix):
    matches = sorted([f for f in glob.glob(os.path.join(dl, f'*{prefix}*'))
                      if not f.endswith('.tmp') and not f.endswith('.crdownload')])
    return matches[-1] if matches else ''

def fc(df, *kws):
    for kw in kws:
        kw = kw.lower()
        for col in df.columns:
            if kw in str(col).lower():
                return col
    return None

def norm_code(series):
    return series.fillna('').astype(str).str.strip().str.lstrip('0').str.replace(r'\.0$', '', regex=True)

def clean_afm(val):
    return str(val).strip().strip('"').lstrip('=').strip('"').strip()

def read_csv_enc(path):
    if not path: return pd.DataFrame()
    if path.endswith('.zip'):
        with zipfile.ZipFile(path) as z:
            csvname = [n for n in z.namelist() if n.endswith('.csv')][0]
            data = z.read(csvname)
        for enc in ['utf-8-sig', 'utf-8', 'iso-8859-7', 'cp1253']:
            try:
                return pd.read_csv(io.BytesIO(data), sep=None, engine='python',
                                   encoding=enc, header=0, dtype=str)
            except Exception: continue
        return pd.DataFrame()
    for enc in ['utf-8-sig', 'utf-8', 'iso-8859-7', 'cp1253']:
        try:
            return pd.read_csv(path, sep=None, engine='python',
                               encoding=enc, header=0, dtype=str)
        except Exception: continue
    return pd.DataFrame()

# ── Φόρτωση αρχείων (μία φορά) ───────────────────────────────────────────────
topoth_path = auto_find('Topothetiseis')
grid_path   = auto_find('gridResults')
stat16_path = auto_find('stat4_16')
stat1_path  = auto_find('stat4_1')
stat2_path  = auto_find('stat4_2')

df_t_raw = pd.read_excel(topoth_path, header=0)
spec_col        = fc(df_t_raw,'κλάδ','ειδικ') or df_t_raw.columns[4]
code_col        = fc(df_t_raw,'κωδικ') or df_t_raw.columns[7]
eponym_col      = fc(df_t_raw,'επώνυμ') or df_t_raw.columns[2]
org_col         = fc(df_t_raw,'σχέση εργ','οργαν') or df_t_raw.columns[5]
topoth_col      = fc(df_t_raw,'σχέση τοποθ') or df_t_raw.columns[6]
school_name_col = fc(df_t_raw,'φορέας τοποθ','φορέας') or df_t_raw.columns[8]
afm_col         = fc(df_t_raw,'α.φ.μ','αφμ') or df_t_raw.columns[1]
status_col      = fc(df_t_raw,'κατάσταση','κατασταση') or df_t_raw.columns[17]
area_mt_col     = fc(df_t_raw,'περιοχή μετάθεσης φορέα','μετάθεσης φορέα') or df_t_raw.columns[19]
am_col   = next((c for c in df_t_raw.columns if 'α.μ' in str(c).lower() and 'φ' not in str(c).lower()), df_t_raw.columns[0])
onoma_col= next((c for c in df_t_raw.columns if ('όνομ' in str(c).lower() or 'ονομ' in str(c).lower()) and 'ονομασ' not in str(c).lower() and 'σχολ' not in str(c).lower()), None)

# Φίλτρα Topothetiseis (κοινά για όλες τις ειδικότητες)
_EO = r'Με άδεια διδασκαλίας για Ξένο Σχολείο|Αναπληρωτής Ιδιωτικής Εκπαίδευσης|Ιδιωτικού Δικαίου Αορίστου Χρόνου'
_ET = r'Υπερωριακά|Μερική Διάθεση|Τοποθέτηση Διοικητικού'
df_t_base = df_t_raw[df_t_raw[status_col].fillna('').astype(str).str.strip() != 'ΠΑΡΗΛΘΕ'].copy()
df_t_base = df_t_base[~df_t_base[org_col].fillna('').astype(str).str.strip().str.contains(_EO, regex=True, na=False)].copy()
df_t_base = df_t_base[df_t_base[area_mt_col].fillna('').astype(str).str.contains(r'Α.{0,2}\s*ΘΕΣΣΑΛΟΝΙΚΗΣ.*Π\.Ε', regex=True, na=False)].copy()
df_t_base = df_t_base[~df_t_base[topoth_col].fillna('').astype(str).str.strip().str.contains(_ET, regex=True, na=False)].copy()
df_t_base['_code'] = norm_code(df_t_base[code_col])
df_t_base['_afm']  = df_t_base[afm_col].fillna('').astype(str).str.strip().str.replace(r'\.0$','',regex=True).str.zfill(9)
df_t_base['_am']   = df_t_base[am_col].fillna('').astype(str).str.strip().str.replace(r'\.0$','',regex=True)

# gridResults
df_g = pd.read_excel(grid_path, header=0)
gc_code = fc(df_g,'κωδικός','κωδ') or df_g.columns[11]
gc_name = fc(df_g,'ονομασ') or df_g.columns[1]
gc_phone= fc(df_g,'τηλ') or df_g.columns[15]
gc_email= fc(df_g,'e-mail','email') or df_g.columns[17]
gc_area = fc(df_g,'περιοχ','τοποθεσ') or df_g.columns[18]
gc_eidos= fc(df_g,'είδος','ειδος')
if gc_eidos:
    df_g = df_g[df_g[gc_eidos].fillna('').astype(str).str.strip() != 'Ιδιωτικά Σχολεία'].copy()
df_g['_code'] = norm_code(df_g[gc_code])
df_g_lu = df_g[['_code',gc_name,gc_phone,gc_email,gc_area]].drop_duplicates('_code').copy()
df_g_lu.columns = ['_code','_school_name','_phone','_school_email','_area']
df_g_lu['_phone'] = df_g_lu['_phone'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
valid_codes = set(df_g_lu['_code'])

# stat4_16
df_s16 = read_csv_enc(stat16_path)
if not df_s16.empty:
    s16_afm_col = next((c for c in df_s16.columns if 'α.μ' in str(c).lower() and 'φ' not in str(c).lower()), df_s16.columns[16])
    s16_abs_col = df_s16.columns[44]
    s16_eos_col = df_s16.columns[47] if len(df_s16.columns) > 47 else None
    df_s16['_afm'] = df_s16[s16_afm_col].apply(clean_afm).str.zfill(9)
    keep16 = ['_afm', s16_abs_col] + ([s16_eos_col] if s16_eos_col else [])
    df_s16_lu = df_s16[keep16].drop_duplicates('_afm').copy()
    ren16 = {s16_abs_col: '_apoysia'}
    if s16_eos_col: ren16[s16_eos_col] = '_eos'
    df_s16_lu = df_s16_lu.rename(columns=ren16)
    if '_eos' not in df_s16_lu: df_s16_lu['_eos'] = ''
else:
    df_s16_lu = pd.DataFrame(columns=['_afm','_apoysia','_eos'])

# stat4_1 + stat4_2
frames = []
for p in [stat1_path, stat2_path]:
    df_tmp = read_csv_enc(p)
    if not df_tmp.empty: frames.append(df_tmp)
if frames:
    df_41_42 = pd.concat(frames, ignore_index=True)
    df_41_42['_afm'] = df_41_42[df_41_42.columns[0]].apply(clean_afm).str.zfill(9)
    s41_psd = df_41_42.columns[12] if len(df_41_42.columns) > 12 else None
    s41_eml = df_41_42.columns[11] if len(df_41_42.columns) > 11 else None
    s41_mob = df_41_42.columns[9]  if len(df_41_42.columns) > 9  else None
    keep41 = ['_afm'] + [c for c in [s41_psd, s41_eml, s41_mob] if c]
    df_41_lu = df_41_42[keep41].drop_duplicates('_afm').copy()
    ren41 = {}
    if s41_psd: ren41[s41_psd] = '_email_psd'
    if s41_eml: ren41[s41_eml] = '_email_personal'
    if s41_mob: ren41[s41_mob] = '_kinito'
    df_41_lu = df_41_lu.rename(columns=ren41)
    for c in ['_email_psd','_email_personal','_kinito']:
        if c not in df_41_lu: df_41_lu[c] = ''
else:
    df_41_lu = pd.DataFrame(columns=['_afm','_email_psd','_email_personal','_kinito'])

OUT_COLS = ['ΑΜ','Επώνυμο','Όνομα','Κύρια Ειδικ.',
            'Email στο ΠΣΔ','Email','Κινητό',
            'Σχέση εργασίας','Σχέση τοποθέτησης','Κατάσταση',
            'Φορέας τοποθέτησης','Τηλέφωνο','e-mail','ΑΠΟΥΣΙΑ','Έως']

today_str = datetime.today().strftime('%Y%m%d')
out_dir   = os.path.join(os.path.expanduser('~'),'Documents','MySchoolChecks',f'results_{today_str}')
os.makedirs(out_dir, exist_ok=True)

# ── Βγάλε Excel για κάθε ειδικότητα ─────────────────────────────────────────
for SPECIALTY in SPECIALTIES:
    df_t = df_t_base[df_t_base['_code'].isin(valid_codes)].copy()
    df_t = df_t[df_t[spec_col].astype(str) == SPECIALTY].copy()

    if df_t.empty:
        print(f'{SPECIALTY}: δεν βρέθηκαν εγγραφές')
        continue

    df_t = df_t.merge(df_g_lu,   on='_code', how='left')
    df_t = df_t.merge(df_s16_lu, on='_afm',  how='left')
    df_t = df_t.merge(df_41_lu,  on='_afm',  how='left')
    df_t['_absent'] = df_t[status_col].fillna('').astype(str).str.strip() == 'ΑΠΟΥΣΙΑ'

    def gcol(col):
        if col is not None and col in df_t.columns:
            return df_t[col].fillna('').astype(str)
        return pd.Series(['']*len(df_t), index=df_t.index)

    out = pd.DataFrame(index=df_t.index)
    out['ΑΜ']                = df_t['_am'].fillna('')
    out['Επώνυμο']            = gcol(eponym_col)
    out['Όνομα']              = gcol(onoma_col)
    out['Κύρια Ειδικ.']      = gcol(spec_col)
    out['Email στο ΠΣΔ']     = df_t['_email_psd'].fillna('')
    out['Email']             = df_t['_email_personal'].fillna('') if '_email_personal' in df_t.columns else ''
    out['Κινητό']            = df_t['_kinito'].fillna('')
    out['Σχέση εργασίας']    = gcol(org_col)
    out['Σχέση τοποθέτησης'] = gcol(topoth_col)
    out['Κατάσταση']         = gcol(status_col)
    out['Φορέας τοποθέτησης']= gcol(school_name_col)
    out['Τηλέφωνο']          = df_t['_phone'].fillna('')
    out['e-mail']            = df_t['_school_email'].fillna('')
    out['ΑΠΟΥΣΙΑ']           = df_t['_apoysia'].fillna('')
    out['Έως']               = df_t['_eos'].fillna('') if '_eos' in df_t.columns else ''
    out['_absent']           = df_t['_absent']
    out.loc[~out['_absent'], 'ΑΠΟΥΣΙΑ'] = ''
    out.loc[~out['_absent'], 'Έως']     = ''
    out = out.sort_values('Επώνυμο', na_position='last').reset_index(drop=True)

    spec_safe = SPECIALTY.replace('/','_')
    out_path  = os.path.join(out_dir, f'Εκπαιδευτικοί_{spec_safe}_{today_str}.xlsx')

    wb = Workbook(); ws = wb.active; ws.title = SPECIALTY
    RED = 'FF0000'
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin,right=thin,top=thin,bottom=thin)
    for ci,col in enumerate(OUT_COLS,1):
        cell = ws.cell(row=1,column=ci,value=col)
        cell.font = Font(name='Arial',bold=True,color='FFFFFF',size=9)
        cell.fill = PatternFill('solid',start_color=RED)
        cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border = border
    alt_fill = PatternFill('solid',start_color='FFF0F0')
    for ri,row in out.iterrows():
        is_absent = bool(row.get('_absent',False))
        for ci,col in enumerate(OUT_COLS,1):
            val = row.get(col,'')
            if pd.isna(val): val = ''
            cell = ws.cell(row=ri+2,column=ci,value=str(val) if val!='' else '')
            if is_absent:
                cell.font = Font(name='Arial',size=9,color=RED,bold=True)
            else:
                cell.font = Font(name='Arial',size=9,color='000000')
                if ri % 2 == 1: cell.fill = alt_fill
            cell.alignment = Alignment(horizontal='left',vertical='center')
            cell.border = border
    for ci,col in enumerate(OUT_COLS,1):
        vals = [str(out.iloc[r][col]) for r in range(min(len(out),50))
                if col in out.columns and not pd.isna(out.iloc[r][col])]
        w = max([len(col)]+[len(v) for v in vals]) if vals else len(col)
        ws.column_dimensions[get_column_letter(ci)].width = min(w+3,42)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    wb.save(out_path)

    print(f'{SPECIALTY}: {len(out)} εγγραφές | Απόντες: {out["_absent"].sum()} | Email ΠΣΔ: {(out["Email στο ΠΣΔ"].str.strip()!="").sum()} | Κινητό: {(out["Κινητό"].str.strip()!="").sum()} | Με Έως: {(out["Έως"].str.strip()!="").sum()}')

print(f'\nΑποθηκεύτηκαν στο: {out_dir}')
import subprocess; subprocess.Popen(['explorer', out_dir])
